# # import eventlet  # REMOVED - Not used anywhere

from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, Response, current_app
from supabase.client import create_client, Client
import csv
import openpyxl
import os
from datetime import datetime, timedelta, date

import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from werkzeug.utils import secure_filename
import random
import string
import io
from dotenv import load_dotenv
from collections import defaultdict, Counter
import json
from auth import AuthManager, require_auth, require_admin, require_cre, require_ps, require_rec
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from security_verification import run_security_verification
import time
import gc
from flask_socketio import SocketIO, emit
import math
# from redis import Redis  # REMOVED - Not needed for local development

# Import optimized operations for faster lead updates
from optimized_lead_operations import create_optimized_operations

# Add this instead:
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from flask import send_file
import tempfile
import matplotlib
matplotlib.use('Agg')  # For headless environments
import matplotlib.pyplot as plt
import seaborn as sns
from matplotlib.backends.backend_pdf import PdfPages
import pytz
import pandas as pd
from werkzeug.security import generate_password_hash, check_password_hash

# Load environment variables from .env file
load_dotenv()

app = Flask(__name__)
socketio = SocketIO(app, cors_allowed_origins="*")

# =============================================================================
# AUTO-ASSIGN SYSTEM INTEGRATION
# =============================================================================

def auto_assign_new_leads_for_source(source):
    """
    Automatically assign new leads for a specific source based on saved configuration.
    This function is integrated directly into Flask app for automatic execution.
    
    Args:
        source (str): The source name to auto-assign leads for
        
    Returns:
        dict: Result with assigned_count and status
    """
    try:
        print(f"🤖 Auto-assigning new leads for source: {source}")
        
        # Get auto-assign configuration for this source
        config_result = supabase.table('auto_assign_config').select('cre_id').eq('source', source).execute()
        configs = config_result.data or []
        
        if not configs:
            print(f"ℹ️ No auto-assign configuration found for {source}")
            return {'success': False, 'message': f'No auto-assign configuration found for {source}', 'assigned_count': 0}
        
        cre_ids = [config['cre_id'] for config in configs]
        print(f"📋 Found {len(cre_ids)} CREs configured for {source}: {cre_ids}")
        
        # Get CRE details for fair distribution
        cre_result = supabase.table('cre_users').select('*').in_('id', cre_ids).execute()
        cres = cre_result.data or []
        
        if not cres:
            print(f"⚠️ No CREs found for source {source}")
            return {'success': False, 'message': f'No CREs found for source {source}', 'assigned_count': 0}
        
        # Get unassigned leads for this source (double-check to avoid race conditions)
        unassigned_result = supabase.table('lead_master').select('*').eq('assigned', 'No').eq('source', source).execute()
        unassigned_leads = unassigned_result.data or []
        
        if not unassigned_leads:
            print(f"ℹ️ No unassigned leads found for {source}")
            return {'success': True, 'message': f'No unassigned leads found for {source}', 'assigned_count': 0}
        
        print(f"📊 Found {len(unassigned_leads)} unassigned leads for {source}")
        
        # Get the last assigned CRE for this source to determine starting point
        last_assigned_result = supabase.table('lead_master').select('cre_name').eq('source', source).order('cre_assigned_at', desc=True).limit(1).execute()
        last_cre_name = None
        if last_assigned_result.data:
            last_cre_name = last_assigned_result.data[0].get('cre_name')
        
        # Find starting index for round-robin
        start_index = 0
        if last_cre_name:
            for i, cre in enumerate(cres):
                if cre['name'] == last_cre_name:
                    start_index = (i + 1) % len(cres)
                    break
        
        # Assign leads using round-robin fair distribution
        assigned_count = 0
        for i, lead in enumerate(unassigned_leads):
            # Double-check if lead is still unassigned before attempting assignment
            current_lead = supabase.table('lead_master').select('assigned').eq('uid', lead['uid']).execute()
            if not current_lead.data or current_lead.data[0].get('assigned') == 'Yes':
                print(f"ℹ️ Lead {lead['uid']} is already assigned, skipping...")
                continue
                
            selected_cre = cres[(start_index + i) % len(cres)]
            
            update_data = {
                'cre_name': selected_cre['name'],
                'assigned': 'Yes',
                'cre_assigned_at': get_ist_timestamp(),
                'lead_status': 'Pending'
            }
            
            try:
                # Update lead assignment with optimistic locking
                result = supabase.table('lead_master').update(update_data).eq('uid', lead['uid']).eq('assigned', 'No').execute()
                
                # Check if update was successful
                if not result.data:
                    print(f"ℹ️ Lead {lead['uid']} was already assigned by another process, skipping...")
                    continue
                
                # Create call attempt history
                call_history_data = {
                    'uid': lead['uid'],
                    'call_no': 'first',
                    'attempt': 1,
                    'status': 'Pending',
                    'cre_name': selected_cre['name'],
                    'call_was_recorded': False,
                    'follow_up_date': None,
                    'remarks': f'Auto-assigned to {selected_cre["name"]} from {source}',
                    'final_status': 'Pending'
                }
                supabase.table('cre_call_attempt_history').insert(call_history_data).execute()
                
                # Update CRE's auto-assign count
                supabase.table('cre_users').update({
                    'auto_assign_count': selected_cre.get('auto_assign_count', 0) + 1
                }).eq('id', selected_cre['id']).execute()
                
                # Log auto-assign history
                history_data = {
                    'lead_uid': lead['uid'],
                    'source': source,
                    'assigned_cre_id': selected_cre['id'],
                    'assigned_cre_name': selected_cre['name'],
                    'cre_total_leads_before': selected_cre.get('auto_assign_count', 0),
                    'cre_total_leads_after': selected_cre.get('auto_assign_count', 0) + 1,
                    'assignment_method': 'fair_distribution',
                    'created_at': get_ist_timestamp()
                }
                supabase.table('auto_assign_history').insert(history_data).execute()
                
                assigned_count += 1
                print(f"🎉 🟢 SUCCESS: Auto-assigned lead {lead['uid']} to {selected_cre['name']} (source: {source})")
                print(f"   📋 Lead Details: UID={lead['uid']}, Customer={lead.get('customer_name', 'N/A')}, Phone={lead.get('customer_phone_number', 'N/A')}")
                print(f"   👤 CRE Details: ID={selected_cre['id']}, Name={selected_cre['name']}, Total Leads Before={selected_cre.get('auto_assign_count', 0)}")
                print(f"   ⏰ Assignment Time: {get_ist_timestamp()}")
                print(f"   🔄 Status: {lead.get('lead_status', 'N/A')} → Pending")
                print(f"   📞 Call History: First call attempt recorded")
                print(f"   📊 Auto-assign Count: {selected_cre.get('auto_assign_count', 0)} → {selected_cre.get('auto_assign_count', 0) + 1}")
                print("   " + "="*80)
                
            except Exception as e:
                print(f"❌ 🔴 FAILED: Auto-assign failed for lead {lead['uid']} (source: {source})")
                print(f"   📋 Lead Details: UID={lead['uid']}, Customer={lead.get('customer_name', 'N/A')}, Phone={lead.get('customer_phone_number', 'N/A')}")
                print(f"   👤 CRE Details: ID={selected_cre['id']}, Name={selected_cre['name']}")
                print(f"   ⏰ Error Time: {get_ist_timestamp()}")
                print(f"   🚨 Error Details: {str(e)}")
                print(f"   🔍 Error Type: {type(e).__name__}")
                print(f"   📊 Current Status: assigned={lead.get('assigned', 'N/A')}, lead_status={lead.get('lead_status', 'N/A')}")
                print("   " + "="*80)
                # Don't continue on error, just log it
                continue
        
        if assigned_count > 0:
            print(f"🎉 🟢 SUCCESS SUMMARY: Auto-assigned {assigned_count} new leads for {source}")
            print(f"   📊 Total leads processed: {len(unassigned_leads)}")
            print(f"   ✅ Successfully assigned: {assigned_count}")
            print(f"   ❌ Failed assignments: {len(unassigned_leads) - assigned_count}")
            print(f"   👥 CREs involved: {[cre['name'] for cre in cres]}")
            print(f"   ⏰ Completion Time: {get_ist_timestamp()}")
            print("   " + "="*80)
        else:
            print(f"ℹ️ INFO: No leads were auto-assigned for {source}")
            print(f"   📊 Total leads processed: {len(unassigned_leads)}")
            print(f"   🔍 Reason: All leads were already assigned or no unassigned leads found")
            print("   " + "="*80)
        
        return {
            'success': True,
            'message': f'Successfully auto-assigned {assigned_count} new leads for {source}',
            'assigned_count': assigned_count
        }
        
    except Exception as e:
        print(f"❌ Error in auto_assign_new_leads_for_source: {e}")
        return {'success': False, 'message': str(e), 'assigned_count': 0}

@app.route('/check_and_assign_new_leads', methods=['POST'])
def check_and_assign_new_leads():
    """
    Check all configured sources and assign new leads automatically.
    This function runs in the background to continuously monitor for new leads.
    """
    try:
        print("🔍 Checking for new leads to auto-assign...")
        
        # Get all auto-assign configurations
        config_result = supabase.table('auto_assign_config').select('*').execute()
        configs = config_result.data or []
        
        if not configs:
            print("ℹ️ No auto-assign configurations found")
            return {'success': False, 'message': 'No auto-assign configurations found', 'total_assigned': 0}
        
        # Group configs by source
        source_configs = {}
        for config in configs:
            source = config['source']
            if source not in source_configs:
                source_configs[source] = []
            source_configs[source].append(config)
        
        # Process each source
        total_assigned = 0
        for source in source_configs.keys():
            result = auto_assign_new_leads_for_source(source)
            if result and result.get('success'):
                total_assigned += result.get('assigned_count', 0)
        
        if total_assigned > 0:
            print(f"🎉 🟢 GLOBAL SUCCESS: Auto-assign completed across all sources")
            print(f"   📊 Total leads assigned: {total_assigned}")
            print(f"   🔧 Sources processed: {list(source_configs.keys())}")
            print(f"   ⏰ Completion Time: {get_ist_timestamp()}")
            print("   " + "="*80)
        else:
            print(f"ℹ️ INFO: No new leads found to assign across all sources")
            print(f"   🔧 Sources checked: {list(source_configs.keys())}")
            print(f"   🔍 Reason: No unassigned leads found in any configured source")
            print("   " + "="*80)
        
        return jsonify({
            'success': True,
            'message': f'Auto-assign completed: {total_assigned} total leads assigned',
            'total_assigned': total_assigned
        })
        
    except Exception as e:
        print(f"❌ Error in check_and_assign_new_leads: {e}")
        return jsonify({'success': False, 'message': str(e), 'total_assigned': 0})

# =============================================================================

# Start background auto-assign thread
import threading
import time

def auto_assign_background_worker():
    """Background worker that continuously checks for new leads to auto-assign"""
    while True:
        try:
            # Check for new leads every 2 minutes
            time.sleep(120)  # 2 minutes
            print("🔄 Background auto-assign check running...")
            print(f"   ⏰ Check Time: {get_ist_timestamp()}")
            print("   " + "="*80)
            
            # Run within Flask application context
            with app.app_context():
                result = check_and_assign_new_leads()
                if result and result.get('success'):
                    print(f"   ✅ Background check completed successfully")
                else:
                    print(f"   ⚠️ Background check completed with issues")
                
        except Exception as e:
            print(f"❌ 🔴 CRITICAL ERROR in background auto-assign worker: {e}")
            print(f"   ⏰ Error Time: {get_ist_timestamp()}")
            print(f"   🚨 Error Type: {type(e).__name__}")
            print(f"   🔍 Error Details: {str(e)}")
            print("   " + "="*80)
            time.sleep(60)  # Wait 1 minute on error before retrying

# Start the background worker thread
auto_assign_thread = threading.Thread(target=auto_assign_background_worker, daemon=True)
auto_assign_thread.start()
print("🚀 Background auto-assign worker started (checking every 2 minutes)")

# Reduce Flask log noise
import logging
log = logging.getLogger('werkzeug')
log.setLevel(logging.WARNING)

# Add custom Jinja2 filters
# Note: Using Flask's built-in tojson filter instead of custom one

# Utility functions
def get_ist_timestamp():
    """Get current timestamp in Indian Standard Time with explicit timezone"""
    ist_time = datetime.now(pytz.timezone('Asia/Kolkata'))
    # Format with explicit timezone offset
    return ist_time.strftime('%Y-%m-%d %H:%M:%S.%f')[:-3] + '+05:30'

def normalize_call_dates(update_data: dict) -> dict:
    """Ensure all *_call_date fields are stored as full IST timestamps.
    If any call date fields are present with any value, overwrite with current IST timestamp.
    """
    try:
        call_keys = [
            'first_call_date', 'second_call_date', 'third_call_date',
            'fourth_call_date', 'fifth_call_date', 'sixth_call_date', 'seventh_call_date'
        ]
        for key in call_keys:
            if key in update_data and update_data[key] is not None:
                # Always overwrite with current timestamp, regardless of what was there
                print(f"DEBUG: Normalizing {key} from '{update_data[key]}' to '{get_ist_timestamp()}'")
                update_data[key] = get_ist_timestamp()
    except Exception as e:
        # Be resilient; on any issue, leave data as-is
        print(f"DEBUG: Error in normalize_call_dates: {e}")
        pass
    return update_data

def is_valid_date(date_string):
    """Validate date string format (YYYY-MM-DD)"""
    try:
        datetime.strptime(date_string, '%Y-%m-%d')
        return True
    except ValueError:
        return False

def is_valid_uid(uid):
    """Validate UID format"""
    if not uid:
        return False
    # Add your UID validation logic here
    return True


# Using Flask's built-in tojson filter

# Get environment variables with fallback values for testing
SUPABASE_URL = os.environ.get('SUPABASE_URL')
SUPABASE_KEY = os.environ.get('SUPABASE_ANON_KEY')
SECRET_KEY = os.environ.get('FLASK_SECRET_KEY', 'fallback-secret-key-change-this')

# Email configuration (add these to your .env file)
SMTP_SERVER = os.environ.get('SMTP_SERVER', 'smtp.gmail.com')
SMTP_PORT = int(os.environ.get('SMTP_PORT', '587'))
EMAIL_USER = os.environ.get('EMAIL_USER', '')
EMAIL_PASSWORD = os.environ.get('EMAIL_PASSWORD', '')

# Debug: Print to check if variables are loaded (remove in production)
print(f"SUPABASE_URL loaded: {SUPABASE_URL is not None}")
print(f"SUPABASE_KEY loaded: {SUPABASE_KEY is not None}")
print(f"SUPABASE_URL: {SUPABASE_URL}")
print(f"SUPABASE_ANON_KEY: {SUPABASE_KEY}")

# Validate required environment variables
if not SUPABASE_URL:
    raise ValueError("SUPABASE_URL environment variable is required. Please check your .env file.")
if not SUPABASE_KEY:
    raise ValueError("SUPABASE_ANON_KEY environment variable is required. Please check your .env file.")

app.secret_key = SECRET_KEY
app.permanent_session_lifetime = timedelta(hours=24)

# Initialize Supabase client
try:
    supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
    print("✅ Supabase client initialized successfully")
    # Removed Supabase warm-up query for local development
except Exception as e:
    print(f"❌ Error initializing Supabase client: {e}")
    raise

# Initialize optimized operations for faster lead updates
try:
    optimized_ops = create_optimized_operations(supabase)
    print("✅ Optimized operations initialized successfully")
except Exception as e:
    print(f"❌ Error initializing optimized operations: {e}")
    # Continue without optimized operations if there's an error
    optimized_ops = None

# Initialize AuthManager
auth_manager = AuthManager(supabase)
# Store auth_manager in app config instead of direct attribute
app.config['AUTH_MANAGER'] = auth_manager

# Initialize rate limiter
limiter = Limiter(
    key_func=get_remote_address,
    app=app,
    default_limits=["1000 per minute"]  # Use in-memory backend for local/dev
)

# Upload folder configuration
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'csv', 'xlsx', 'xls'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max file size

# Create upload folder if it doesn't exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def send_email_to_ps(ps_email, ps_name, lead_data, cre_name):
    """Send email notification to PS when a lead is assigned"""
    try:
        if not EMAIL_USER or not EMAIL_PASSWORD:
            print("Email credentials not configured. Skipping email notification.")
            return False

        # Create message
        msg = MIMEMultipart()
        msg['From'] = EMAIL_USER
        msg['To'] = ps_email
        msg['Subject'] = f"New Lead Assigned - {lead_data['customer_name']}"

        # Email body
        body = f"""
        Dear {ps_name},

        A new lead has been assigned to you by {cre_name}.

        Lead Details:
        - Customer Name: {lead_data['customer_name']}
        - Mobile Number: {lead_data['customer_mobile_number']}
        - Source: {lead_data['source']}
        - Lead Category: {lead_data.get('lead_category', 'Not specified')}
        - Model Interested: {lead_data.get('model_interested', 'Not specified')}
        - Branch: {lead_data.get('branch', 'Not specified')}

        Please log in to the CRM system to view and update this lead.

        Best regards,
        Ather CRM System
        """

        msg.attach(MIMEText(body, 'plain'))

        # Send email
        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        server.starttls()
        server.login(EMAIL_USER, EMAIL_PASSWORD)
        text = msg.as_string()
        server.sendmail(EMAIL_USER, ps_email, text)
        server.quit()

        print(f"Email sent successfully to {ps_email}")
        return True

    except Exception as e:
        print(f"Error sending email: {e}")
        return False


def read_csv_file(filepath):
    """Read CSV file and return list of dictionaries with memory optimization"""
    data = []
    try:
        with open(filepath, 'r', encoding='utf-8') as file:
            csv_reader = csv.DictReader(file)
            for row_num, row in enumerate(csv_reader):
                if row_num >= 10000:  # Limit to 10,000 rows
                    print(f"Warning: File contains more than 10,000 rows. Only processing first 10,000.")
                    break

                # Clean and validate row data
                cleaned_row = {}
                for key, value in row.items():
                    if key and value:  # Only include non-empty keys and values
                        cleaned_row[key.strip()] = str(value).strip()

                if cleaned_row:  # Only add non-empty rows
                    data.append(cleaned_row)

                # Memory management for large files
                if row_num % 1000 == 0 and row_num > 0:
                    print(f"Processed {row_num} rows...")

    except Exception as e:
        print(f"Error reading CSV file: {e}")
        raise

    return data


def read_excel_file(filepath):
    """Read Excel file and return list of dictionaries with memory optimization"""
    data = []
    try:
        workbook = openpyxl.load_workbook(filepath, read_only=True)  # Read-only mode for memory efficiency
        sheet = workbook.active

        # Get headers from first row
        headers = []
        if sheet and sheet[1]:
            for cell in sheet[1]:
                if cell and cell.value:
                    headers.append(str(cell.value).strip())
                else:
                    headers.append(None)

        # Read data rows with limit
        row_count = 0
        if sheet:
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if row_count >= 10000:  # Limit to 10,000 rows
                    print(f"Warning: File contains more than 10,000 rows. Only processing first 10,000.")
                    break

                row_data = {}
                has_data = False

                for i, value in enumerate(row):
                    if i < len(headers) and headers[i] and value is not None:
                        row_data[headers[i]] = str(value).strip()
                        has_data = True

                if has_data:  # Only add rows with actual data
                    data.append(row_data)
                    row_count += 1

                # Memory management for large files
                if row_count % 1000 == 0 and row_count > 0:
                    print(f"Processed {row_count} rows...")

        workbook.close()  # Explicitly close workbook

    except Exception as e:
        print(f"Error reading Excel file: {e}")
        raise

    return data


def batch_insert_leads(leads_data, batch_size=100):
    """Insert leads in batches to avoid overwhelming the database"""
    total_inserted = 0
    total_batches = (len(leads_data) + batch_size - 1) // batch_size

    print(f"Starting batch insert: {len(leads_data)} leads in {total_batches} batches")

    for i in range(0, len(leads_data), batch_size):
        batch = leads_data[i:i + batch_size]
        batch_num = (i // batch_size) + 1

        try:
            # Insert batch
            result = supabase.table('lead_master').insert(batch).execute()

            if result.data:
                batch_inserted = len(result.data)
                total_inserted += batch_inserted
                print(f"Batch {batch_num}/{total_batches}: Inserted {batch_inserted} leads")
            else:
                print(f"Batch {batch_num}/{total_batches}: No data returned from insert")

            # Small delay to prevent overwhelming the database
            time.sleep(0.1)  # CHANGED from eventlet.sleep(0.1)

            # Force garbage collection every 10 batches
            if batch_num % 10 == 0:
                gc.collect()

        except Exception as e:
            print(f"Error inserting batch {batch_num}: {e}")
            # Continue with next batch instead of failing completely
            continue

    print(f"Batch insert completed: {total_inserted} total leads inserted")
    return total_inserted


def generate_uid(source, mobile_number, sequence):
    """Generate UID based on source, mobile number, and sequence"""
    source_map = {
        'Google': 'G',
        'Meta': 'M',
        'Affiliate': 'A',
        'Know': 'K',
        'Whatsapp': 'W',
        'Tele': 'T',
        'Activity': 'AC',
        'Walk-in': 'W',  # Walk-in mapping
        'Walkin': 'W'    # Walkin mapping (without hyphen)
    }

    source_char = source_map.get(source, 'X')

    # Get sequence character (A-Z)
    sequence_char = chr(65 + (sequence % 26))  # A=65 in ASCII

    # Get last 4 digits of mobile number
    mobile_str = str(mobile_number).replace(' ', '').replace('-', '')
    mobile_last4 = mobile_str[-4:] if len(mobile_str) >= 4 else mobile_str.zfill(4)

    # Generate sequence number (0001, 0002, etc.)
    seq_num = f"{(sequence % 9999) + 1:04d}"

    return f"{source_char}{sequence_char}-{mobile_last4}-{seq_num}"


def get_next_call_info(lead_data):
    """Determine the next available call number and which calls are completed"""
    call_order = ['first', 'second', 'third', 'fourth', 'fifth', 'sixth', 'seventh']
    completed_calls = []
    next_call = 'first'

    for call_num in call_order:
        call_date_key = f'{call_num}_call_date'
        if lead_data.get(call_date_key):
            completed_calls.append(call_num)
        else:
            next_call = call_num
            break

    return next_call, completed_calls


def get_next_ps_call_info(ps_data):
    """Determine the next available PS call number and which calls are completed (now 7 calls)"""
    call_order = ['first', 'second', 'third', 'fourth', 'fifth', 'sixth', 'seventh']
    completed_calls = []
    next_call = 'first'

    for call_num in call_order:
        call_date_key = f'{call_num}_call_date'
        if ps_data.get(call_date_key):
            completed_calls.append(call_num)
        else:
            next_call = call_num
            break

    return next_call, completed_calls


def get_accurate_count(table_name, filters=None):
    """Get accurate count from Supabase table"""
    try:
        query = supabase.table(table_name).select('id')

        if filters:
            for key, value in filters.items():
                if value is not None:
                    query = query.eq(key, value)

        result = query.execute()

        # Count the returned data
        return len(result.data) if result.data else 0

    except Exception as e:
        print(f"Error getting count from {table_name}: {e}")
        return 0


def safe_get_data(table_name, filters=None, select_fields='*', limit=10000):
    """Safely get data from Supabase with error handling"""
    try:
        query = supabase.table(table_name).select(select_fields)

        if filters:
            for key, value in filters.items():
                if value is not None:
                    query = query.eq(key, value)

        # Add limit to prevent default 1000 row limitation
        if limit:
            query = query.limit(limit)

        result = query.execute()
        return result.data or []
    except Exception as e:
        print(f"Error fetching data from {table_name}: {e}")
        return []


def sync_test_drive_to_alltest_drive(source_table, original_id, lead_data):
    """
    Sync test drive data to alltest_drive table when test_drive_done is updated
    """
    try:
        # Check if test_drive_done is not null and is Yes/No or True/False
        test_drive_done = lead_data.get('test_drive_done')
        
        # Handle both boolean and string values
        if test_drive_done is None:
            return
        
        # Convert boolean to string if needed
        if test_drive_done is True:
            test_drive_done = 'Yes'
        elif test_drive_done is False:
            test_drive_done = 'No'
        elif test_drive_done not in ['Yes', 'No']:
            return
        
        # Check if record already exists in alltest_drive
        existing_record = supabase.table('alltest_drive').select('*').eq('source_table', source_table).eq('original_id', str(original_id)).execute()
        
        # Prepare data for alltest_drive table
        alltest_drive_data = {
            'source_table': source_table,
            'original_id': str(original_id),
            'test_drive_done': test_drive_done,
            'updated_at': datetime.now().isoformat()
        }
        
        # Map fields based on source table
        if source_table == 'walkin_table':
            alltest_drive_data.update({
                'customer_name': lead_data.get('customer_name'),
                'mobile_number': lead_data.get('mobile_number'),
                'lead_status': lead_data.get('lead_status'),
                'lead_category': lead_data.get('lead_category'),
                'model_interested': lead_data.get('model_interested'),
                'final_status': lead_data.get('status'),
                'ps_name': lead_data.get('ps_assigned'),
                'branch': lead_data.get('branch'),
                'created_at': lead_data.get('created_at')
            })
            # For walkin_table, use uid instead of id
            alltest_drive_data['original_id'] = lead_data.get('uid', str(original_id))
        elif source_table == 'ps_followup_master':
            alltest_drive_data.update({
                'customer_name': lead_data.get('customer_name'),
                'mobile_number': lead_data.get('customer_mobile_number'),
                'lead_status': lead_data.get('lead_status'),
                'lead_category': lead_data.get('lead_category'),
                'model_interested': lead_data.get('model_interested'),
                'final_status': lead_data.get('final_status'),
                'ps_name': lead_data.get('ps_name'),
                'branch': lead_data.get('branch') or lead_data.get('ps_branch'),
                'created_at': lead_data.get('created_at'),
                'lead_source': lead_data.get('lead_source'),
                'cre_name': lead_data.get('cre_name')
            })
        elif source_table == 'activity_leads':
            alltest_drive_data.update({
                'customer_name': lead_data.get('customer_name'),
                'mobile_number': lead_data.get('customer_phone_number'),
                'lead_status': lead_data.get('lead_status'),
                'lead_category': lead_data.get('lead_category'),
                'model_interested': lead_data.get('interested_model'),
                'final_status': lead_data.get('final_status'),
                'ps_name': lead_data.get('ps_name'),
                'branch': lead_data.get('location'),
                'created_at': lead_data.get('created_at'),
                'remarks': lead_data.get('remarks'),
                'activity_name': lead_data.get('activity_name'),
                'activity_location': lead_data.get('activity_location'),
                'customer_location': lead_data.get('customer_location'),
                'customer_profession': lead_data.get('customer_profession'),
                'gender': lead_data.get('gender')
            })
        
        # Insert or update record in alltest_drive table
        if existing_record.data:
            # Update existing record
            supabase.table('alltest_drive').update(alltest_drive_data).eq('source_table', source_table).eq('original_id', str(original_id)).execute()
        else:
            # Insert new record
            supabase.table('alltest_drive').insert(alltest_drive_data).execute()
            
        print(f"Successfully synced test drive data for {source_table} - {original_id}")
        
    except Exception as e:
        print(f"Error syncing test drive data to alltest_drive: {e}")


def create_or_update_ps_followup(lead_data, ps_name, ps_branch):
    from datetime import datetime
    try:
        existing = supabase.table('ps_followup_master').select('*').eq('lead_uid', lead_data['uid']).execute()
        ps_followup_data = {
            'lead_uid': lead_data['uid'],
            'ps_name': ps_name,
            'ps_branch': ps_branch,
            'customer_name': lead_data.get('customer_name'),
            'customer_mobile_number': lead_data.get('customer_mobile_number'),
            'source': lead_data.get('source'),
            'cre_name': lead_data.get('cre_name'),
            'lead_category': lead_data.get('lead_category'),
            'model_interested': lead_data.get('model_interested'),
            'final_status': 'Pending',
            'ps_assigned_at': datetime.now().isoformat(),  # Always set when PS is assigned
            'created_at': lead_data.get('created_at') or datetime.now().isoformat(),
            'first_call_date': None  # Ensure fresh leads start without first_call_date
        }
        if existing.data:
            supabase.table('ps_followup_master').update(ps_followup_data).eq('lead_uid', lead_data['uid']).execute()
        else:
            supabase.table('ps_followup_master').insert(ps_followup_data).execute()
    except Exception as e:
        print(f"Error creating/updating PS followup: {e}")

def track_cre_call_attempt(uid, cre_name, call_no, lead_status, call_was_recorded=False, follow_up_date=None, remarks=None):
    """Track CRE call attempt in the history table and update TAT for first attempt"""
    try:
        # Get the next attempt number for this call
        attempt_result = supabase.table('cre_call_attempt_history').select('attempt').eq('uid', uid).eq('call_no', call_no).order('attempt', desc=True).limit(1).execute()
        next_attempt = 1
        if attempt_result.data:
            next_attempt = attempt_result.data[0]['attempt'] + 1

        # Fetch the current final_status from lead_master
        final_status = None
        lead_result = supabase.table('lead_master').select('final_status').eq('uid', uid).limit(1).execute()
        if lead_result.data and 'final_status' in lead_result.data[0]:
            final_status = lead_result.data[0]['final_status']

        # Prepare attempt data
        attempt_data = {
            'uid': uid,
            'call_no': call_no,
            'attempt': next_attempt,
            'status': lead_status,
            'cre_name': cre_name,
            'call_was_recorded': call_was_recorded,
            'follow_up_date': follow_up_date,
            'remarks': remarks,
            'final_status': final_status
        }

        # Insert the attempt record
        insert_result = supabase.table('cre_call_attempt_history').insert(attempt_data).execute()
        print(f"Tracked call attempt: {uid} - {call_no} call, attempt {next_attempt}, status: {lead_status}")

        # --- TAT Calculation and Update ---
        if call_no == 'first' and next_attempt == 1:
            # Fetch lead's created_at
            lead_result = supabase.table('lead_master').select('created_at').eq('uid', uid).limit(1).execute()
            if lead_result.data and lead_result.data[0].get('created_at'):
                created_at_str = lead_result.data[0]['created_at']
                from datetime import datetime
                try:
                    if 'T' in created_at_str:
                        created_at = datetime.fromisoformat(created_at_str.replace('Z', '+00:00'))
                    else:
                        created_at = datetime.strptime(created_at_str, '%Y-%m-%d %H:%M:%S')
                except Exception:
                    created_at = None
                # Get updated_at from inserted attempt (if available), else use now
                updated_at_str = None
                if insert_result.data and insert_result.data[0].get('updated_at'):
                    updated_at_str = insert_result.data[0]['updated_at']
                else:
                    from datetime import datetime
                    updated_at_str = datetime.now().isoformat()
                try:
                    if 'T' in updated_at_str:
                        updated_at = datetime.fromisoformat(updated_at_str.replace('Z', '+00:00'))
                    else:
                        updated_at = datetime.strptime(updated_at_str, '%Y-%m-%d %H:%M:%S')
                except Exception:
                    updated_at = datetime.now()
                if created_at:
                    tat_seconds = (updated_at - created_at).total_seconds()
                    # Update lead_master with TAT
                    supabase.table('lead_master').update({'tat': tat_seconds}).eq('uid', uid).execute()
                    print(f"TAT updated for lead {uid}: {tat_seconds} seconds")
    except Exception as e:
        print(f"Error tracking CRE call attempt: {e}")

def track_ps_call_attempt(uid, ps_name, call_no, lead_status, call_was_recorded=False, follow_up_date=None, remarks=None):
    """Track PS call attempt in the history table"""
    try:
        # Get the next attempt number for this call
        attempt_result = supabase.table('ps_call_attempt_history').select('attempt').eq('uid', uid).eq('call_no', call_no).order('attempt', desc=True).limit(1).execute()
        next_attempt = 1
        if attempt_result.data:
            next_attempt = attempt_result.data[0]['attempt'] + 1

        # Fetch the current final_status from multiple tables
        final_status = None
        
        # Try all possible tables and field combinations
        tables_to_check = [
            # (table_name, uid_field, status_field, uid_value)
            ('ps_followup_master', 'lead_uid', 'final_status', uid),
            ('lead_master', 'uid', 'final_status', uid),
            ('walkin_table', 'uid', 'status', uid),
            ('activity_leads', 'activity_uid', 'final_status', uid),
            # Also try with different UID formats for walkin and activity
            ('walkin_table', 'uid', 'status', uid.replace('WB-', 'W')),
            ('activity_leads', 'activity_uid', 'final_status', uid.replace('WB-', 'A'))
        ]
        
        for table_name, uid_field, status_field, uid_value in tables_to_check:
            try:
                result = supabase.table(table_name).select(status_field).eq(uid_field, uid_value).limit(1).execute()
                if result.data and result.data[0].get(status_field):
                    final_status = result.data[0][status_field]
                    print(f"Found final_status '{final_status}' in {table_name} for {uid_value}")
                    break
            except Exception as e:
                print(f"Error checking {table_name}: {e}")
                continue
        
        # If still not found, set a default
        if not final_status:
            final_status = 'Pending'
            print(f"No final_status found for {uid}, defaulting to 'Pending'")

        # Prepare attempt data
        attempt_data = {
            'uid': uid,
            'call_no': call_no,
            'attempt': next_attempt,
            'status': lead_status,
            'ps_name': ps_name,
            'call_was_recorded': call_was_recorded,
            'follow_up_date': follow_up_date,
            'remarks': remarks,
            'final_status': final_status
        }

        # Insert the attempt record
        supabase.table('ps_call_attempt_history').insert(attempt_data).execute()
        print(f"Tracked PS call attempt: {uid} - {call_no} call, attempt {next_attempt}, status: {lead_status}")
    except Exception as e:
        print(f"Error tracking PS call attempt: {e}")


def filter_leads_by_date(leads, filter_type, date_field='created_at'):
    """Filter leads based on date range"""
    if filter_type == 'all':
        return leads

    today = datetime.now().date()

    if filter_type == 'today':
        start_date = today
        end_date = today
    elif filter_type == 'mtd':  # Month to Date
        start_date = today.replace(day=1)
        end_date = today
    elif filter_type == 'week':
        start_date = today - timedelta(days=today.weekday())  # Start of current week (Monday)
        end_date = today
    elif filter_type == 'month':
        start_date = today - timedelta(days=30)
        end_date = today
    elif filter_type == 'quarter':
        start_date = today - timedelta(days=90)
        end_date = today
    elif filter_type == 'year':
        start_date = today - timedelta(days=365)
        end_date = today
    else:
        return leads

    filtered_leads = []
    for lead in leads:
        lead_date_str = lead.get(date_field)
        if lead_date_str:
            try:
                # Handle different date formats
                if 'T' in lead_date_str:  # ISO format with time
                    lead_date = datetime.fromisoformat(lead_date_str.replace('Z', '+00:00')).date()
                else:  # Date only format
                    lead_date = datetime.strptime(lead_date_str, '%Y-%m-%d').date()

                if start_date <= lead_date <= end_date:
                    filtered_leads.append(lead)
            except (ValueError, TypeError):
                # If date parsing fails, include the lead
                filtered_leads.append(lead)
        else:
            # If no date field, include the lead
            filtered_leads.append(lead)

    return filtered_leads


def fix_missing_timestamps():
    """
    Fix missing timestamps for existing leads that have final_status but missing won_timestamp or lost_timestamp
    """
    try:
        # Fix lead_master table
        # Get leads with final_status = 'Won' but no won_timestamp
        won_leads = supabase.table('lead_master').select('uid, final_status, won_timestamp, lost_timestamp, updated_at').eq('final_status', 'Won').is_('won_timestamp', 'null').execute()
        
        for lead in won_leads.data:
            supabase.table('lead_master').update({
                'won_timestamp': lead.get('updated_at') or datetime.now().isoformat()
            }).eq('uid', lead['uid']).execute()
        
        # Get leads with final_status = 'Lost' but no lost_timestamp
        lost_leads = supabase.table('lead_master').select('uid, final_status, won_timestamp, lost_timestamp, updated_at').eq('final_status', 'Lost').is_('lost_timestamp', 'null').execute()
        
        for lead in lost_leads.data:
            supabase.table('lead_master').update({
                'lost_timestamp': lead.get('updated_at') or datetime.now().isoformat()
            }).eq('uid', lead['uid']).execute()
        
        # Fix ps_followup_master table
        # Get PS leads with final_status = 'Won' but no won_timestamp
        ps_won_leads = supabase.table('ps_followup_master').select('lead_uid, final_status, won_timestamp, lost_timestamp, updated_at').eq('final_status', 'Won').is_('won_timestamp', 'null').execute()
        
        for lead in ps_won_leads.data:
            supabase.table('ps_followup_master').update({
                'won_timestamp': lead.get('updated_at') or datetime.now().isoformat()
            }).eq('lead_uid', lead['lead_uid']).execute()
        
        # Get PS leads with final_status = 'Lost' but no lost_timestamp
        ps_lost_leads = supabase.table('ps_followup_master').select('lead_uid, final_status, won_timestamp, lost_timestamp, updated_at').eq('final_status', 'Lost').is_('lost_timestamp', 'null').execute()
        
        for lead in ps_lost_leads.data:
            supabase.table('ps_followup_master').update({
                'lost_timestamp': lead.get('updated_at') or datetime.now().isoformat()
            }).eq('lead_uid', lead['lead_uid']).execute()
        
        print(f"Fixed {len(won_leads.data)} won leads, {len(lost_leads.data)} lost leads in lead_master")
        print(f"Fixed {len(ps_won_leads.data)} won leads, {len(ps_lost_leads.data)} lost leads in ps_followup_master")
        
    except Exception as e:
        print(f"Error fixing timestamps: {str(e)}")


@app.route('/')
def index():
    session.clear()  # Ensure no session data is present
    return render_template('index.html')


@app.route('/unified_login', methods=['POST'])
@limiter.limit("100000 per minute")
def unified_login() -> Response:
    start_time = time.time()
    username = request.form.get('username', '').strip()
    password = request.form.get('password', '').strip()
    user_type = request.form.get('user_type', '').strip().lower()

    valid_user_types = ['admin', 'cre', 'ps', 'rec']
    if user_type not in valid_user_types:
        flash('Please select a valid role (Admin, CRE, PS, or Receptionist)', 'error')
        return redirect(url_for('index'))

    # Branch Head login removed

    elif user_type == 'rec':
        # Receptionist authentication
        rec_user = supabase.table('rec_users').select('*').eq('username', username).execute().data
        if not rec_user:
            flash('Invalid username or password', 'error')
            return redirect(url_for('index'))
        rec_user = rec_user[0]
        if not rec_user.get('is_active', True):
            flash('User is inactive', 'error')
            return redirect(url_for('index'))
        # Check password using werkzeug
        if not check_password_hash(rec_user['password_hash'], password):
            flash('Incorrect password', 'error')
            return redirect(url_for('index'))
        session.clear()
        session['rec_user_id'] = rec_user['id']
        session['rec_branch'] = rec_user['branch']
        session['rec_name'] = rec_user.get('name', username)
        session['user_type'] = 'rec'
        session['username'] = username
        flash('Welcome! Logged in as Receptionist', 'success')
        return redirect(url_for('add_walkin_lead'))

    # Existing logic for admin, cre, ps
    t_user = time.time()
    success, message, user_data = auth_manager.authenticate_user(username, password, user_type)
    print(f"[PERF] unified_login: authenticate_user({user_type}) took {time.time() - t_user:.3f} seconds")
    if success:
        t2 = time.time()
        session_id = auth_manager.create_session(user_data['id'], user_type, user_data)
        print(f"DEBUG: Logged in as user_type={user_type}, session.user_type={session.get('user_type')}")
        print(f"[PERF] unified_login: create_session took {time.time() - t2:.3f} seconds")
        if session_id:
            flash(f'Welcome! Logged in as {user_type.upper()}', 'success')
            t3 = time.time()
            # Redirect to appropriate dashboard
            if user_type == 'admin':
                print(f"[PERF] unified_login: redirect to admin_dashboard after {time.time() - t3:.3f} seconds")
                print(f"[PERF] unified_login TOTAL took {time.time() - start_time:.3f} seconds")
                return redirect(url_for('admin_dashboard'))
            elif user_type == 'cre':
                print(f"[PERF] unified_login: redirect to cre_dashboard after {time.time() - t3:.3f} seconds")
                print(f"[PERF] unified_login TOTAL took {time.time() - start_time:.3f} seconds")
                return redirect(url_for('cre_dashboard'))
            elif user_type == 'ps':
                print(f"[PERF] unified_login: redirect to ps_dashboard after {time.time() - t3:.3f} seconds")
                print(f"[PERF] unified_login TOTAL took {time.time() - start_time:.3f} seconds")
                return redirect(url_for('ps_dashboard'))
        else:
            flash('Error creating session', 'error')
            print(f"[PERF] unified_login: session creation failed after {time.time() - t2:.3f} seconds")
            print(f"[PERF] unified_login TOTAL took {time.time() - start_time:.3f} seconds")
            return redirect(url_for('index'))
    else:
        flash('Invalid username or password', 'error')
        print(f"[PERF] unified_login TOTAL (invalid login) took {time.time() - start_time:.3f} seconds")
        return redirect(url_for('index'))
# Keep the old login routes for backward compatibility (redirect to unified login)
@app.route('/admin_login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        return unified_login()
    return redirect(url_for('index'))


@app.route('/cre_login', methods=['GET', 'POST'])
def cre_login():
    if request.method == 'POST':
        return unified_login()
    return redirect(url_for('index'))


@app.route('/ps_login', methods=['GET', 'POST'])
def ps_login():
    if request.method == 'POST':
        return unified_login()
    return redirect(url_for('index'))


@app.route('/password_reset_request', methods=['GET', 'POST'])
@limiter.limit("5 per minute")
def password_reset_request():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        user_type = request.form.get('user_type', '').strip()

        if not username or not user_type:
            flash('Please enter username and select user type', 'error')
            return render_template('password_reset_request.html')

        success, message, token = auth_manager.generate_password_reset_token(username, user_type)

        if success:
            # Send the reset link via email
            # Fetch user email from the appropriate table
            user_email = None
            try:
                if user_type == 'admin':
                    user_result = supabase.table('admin_users').select('email').eq('username', username).execute()
                elif user_type == 'cre':
                    user_result = supabase.table('cre_users').select('email').eq('username', username).execute()
                elif user_type == 'ps':
                    user_result = supabase.table('ps_users').select('email').eq('username', username).execute()
                else:
                    user_result = None
                if user_result and user_result.data and user_result.data[0].get('email'):
                    user_email = user_result.data[0]['email']
            except Exception as e:
                print(f"Error fetching user email for password reset: {e}")
                user_email = None

            reset_url = url_for('password_reset', token=token, _external=True)
            email_sent = False
            if user_email:
                try:
                    msg = MIMEMultipart()
                    msg['From'] = EMAIL_USER
                    msg['To'] = user_email
                    msg['Subject'] = 'Ather CRM Password Reset Request'
                    body = f"""
                    Dear {username},

                    We received a request to reset your password for your Ather CRM account.

                    Please click the link below to reset your password:
                    {reset_url}

                    If you did not request this, please ignore this email.

                    Best regards,\nAther CRM System
                    """
                    msg.attach(MIMEText(body, 'plain'))
                    server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
                    server.starttls()
                    server.login(EMAIL_USER, EMAIL_PASSWORD)
                    text = msg.as_string()
                    server.sendmail(EMAIL_USER, user_email, text)
                    server.quit()
                    print(f"Password reset email sent to {user_email}")
                    email_sent = True
                except Exception as e:
                    print(f"Error sending password reset email: {e}")
                    email_sent = False
            if email_sent:
                flash('If the username exists and is valid, a password reset link has been sent to the registered email address.', 'success')
            else:
                flash('If the username exists and is valid, a password reset link has been sent to the registered email address.', 'success')
                # Optionally, log or alert admin if email sending failed
        else:
            flash(message, 'error')

    return render_template('password_reset_request.html')


@app.route('/password_reset/<token>', methods=['GET', 'POST'])
def password_reset(token):
    if request.method == 'POST':
        new_password = request.form.get('new_password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()

        if not new_password or not confirm_password:
            flash('Please enter and confirm your new password', 'error')
            return render_template('password_reset.html', token=token)

        if new_password != confirm_password:
            flash('Passwords do not match', 'error')
            return render_template('password_reset.html', token=token)

        success, message = auth_manager.reset_password_with_token(token, new_password)

        if success:
            flash('Password reset successfully. Please log in with your new password.', 'success')
            return redirect(url_for('index'))
        else:
            flash(message, 'error')

    return render_template('password_reset.html', token=token)


@app.route('/change_password', methods=['POST'])
@require_auth()
def change_password():
    current_password = request.form.get('current_password', '').strip()
    new_password = request.form.get('new_password', '').strip()
    confirm_password = request.form.get('confirm_password', '').strip()

    if not all([current_password, new_password, confirm_password]):
        flash('All fields are required', 'error')
        return redirect(url_for('security_settings'))

    if new_password != confirm_password:
        flash('New passwords do not match', 'error')
        return redirect(url_for('security_settings'))

    user_id = session.get('user_id')
    user_type = session.get('user_type')

    if not user_id or not user_type:
        flash('Session information not found', 'error')
        return redirect(url_for('security_settings'))

    success, message = auth_manager.change_password(user_id, user_type, current_password, new_password)

    if success:
        flash('Password changed successfully', 'success')
    else:
        flash(message, 'error')

    return redirect(url_for('security_settings'))


@app.route('/change_cre_password', methods=['GET', 'POST'])
@require_cre
def change_cre_password():
    if request.method == 'POST':
        current_password = request.form.get('current_password', '').strip()
        new_password = request.form.get('new_password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()

        if not all([current_password, new_password, confirm_password]):
            flash('All fields are required', 'error')
            return render_template('change_cre_password.html')

        if new_password != confirm_password:
            flash('New passwords do not match', 'error')
            return render_template('change_cre_password.html')

        user_id = session.get('user_id')
        user_type = session.get('user_type')

        if not user_id or not user_type:
            flash('Session information not found', 'error')
            return render_template('change_cre_password.html')

        success, message = auth_manager.change_password(user_id, user_type, current_password, new_password)

        if success:
            flash('Password changed successfully', 'success')
            return redirect(url_for('cre_dashboard'))
        else:
            flash(message, 'error')

    return render_template('change_cre_password.html')


@app.route('/change_ps_password', methods=['GET', 'POST'])
@require_ps
def change_ps_password():
    if request.method == 'POST':
        current_password = request.form.get('current_password', '').strip()
        new_password = request.form.get('new_password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()

        if not all([current_password, new_password, confirm_password]):
            flash('All fields are required', 'error')
            return render_template('change_ps_password.html')

        if new_password != confirm_password:
            flash('New passwords do not match', 'error')
            return render_template('change_ps_password.html')

        user_id = session.get('user_id')
        user_type = session.get('user_type')

        if not user_id or not user_type:
            flash('Session information not found', 'error')
            return render_template('change_ps_password.html')

        success, message = auth_manager.change_password(user_id, user_type, current_password, new_password)

        if success:
            flash('Password changed successfully', 'success')
            return redirect(url_for('ps_dashboard'))
        else:
            flash(message, 'error')

    return render_template('change_ps_password.html')


@app.route('/security_settings')
@require_auth()
def security_settings():
    user_id = session.get('user_id')
    user_type = session.get('user_type')

    if not user_id or not user_type:
        flash('Session information not found', 'error')
        return redirect(url_for('index'))

    # Get active sessions
    sessions = auth_manager.get_user_sessions(user_id, user_type)

    # Get audit logs
    audit_logs = auth_manager.get_audit_logs(user_id, user_type, limit=20)

    return render_template('security_settings.html', sessions=sessions, audit_logs=audit_logs)


@app.route('/security_audit')
@require_admin
def security_audit():
    """Security audit dashboard"""
    return render_template('security_audit.html')


@app.route('/run_security_audit', methods=['POST'])
@require_admin
def run_security_audit():
    """Run comprehensive security audit"""
    try:
        # Run security verification
        audit_results = run_security_verification(supabase)

        # Log the security audit
        user_id = session.get('user_id')
        user_type = session.get('user_type')
        
        if user_id and user_type:
            auth_manager.log_audit_event(
                user_id=user_id,
                user_type=user_type,
                action='SECURITY_AUDIT_RUN',
                resource='security_audit',
                details={'overall_score': audit_results.get('overall_score', 0)}
            )

        return jsonify({
            'success': True,
            'results': audit_results
        })
    except Exception as e:
        print(f"Error running security audit: {e}")
        return jsonify({
            'success': False,
            'message': str(e)
        })


@app.route('/terminate_session', methods=['POST'])
@require_auth()
def terminate_session():
    try:
        data = request.get_json()
        session_id = data.get('session_id')

        if not session_id:
            return jsonify({'success': False, 'message': 'Session ID required'})

        auth_manager.deactivate_session(session_id)

        user_id = session.get('user_id')
        user_type = session.get('user_type')
        
        if user_id and user_type:
            auth_manager.log_audit_event(
                user_id=user_id,
                user_type=user_type,
                action='SESSION_TERMINATED',
                details={'terminated_session': session_id}
            )

        return jsonify({'success': True, 'message': 'Session terminated successfully'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/terminate_all_sessions', methods=['POST'])
@require_auth()
def terminate_all_sessions():
    try:
        user_id = session.get('user_id')
        user_type = session.get('user_type')
        current_session = session.get('session_id')

        if user_id and user_type and current_session:
            auth_manager.deactivate_all_user_sessions(user_id, user_type, current_session)
            auth_manager.log_audit_event(
                user_id=user_id,
                user_type=user_type,
                action='ALL_SESSIONS_TERMINATED',
                details={'except_session': current_session}
            )
        else:
            return jsonify({'success': False, 'message': 'Session information not found'})

        return jsonify({'success': True, 'message': 'All other sessions terminated successfully'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/admin_dashboard')
@require_admin
def admin_dashboard():
    # Get counts for dashboard with better error handling and actual queries
    try:
        # Get actual counts from database with proper queries
        cre_count = get_accurate_count('cre_users')
        ps_count = get_accurate_count('ps_users')
        leads_count = get_accurate_count('lead_master')
        unassigned_leads = get_accurate_count('lead_master', {'assigned': 'No'})

        print(
            f"Dashboard counts - CRE: {cre_count}, PS: {ps_count}, Total Leads: {leads_count}, Unassigned: {unassigned_leads}")

    except Exception as e:
        print(f"Error getting dashboard counts: {e}")
        cre_count = ps_count = leads_count = unassigned_leads = 0

    # Log dashboard access
    user_id = session.get('user_id')
    user_type = session.get('user_type')
    if user_id and user_type:
        auth_manager.log_audit_event(
            user_id=user_id,
            user_type=user_type,
            action='DASHBOARD_ACCESS',
            resource='admin_dashboard'
        )

    return render_template('admin_dashboard.html',
                           cre_count=cre_count,
                           ps_count=ps_count,
                           leads_count=leads_count,
                           unassigned_leads=unassigned_leads)


@app.route('/upload_data', methods=['GET', 'POST'])
@require_admin
def upload_data():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file selected', 'error')
            return redirect(request.url)

        file = request.files['file']
        source = request.form.get('source', '').strip()

        if not source:
            flash('Please select a data source', 'error')
            return redirect(request.url)

        if file.filename == '':
            flash('No file selected', 'error')
            return redirect(request.url)

        if file and file.filename and allowed_file(file.filename):
            filename = secure_filename(str(file.filename))
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)

            try:
                file.save(filepath)

                # Check file size
                file_size = os.path.getsize(filepath)
                if file_size > 50 * 1024 * 1024:  # 50MB limit
                    flash('File too large. Maximum size is 50MB.', 'error')
                    os.remove(filepath)
                    return redirect(request.url)

                print(f"Processing file: {filename} ({file_size / 1024 / 1024:.2f} MB)")

                # Read the file based on extension
                if filename.lower().endswith('.csv'):
                    data = read_csv_file(filepath)
                else:
                    data = read_excel_file(filepath)

                if not data:
                    flash('No valid data found in file', 'error')
                    os.remove(filepath)
                    return redirect(request.url)

                print(f"Read {len(data)} rows from file")

                # Get current sequence number for UID generation
                result = supabase.table('lead_master').select('uid').execute()
                current_count = len(result.data) if result.data else 0

                # Prepare leads data for batch insert
                leads_to_insert = []
                skipped_rows = 0

                for index, row in enumerate(data):
                    try:
                        # Validate required fields
                        required_fields = ['customer_name', 'customer_mobile_number', 'date']
                        if not all(key in row and str(row[key]).strip() for key in required_fields):
                            skipped_rows += 1
                            continue

                        uid = generate_uid(source, row['customer_mobile_number'],
                                           current_count + len(leads_to_insert) + 1)

                        lead_data = {
                            'uid': uid,
                            'date': str(row['date']).strip(),
                            'customer_name': str(row['customer_name']).strip(),
                            'customer_mobile_number': str(row['customer_mobile_number']).strip(),
                            'source': source,
                            'assigned': 'No',
                            'final_status': 'Pending'
                        }

                        leads_to_insert.append(lead_data)


                    except Exception as e:
                        print(f"Error processing row {index}: {e}")
                        skipped_rows += 1
                        continue

                if not leads_to_insert:
                    flash('No valid leads found to insert', 'error')
                    os.remove(filepath)
                    return redirect(request.url)

                print(f"Prepared {len(leads_to_insert)} leads for insertion")

                # Batch insert leads
                success_count = batch_insert_leads(leads_to_insert)

                # Log data upload
                auth_manager.log_audit_event(
                    user_id=session.get('user_id'),
                    user_type=session.get('user_type'),
                    action='DATA_UPLOAD',
                    resource='lead_master',
                    details={
                        'source': source,
                        'records_uploaded': success_count,
                        'filename': filename,
                        'file_size_mb': round(file_size / 1024 / 1024, 2),
                        'skipped_rows': skipped_rows
                    }
                )

                # Create success message
                message = f'Successfully uploaded {success_count} leads'
                if skipped_rows > 0:
                    message += f' ({skipped_rows} rows skipped due to missing data)'
                message += '. Please go to "Assign Leads" to assign them to CREs.'

                flash(message, 'success')

                # Clean up uploaded file
                os.remove(filepath)

            except Exception as e:
                print(f"Error processing file: {e}")
                flash(f'Error processing file: {str(e)}', 'error')
                if os.path.exists(filepath):
                    os.remove(filepath)
        else:
            flash('Invalid file format. Please upload CSV or Excel files only.', 'error')

    return render_template('upload_data.html')

@app.route('/assign_leads')
@require_admin
def assign_leads():
    try:
        # Fetch all unassigned leads in batches of 1000
        all_unassigned_leads = []
        batch_size = 1000
        offset = 0
        while True:
            result = supabase.table('lead_master').select('*').eq('assigned', 'No').range(offset, offset + batch_size - 1).execute()
            batch = result.data or []
            all_unassigned_leads.extend(batch)
            if len(batch) < batch_size:
                break
            offset += batch_size

        # Organize by source
        leads_by_source = {}
        for lead in all_unassigned_leads:
            source = lead.get('source', 'Unknown')
            leads_by_source.setdefault(source, []).append(lead)

        # Get CREs
        cres = safe_get_data('cre_users')

        # Get accurate total unassigned count
        actual_unassigned_count = get_accurate_count('lead_master', {'assigned': 'No'})

        # Get accurate per-source unassigned counts
        source_unassigned_counts = {}
        for source in leads_by_source.keys():
            source_unassigned_counts[source] = get_accurate_count('lead_master', {'assigned': 'No', 'source': source})

        # Get all sources that have auto-assign configurations, even if they have zero unassigned leads
        auto_assign_sources = set()
        try:
            auto_assign_result = supabase.table('auto_assign_config').select('source').execute()
            if auto_assign_result.data:
                for config in auto_assign_result.data:
                    auto_assign_sources.add(config['source'])
        except Exception as e:
            print(f"Error fetching auto-assign sources: {e}")

        # Add sources with auto-assign configs to source_unassigned_counts if they're not already there
        for source in auto_assign_sources:
            if source not in source_unassigned_counts:
                source_unassigned_counts[source] = 0

        return render_template('assign_leads.html',
                               unassigned_leads=all_unassigned_leads,
                               actual_unassigned_count=actual_unassigned_count,
                               cres=cres,
                               leads_by_source=leads_by_source,
                               source_unassigned_counts=source_unassigned_counts)
    except Exception as e:
        print(f"Error loading assign leads data: {e}")
        flash(f'Error loading data: {str(e)}', 'error')
        return redirect(url_for('admin_dashboard'))


@app.route('/assign_leads_dynamic_action', methods=['POST'])
@require_admin
def assign_leads_dynamic_action():
    try:
        data = request.get_json()
        assignments = data.get('assignments', [])

        if not assignments:
            return jsonify({'success': False, 'message': 'No assignments provided'}), 400

        # Fetch all unassigned leads in batches
        all_unassigned = []
        batch_size = 1000
        offset = 0
        while True:
            result = supabase.table('lead_master').select('*').eq('assigned', 'No').range(offset, offset + batch_size - 1).execute()
            batch = result.data or []
            all_unassigned.extend(batch)
            if len(batch) < batch_size:
                break
            offset += batch_size

        leads_by_source = {}
        for lead in all_unassigned:
            source = lead.get('source', 'Unknown')
            leads_by_source.setdefault(source, []).append(lead)

        total_assigned = 0

        for assignment in assignments:
            cre_id = assignment.get('cre_id')
            source = assignment.get('source')
            quantity = assignment.get('quantity')

            if not cre_id or not source or not quantity:
                continue

            # Get CRE details
            cre_data = supabase.table('cre_users').select('*').eq('id', cre_id).execute()
            if not cre_data.data:
                continue

            cre = cre_data.data[0]
            leads = leads_by_source.get(source, [])

            if not leads:
                print(f"No unassigned leads found for source {source}")
                continue

            random.shuffle(leads)
            leads_to_assign = leads[:quantity]
            leads_by_source[source] = leads[quantity:]  # Remove assigned leads

            for lead in leads_to_assign:
                update_data = {
                    'cre_name': cre['name'],
                    'assigned': 'Yes',
                    'cre_assigned_at': datetime.now().isoformat()

                }

                try:
                    supabase.table('lead_master').update(update_data).eq('uid', lead['uid']).execute()
                    total_assigned += 1
                    print(f"Assigned lead {lead['uid']} to CRE {cre['name']} for source {source}")
                    if total_assigned % 100 == 0:
                        time.sleep(0.1)
                except Exception as e:
                    print(f"Error assigning lead {lead['uid']}: {e}")

        print(f"Total leads assigned: {total_assigned}")
        return jsonify({'success': True, 'message': f'Total {total_assigned} leads assigned successfully'})

    except Exception as e:
        print(f"Error in dynamic lead assignment: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500
@app.route('/add_cre', methods=['GET', 'POST'])
@require_admin
def add_cre():
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        phone = request.form.get('phone', '').strip()
        email = request.form.get('email', '').strip()

        if not all([name, username, password, phone, email]):
            flash('All fields are required', 'error')
            return render_template('add_cre.html')

        if not auth_manager.validate_password_strength(password):
            flash(
                'Password must be at least 8 characters long and contain uppercase, lowercase, number, and special character',
                'error')
            return render_template('add_cre.html')

        try:
            # Check if username already exists
            existing = supabase.table('cre_users').select('username').eq('username', username).execute()
            if existing.data:
                flash('Username already exists', 'error')
                return render_template('add_cre.html')

            # Hash password
            password_hash, salt = auth_manager.hash_password(password)

            # Replace the existing cre_data creation with this:
            cre_data = {
                'name': name,
                'username': username,
                'password': password,  # Keep for backward compatibility
                'password_hash': password_hash,
                'salt': salt,
                'phone': phone,
                'email': email,
                'is_active': True,
                'role': 'cre',
                'failed_login_attempts': 0
            }

            result = supabase.table('cre_users').insert(cre_data).execute()

            # Log CRE creation
            auth_manager.log_audit_event(
                user_id=session.get('user_id'),
                user_type=session.get('user_type'),
                action='CRE_CREATED',
                resource='cre_users',
                resource_id=str(result.data[0]['id']) if result.data else None,
                details={'cre_name': name, 'username': username}
            )

            flash('CRE added successfully', 'success')
            return redirect(url_for('manage_cre'))
        except Exception as e:
            flash(f'Error adding CRE: {str(e)}', 'error')

    return render_template('add_cre.html')


@app.route('/add_ps', methods=['GET', 'POST'])
@require_admin
def add_ps():
    branches = ['PORUR', 'NUNGAMBAKKAM', 'TIRUVOTTIYUR']

    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        phone = request.form.get('phone', '').strip()
        email = request.form.get('email', '').strip()
        branch = request.form.get('branch', '').strip()

        if not all([name, username, password, phone, email, branch]):
            flash('All fields are required', 'error')
            return render_template('add_ps.html', branches=branches)

        if not auth_manager.validate_password_strength(password):
            flash(
                'Password must be at least 8 characters long and contain uppercase, lowercase, number, and special character',
                'error')
            return render_template('add_ps.html', branches=branches)

        try:
            # Check if username already exists
            existing = supabase.table('ps_users').select('username').eq('username', username).execute()
            if existing.data:
                flash('Username already exists', 'error')
                return render_template('add_ps.html', branches=branches)

            # Hash password
            password_hash, salt = auth_manager.hash_password(password)

            # Replace the existing ps_data creation with this:
            ps_data = {
                'name': name,
                'username': username,
                'password': password,  # Keep for backward compatibility
                'password_hash': password_hash,
                'salt': salt,
                'phone': phone,
                'email': email,
                'branch': branch,
                'is_active': True,
                'role': 'ps',
                'failed_login_attempts': 0
            }

            result = supabase.table('ps_users').insert(ps_data).execute()

            # Log PS creation
            auth_manager.log_audit_event(
                user_id=session.get('user_id'),
                user_type=session.get('user_type'),
                action='PS_CREATED',
                resource='ps_users',
                resource_id=str(result.data[0]['id']) if result.data else None,
                details={'ps_name': name, 'username': username, 'branch': branch}
            )

            flash('Product Specialist added successfully', 'success')
            return redirect(url_for('manage_ps'))
        except Exception as e:
            flash(f'Error adding Product Specialist: {str(e)}', 'error')

    return render_template('add_ps.html', branches=branches)


@app.route('/manage_cre')
@require_admin
def manage_cre():
    try:
        cre_users = safe_get_data('cre_users')
        return render_template('manage_cre.html', cre_users=cre_users)
    except Exception as e:
        flash(f'Error loading CRE users: {str(e)}', 'error')
        return render_template('manage_cre.html', cre_users=[])


@app.route('/manage_ps')
@require_admin
def manage_ps():
    try:
        ps_users = safe_get_data('ps_users')
        return render_template('manage_ps.html', ps_users=ps_users)
    except Exception as e:
        flash(f'Error loading PS users: {str(e)}', 'error')
        return render_template('manage_ps.html', ps_users=[])


@app.route('/toggle_ps_status/<int:ps_id>', methods=['POST'])
@require_admin
def toggle_ps_status(ps_id):
    try:
        data = request.get_json()
        active_status = data.get('active', True)

        # Update PS status
        result = supabase.table('ps_users').update({
            'is_active': active_status
        }).eq('id', ps_id).execute()

        if result.data:
            # Log status change
            auth_manager.log_audit_event(
                user_id=session.get('user_id'),
                user_type=session.get('user_type'),
                action='PS_STATUS_CHANGED',
                resource='ps_users',
                resource_id=str(ps_id),
                details={'new_status': 'active' if active_status else 'inactive'}
            )

            return jsonify({'success': True, 'message': 'PS status updated successfully'})
        else:
            return jsonify({'success': False, 'message': 'PS not found'})

    except Exception as e:
        print(f"Error toggling PS status: {e}")
        return jsonify({'success': False, 'message': str(e)})


@app.route('/manage_leads')
@require_admin
def manage_leads():
    try:
        cres = safe_get_data('cre_users')
        cre_id = request.args.get('cre_id')
        source = request.args.get('source')
        qualification = request.args.get('qualification', 'all')
        date_filter = request.args.get('date_filter', 'all')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        final_status = request.args.get('final_status', '')
        page = int(request.args.get('page', 1))
        per_page = 50
        search_uid = request.args.get('search_uid', '').strip()
        selected_cre = None
        leads = []
        sources = []
        total_leads = 0
        if cre_id:
            # Find selected CRE
            selected_cre = next((cre for cre in cres if str(cre.get('id')) == str(cre_id)), None)
            # Fetch leads for this CRE
            filters = {'cre_name': selected_cre['name']} if selected_cre else {}
            if source:
                filters['source'] = source
            leads = safe_get_data('lead_master', filters)
            # UID substring search for this CRE
            if search_uid:
                leads = [lead for lead in leads if search_uid.lower() in str(lead.get('uid', '')).lower()]
            # Qualification filter
            if qualification == 'qualified':
                leads = [lead for lead in leads if lead.get('first_call_date')]
            elif qualification == 'unqualified':
                leads = [lead for lead in leads if not lead.get('first_call_date')]
            # Final status filter
            if final_status:
                leads = [lead for lead in leads if (lead.get('final_status') or '') == final_status]
            # Date filtering
            if date_filter == 'today':
                today_str = datetime.now().strftime('%Y-%m-%d')
                leads = [lead for lead in leads if lead.get('cre_assigned_at') and str(lead.get('cre_assigned_at')).startswith(today_str)]
            elif date_filter == 'range' and start_date and end_date:
                def in_range(ld):
                    dt = ld.get('cre_assigned_at')
                    if not dt:
                        return False
                    try:
                        dt_val = dt[:10]
                        return start_date <= dt_val <= end_date
                    except Exception:
                        return False
                leads = [lead for lead in leads if in_range(lead)]
            # Get all unique sources for this CRE's leads
            sources = sorted(list(set(lead.get('source', 'Unknown') for lead in leads)))
            # Pagination
            total_leads = len(leads)
            total_pages = (total_leads + per_page - 1) // per_page
            start_idx = (page - 1) * per_page
            end_idx = start_idx + per_page
            leads = leads[start_idx:end_idx]
        else:
            if search_uid:
                # Search by UID substring across all leads
                all_leads = safe_get_data('lead_master')
                leads = [lead for lead in all_leads if search_uid.lower() in str(lead.get('uid', '')).lower()]
                sources = sorted(list(set(lead.get('source', 'Unknown') for lead in leads)))
                total_leads = len(leads)
                total_pages = 1
                page = 1
            else:
                total_pages = 1
                page = 1
        # Add formatted CRE TAT for display
        def format_cre_tat(tat):
            try:
                tat = float(tat)
            except (TypeError, ValueError):
                return 'N/A'
            if tat < 60:
                return f"{int(tat)}s"
            elif tat < 3600:
                m = int(tat // 60)
                s = int(tat % 60)
                return f"{m}m {s}s"
            elif tat < 86400:
                h = int(tat // 3600)
                m = int((tat % 3600) // 60)
                s = int(tat % 60)
                return f"{h}h {m}m {s}s"
            else:
                d = int(tat // 86400)
                h = int((tat % 86400) // 3600)
                return f"{d} Days {h}h"
        for lead in leads:
            lead['cre_tat_display'] = format_cre_tat(lead.get('tat'))
        return render_template('manage_leads.html', cres=cres, selected_cre=selected_cre, leads=leads, sources=sources, selected_source=source, qualification=qualification, date_filter=date_filter, start_date=start_date, end_date=end_date, page=page, total_pages=total_pages, total_leads=total_leads, final_status=final_status)
    except Exception as e:
        flash(f'Error loading leads: {str(e)}', 'error')
        return render_template('manage_leads.html', cres=[], selected_cre=None, leads=[], sources=[], selected_source=None, qualification='all', date_filter='all', start_date=None, end_date=None, page=1, total_pages=1, total_leads=0, final_status='')


@app.route('/delete_leads', methods=['POST'])
@require_admin
def delete_leads():
    try:
        delete_type = request.form.get('delete_type')

        if delete_type == 'single':
            uid = request.form.get('uid')
            if not uid:
                return jsonify({'success': False, 'message': 'No UID provided'})

            # Delete from ps_followup_master first (foreign key constraint)
            supabase.table('ps_followup_master').delete().eq('lead_uid', uid).execute()

            # Delete from lead_master
            result = supabase.table('lead_master').delete().eq('uid', uid).execute()

            # Log deletion
            auth_manager.log_audit_event(
                user_id=session.get('user_id'),
                user_type=session.get('user_type'),
                action='LEAD_DELETED',
                resource='lead_master',
                resource_id=uid,
                details={'delete_type': 'single'}
            )

            return jsonify({'success': True, 'message': 'Lead deleted successfully'})

        elif delete_type == 'bulk':
            uids = request.form.getlist('uids')
            if not uids:
                return jsonify({'success': False, 'message': 'No leads selected'})

            # Delete from ps_followup_master first
            for uid in uids:
                supabase.table('ps_followup_master').delete().eq('lead_uid', uid).execute()

            # Delete from lead_master
            for uid in uids:
                supabase.table('lead_master').delete().eq('uid', uid).execute()

            # Log bulk deletion
            auth_manager.log_audit_event(
                user_id=session.get('user_id'),
                user_type=session.get('user_type'),
                action='LEADS_BULK_DELETED',
                resource='lead_master',
                details={'delete_type': 'bulk', 'count': len(uids), 'uids': uids}
            )

            return jsonify({'success': True, 'message': f'{len(uids)} leads deleted successfully'})

        else:
            return jsonify({'success': False, 'message': 'Invalid delete type'})

    except Exception as e:
        print(f"Error deleting leads: {e}")
        return jsonify({'success': False, 'message': f'Error deleting leads: {str(e)}'})


@app.route('/bulk_unassign_leads', methods=['POST'])
@require_admin
def bulk_unassign_leads():
    try:
        uids = request.form.getlist('uids')
        if not uids:
            return jsonify({'success': False, 'message': 'No leads selected'})

        # Update leads to unassigned
        for uid in uids:
            supabase.table('lead_master').update({
                'cre_name': None,
                'assigned': 'No',
                'ps_name': None
            }).eq('uid', uid).execute()

            # Also remove from PS followup if exists
            supabase.table('ps_followup_master').delete().eq('lead_uid', uid).execute()

        # Log bulk unassignment
        auth_manager.log_audit_event(
            user_id=session.get('user_id'),
            user_type=session.get('user_type'),
            action='LEADS_BULK_UNASSIGNED',
            resource='lead_master',
            details={'count': len(uids), 'uids': uids}
        )

        return jsonify({'success': True, 'message': f'{len(uids)} leads unassigned successfully'})

    except Exception as e:
        print(f"Error unassigning leads: {e}")
        return jsonify({'success': False, 'message': f'Error unassigning leads: {str(e)}'})


@app.route('/delete_cre/<int:cre_id>')
@require_admin
def delete_cre(cre_id):
    try:
        # Get the CRE details first
        cre_result = supabase.table('cre_users').select('*').eq('id', cre_id).execute()
        if not cre_result.data:
            flash('CRE not found', 'error')
            return redirect(url_for('manage_cre'))
        
        cre = cre_result.data[0]
        cre_name = cre.get('name')
        
        # Check if CRE has any pending leads in lead_master
        pending_leads_result = supabase.table('lead_master').select('id').eq('cre_name', cre_name).eq('final_status', 'Pending').execute()
        pending_count = len(pending_leads_result.data) if pending_leads_result.data else 0
        
        # Check if CRE has any pending leads in ps_followup_master
        ps_pending_result = supabase.table('ps_followup_master').select('id').eq('cre_name', cre_name).eq('final_status', 'Pending').execute()
        ps_pending_count = len(ps_pending_result.data) if ps_pending_result.data else 0
        
        total_pending = pending_count + ps_pending_count
        
        if total_pending > 0:
            flash(f'Cannot delete CRE {cre_name}. They have {total_pending} pending leads ({pending_count} in lead_master, {ps_pending_count} in ps_followup). Please transfer or close these leads first.', 'error')
            return redirect(url_for('manage_cre'))
        
        # If no pending leads, proceed with deletion
        # Update leads assigned to this CRE to unassigned
        supabase.table('lead_master').update({
            'cre_name': None,
            'assigned': 'No'
        }).eq('cre_name', cre_name).execute()
        
        # Update ps_followup_master leads
        supabase.table('ps_followup_master').update({
            'cre_name': None
        }).eq('cre_name', cre_name).execute()

        # Delete the CRE user
        supabase.table('cre_users').delete().eq('id', cre_id).execute()

        # Log CRE deletion
        auth_manager.log_audit_event(
            user_id=session.get('user_id'),
            user_type=session.get('user_type'),
            action='CRE_DELETED',
            resource='cre_users',
            resource_id=str(cre_id),
            details={'cre_name': cre_name}
        )

        flash(f'CRE {cre_name} has been deleted successfully', 'success')
        
    except Exception as e:
        print(f"Error deleting CRE: {str(e)}")
        flash('Error deleting CRE', 'error')

    return redirect(url_for('manage_cre'))


@app.route('/delete_ps/<int:ps_id>')
@require_admin
def delete_ps(ps_id):
    try:
        # Get the PS details first
        ps_result = supabase.table('ps_users').select('*').eq('id', ps_id).execute()
        if not ps_result.data:
            flash('PS not found', 'error')
            return redirect(url_for('manage_ps'))
        
        ps = ps_result.data[0]
        ps_name = ps.get('name')
        
        # Check if PS has any pending leads in ps_followup_master
        ps_pending_result = supabase.table('ps_followup_master').select('id').eq('ps_name', ps_name).eq('final_status', 'Pending').execute()
        ps_pending_count = len(ps_pending_result.data) if ps_pending_result.data else 0
        
        # Check if PS has any pending leads in walkin_table
        walkin_pending_result = supabase.table('walkin_table').select('id').eq('ps_assigned', ps_name).eq('status', 'Pending').execute()
        walkin_pending_count = len(walkin_pending_result.data) if walkin_pending_result.data else 0
        
        # Check if PS has any pending leads in activity_leads
        activity_pending_result = supabase.table('activity_leads').select('id').eq('ps_name', ps_name).eq('final_status', 'Pending').execute()
        activity_pending_count = len(activity_pending_result.data) if activity_pending_result.data else 0
        
        total_pending = ps_pending_count + walkin_pending_count + activity_pending_count
        
        if total_pending > 0:
            flash(f'Cannot delete PS {ps_name}. They have {total_pending} pending leads ({ps_pending_count} in ps_followup, {walkin_pending_count} in walkin, {activity_pending_count} in activity). Please transfer or close these leads first.', 'error')
            return redirect(url_for('manage_ps'))
        
        # If no pending leads, proceed with deletion
        # Update leads assigned to this PS to unassigned
        supabase.table('lead_master').update({
            'ps_name': None
        }).eq('ps_name', ps_name).execute()
        
        # Update ps_followup_master leads
        supabase.table('ps_followup_master').update({
            'ps_name': None
        }).eq('ps_name', ps_name).execute()
        
        # Update walkin_table leads
        supabase.table('walkin_table').update({
            'ps_assigned': None
        }).eq('ps_assigned', ps_name).execute()
        
        # Update activity_leads
        supabase.table('activity_leads').update({
            'ps_name': None
        }).eq('ps_name', ps_name).execute()

        # Delete the PS user
        supabase.table('ps_users').delete().eq('id', ps_id).execute()

        # Log PS deletion
        auth_manager.log_audit_event(
            user_id=session.get('user_id'),
            user_type=session.get('user_type'),
            action='PS_DELETED',
            resource='ps_users',
            resource_id=str(ps_id),
            details={'ps_name': ps_name}
        )

        flash(f'PS {ps_name} has been deleted successfully', 'success')
        
    except Exception as e:
        print(f"Error deleting PS: {str(e)}")
        flash('Error deleting PS', 'error')

    return redirect(url_for('manage_ps'))


@app.route('/edit_cre/<int:cre_id>', methods=['GET', 'POST'])
@require_admin
def edit_cre(cre_id):
    """Edit CRE user details"""
    try:
        if request.method == 'POST':
            email = request.form.get('email')
            phone = request.form.get('phone')
            
            if not email or not phone:
                flash('Email and phone are required', 'error')
                return redirect(url_for('edit_cre', cre_id=cre_id))
            
            # Update the CRE
            supabase.table('cre_users').update({
                'email': email,
                'phone': phone
            }).eq('id', cre_id).execute()
            
            flash('CRE details updated successfully', 'success')
            return redirect(url_for('manage_cre'))
        
        # Get CRE details for editing
        cre_result = supabase.table('cre_users').select('*').eq('id', cre_id).execute()
        if not cre_result.data:
            flash('CRE not found', 'error')
            return redirect(url_for('manage_cre'))
        
        cre = cre_result.data[0]
        return render_template('edit_cre.html', cre=cre)
        
    except Exception as e:
        print(f"Error editing CRE: {str(e)}")
        flash('Error editing CRE', 'error')
        return redirect(url_for('manage_cre'))


@app.route('/edit_ps/<int:ps_id>', methods=['GET', 'POST'])
@require_admin
def edit_ps(ps_id):
    """Edit PS user details"""
    try:
        if request.method == 'POST':
            email = request.form.get('email')
            phone = request.form.get('phone')
            
            if not email or not phone:
                flash('Email and phone are required', 'error')
                return redirect(url_for('edit_ps', ps_id=ps_id))
            
            # Update the PS
            supabase.table('ps_users').update({
                'email': email,
                'phone': phone
            }).eq('id', ps_id).execute()
            
            flash('PS details updated successfully', 'success')
            return redirect(url_for('manage_ps'))
        
        # Get PS details for editing
        ps_result = supabase.table('ps_users').select('*').eq('id', ps_id).execute()
        if not ps_result.data:
            flash('PS not found', 'error')
            return redirect(url_for('manage_ps'))
        
        ps = ps_result.data[0]
        return render_template('edit_ps.html', ps=ps)
        
    except Exception as e:
        print(f"Error editing PS: {str(e)}")
        flash('Error editing PS', 'error')
        return redirect(url_for('manage_ps'))


@app.route('/manage_rec', methods=['GET', 'POST'])
@require_admin
def manage_rec():
    branches = ["PORUR", "NUNGAMBAKKAM", "TIRUVOTTIYUR"]
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        name = request.form.get('name', '').strip()
        branch = request.form.get('branch', '').strip()
        if not all([username, password, name, branch]):
            flash('All fields are required', 'error')
        else:
            existing = supabase.table('rec_users').select('username').eq('username', username).execute()
            if existing.data:
                flash('Username already exists', 'error')
            else:
                password_hash = generate_password_hash(password)
                rec_data = {
                    'username': username,
                    'password_hash': password_hash,
                    'password': password,  # Store plain password as well
                    'name': name,
                    'branch': branch,
                    'is_active': True
                }
                try:
                    supabase.table('rec_users').insert(rec_data).execute()
                    flash('Receptionist user added successfully', 'success')
                except Exception as e:
                    flash(f'Error adding receptionist: {str(e)}', 'error')
    rec_users = safe_get_data('rec_users')
    return render_template('manage_rec.html', rec_users=rec_users, branches=branches)


@app.route('/delete_rec/<int:rec_id>')
@require_admin
def delete_rec(rec_id):
    try:
        supabase.table('rec_users').delete().eq('id', rec_id).execute()
        flash('Receptionist user deleted successfully', 'success')
    except Exception as e:
        flash(f'Error deleting receptionist: {str(e)}', 'error')
    return redirect(url_for('manage_rec'))


@app.route('/check_duplicate_lead', methods=['POST'])
@require_auth(['admin', 'cre'])
def check_duplicate_lead():
    """
    Check if a lead with the same phone number already exists.
    Returns duplicate information if found.
    """
    try:
        data = request.get_json()
        phone_number = data.get('phone_number', '').strip()
        source = data.get('source', '').strip()
        subsource = data.get('subsource', '').strip()
        
        if not phone_number:
            return jsonify({'success': False, 'message': 'Phone number is required'}), 400
        
        # Normalize phone number (remove all non-digits)
        normalized_phone = ''.join(filter(str.isdigit, phone_number))
        
        # Check in lead_master table
        result = supabase.table('lead_master').select('*').eq('customer_mobile_number', normalized_phone).execute()
        existing_leads = result.data or []
        
        if existing_leads:
            # Found existing lead(s)
            existing_lead = existing_leads[0]  # Get the first one
            
            # Check if this exact source-subsource combination already exists
            exact_match = any(
                lead.get('source') == source and lead.get('sub_source') == subsource 
                for lead in existing_leads
            )
            
            if exact_match:
                # This is a true duplicate - same phone, same source, same subsource
                return jsonify({
                    'success': True,
                    'is_duplicate': True,
                    'existing_lead': existing_lead,
                    'duplicate_type': 'exact_match',
                    'message': 'Lead with this phone number and source-subsource combination already exists'
                })
            else:
                # Phone exists but with different source/subsource
                # Get all existing sources for this phone number
                existing_sources = []
                for lead in existing_leads:
                    if lead.get('source') and lead.get('sub_source'):
                        existing_sources.append({
                            'source': lead.get('source'),
                            'sub_source': lead.get('sub_source')
                        })
                
                return jsonify({
                    'success': True,
                    'is_duplicate': True,
                    'existing_lead': existing_lead,
                    'existing_sources': existing_sources,
                    'duplicate_type': 'new_source',
                    'message': 'Phone number exists with different sources'
                })
        
        # Check in duplicate_leads table
        duplicate_result = supabase.table('duplicate_leads').select('*').eq('customer_mobile_number', normalized_phone).execute()
        duplicate_leads = duplicate_result.data or []
        
        if duplicate_leads:
            # Found in duplicate_leads table
            duplicate_lead = duplicate_leads[0]
            
            # Check if this exact source-subsource combination already exists in any slot
            exact_match = False
            existing_sources = []
            
            # Check all source slots (source1 to source10)
            for i in range(1, 11):
                source_field = f'source{i}'
                sub_source_field = f'sub_source{i}'
                
                if duplicate_lead.get(source_field) and duplicate_lead.get(sub_source_field):
                    existing_sources.append({
                        'source': duplicate_lead.get(source_field),
                        'sub_source': duplicate_lead.get(sub_source_field)
                    })
                    
                    # Check if this slot matches the new source/subsource
                    if (duplicate_lead.get(source_field) == source and 
                        duplicate_lead.get(sub_source_field) == subsource):
                        exact_match = True
            
            if exact_match:
                return jsonify({
                    'success': True,
                    'is_duplicate': True,
                    'existing_lead': duplicate_lead,
                    'duplicate_type': 'exact_match',
                    'message': 'Lead with this phone number and source-subsource combination already exists in duplicates'
                })
            else:
                return jsonify({
                    'success': True,
                    'is_duplicate': True,
                    'existing_lead': duplicate_lead,
                    'existing_sources': existing_sources,
                    'duplicate_type': 'new_source',
                    'message': 'Phone number exists in duplicates with different sources'
                })
        
        # No duplicate found
        return jsonify({
            'success': True,
            'is_duplicate': False,
            'message': 'No duplicate found'
        })
        
    except Exception as e:
        print(f"Error checking duplicate lead: {e}")
        return jsonify({'success': False, 'message': f'Error checking duplicate: {str(e)}'}), 500

@app.route('/add_lead', methods=['GET', 'POST'])
@require_cre
def add_lead():
    from datetime import datetime, date
    branches = ['PORUR', 'NUNGAMBAKKAM', 'TIRUVOTTIYUR']
    ps_users = safe_get_data('ps_users')
    if request.method == 'POST':
        customer_name = request.form.get('customer_name', '').strip()
        customer_mobile_number = request.form.get('customer_mobile_number', '').strip()
        source = request.form.get('source', '').strip()
        subsource = request.form.get('subsource', '').strip()
        lead_status = request.form.get('lead_status', '').strip()
        lead_category = request.form.get('lead_category', '').strip()
        model_interested = request.form.get('model_interested', '').strip()
        branch = request.form.get('branch', '').strip()
        ps_name = request.form.get('ps_name', '').strip()
        final_status = request.form.get('final_status', 'Pending').strip()
        follow_up_date = request.form.get('follow_up_date', '').strip()
        remark = request.form.get('remark', '').strip()
        is_duplicate_new_source = request.form.get('is_duplicate_new_source', '').strip()
        date_now = datetime.now().strftime('%Y-%m-%d')
        
        # Validation
        if not customer_name or not customer_mobile_number or not source or not subsource:
            flash('Please fill all required fields', 'error')
            return render_template('add_lead.html', branches=branches, ps_users=ps_users)
        
        # Validate follow_up_date is required when final_status is Pending
        if final_status == 'Pending' and not follow_up_date:
            flash('Follow-up date is required when final status is Pending', 'error')
            return render_template('add_lead.html', branches=branches, ps_users=ps_users)
        
        # Normalize phone number
        normalized_phone = ''.join(filter(str.isdigit, customer_mobile_number))
        
        # Check for duplicates if not already confirmed as new source
        if not is_duplicate_new_source:
            try:
                # Check in lead_master
                result = supabase.table('lead_master').select('*').eq('customer_mobile_number', normalized_phone).execute()
                existing_leads = result.data or []
                
                if existing_leads:
                    # Check for exact source-subsource match
                    exact_match = any(
                        lead.get('source') == source and lead.get('sub_source') == subsource 
                        for lead in existing_leads
                    )
                    
                    if exact_match:
                        flash('Lead with this phone number and source-subsource combination already exists!', 'error')
                        return render_template('add_lead.html', branches=branches, ps_users=ps_users)
                
                # Check in duplicate_leads
                duplicate_result = supabase.table('duplicate_leads').select('*').eq('customer_mobile_number', normalized_phone).execute()
                duplicate_leads = duplicate_result.data or []
                
                if duplicate_leads:
                    # Check if this exact source-subsource combination already exists in any slot
                    exact_match = False
                    for duplicate_lead in duplicate_leads:
                        # Check all source slots (source1 to source10)
                        for i in range(1, 11):
                            source_field = f'source{i}'
                            sub_source_field = f'sub_source{i}'
                            
                            if (duplicate_lead.get(source_field) == source and 
                                duplicate_lead.get(sub_source_field) == subsource):
                                exact_match = True
                                break
                        if exact_match:
                            break
                    
                    if exact_match:
                        flash('Lead with this phone number and source-subsource combination already exists in duplicates!', 'error')
                        return render_template('add_lead.html', branches=branches, ps_users=ps_users)
                        
            except Exception as e:
                print(f"Error checking duplicates: {e}")
                flash('Error checking for duplicates. Please try again.', 'error')
                return render_template('add_lead.html', branches=branches, ps_users=ps_users)
        
        # UID: Source initial (uppercase) + '-' + first 5 letters of name (no spaces, uppercase) + last 5 digits of phone
        src_initial = source[0].upper() if source else 'X'
        name_part = ''.join(customer_name.split()).upper()[:5]
        phone_part = normalized_phone[-5:] if len(normalized_phone) >= 5 else normalized_phone
        uid = f"{src_initial}-{name_part}{phone_part}"
        
        # CRE name from session
        cre_name = session.get('cre_name')
        
        # Prepare lead data
        lead_data = {
            'uid': uid,
            'date': date_now,
            'customer_name': customer_name,
            'customer_mobile_number': normalized_phone,
            'source': source,
            'sub_source': subsource,
            'lead_status': lead_status,
            'lead_category': lead_category,
            'model_interested': model_interested,
            'branch': branch,
            'ps_name': ps_name if ps_name else None,
            'final_status': final_status,
            'follow_up_date': follow_up_date if follow_up_date else None,
            'assigned': 'Yes' if cre_name else 'No',  # Set to Yes if CRE is adding the lead
            'cre_assigned_at': datetime.now().isoformat() if cre_name else None,
            'ps_assigned_at': datetime.now().isoformat() if ps_name else None,
            'created_at': datetime.now().isoformat(),
            'updated_at': datetime.now().isoformat(),
            'first_remark': remark,
            'cre_name': cre_name,
            'first_call_date': get_ist_timestamp()
        }
        try:
            # If this is a duplicate with new source, add to duplicate_leads table
            if is_duplicate_new_source:
                # Check if there's already a duplicate record
                existing_duplicate = supabase.table('duplicate_leads').select('*').eq('customer_mobile_number', normalized_phone).execute()
                
                if existing_duplicate.data:
                    # Add to existing duplicate record
                    duplicate_record = existing_duplicate.data[0]
                    # Find next available slot
                    next_slot = None
                    for i in range(1, 11):
                        source_field = f'source{i}'
                        if not duplicate_record.get(source_field):
                            next_slot = i
                            break
                    
                    if next_slot:
                        # Update the existing duplicate record
                        update_data = {
                            f'source{next_slot}': source,
                            f'sub_source{next_slot}': subsource,
                            f'date{next_slot}': date_now,
                            'duplicate_count': duplicate_record.get('duplicate_count', 0) + 1,
                            'updated_at': datetime.now().isoformat()
                        }
                        supabase.table('duplicate_leads').update(update_data).eq('id', duplicate_record['id']).execute()
                        flash(f'Lead added to existing duplicate record with new source: {source} - {subsource}', 'success')
                    else:
                        flash('Error: Duplicate record is full (max 10 sources reached)', 'error')
                else:
                    # Create new duplicate record
                    original_lead = supabase.table('lead_master').select('*').eq('customer_mobile_number', normalized_phone).execute()
                    if original_lead.data:
                        original = original_lead.data[0]
                        # Create duplicate record with proper structure
                        duplicate_data = {
                            'uid': uid,
                            'customer_mobile_number': normalized_phone,
                            'customer_name': customer_name,
                            'original_lead_id': original['id'],
                            'source1': original['source'],
                            'sub_source1': original.get('sub_source'),
                            'date1': original['date'],
                            'source2': source,
                            'sub_source2': subsource,
                            'date2': date_now,
                            'duplicate_count': 2,
                            'created_at': datetime.now().isoformat(),
                            'updated_at': datetime.now().isoformat()
                        }
                        supabase.table('duplicate_leads').insert(duplicate_data).execute()
                        flash(f'Lead added to duplicates with new source: {source} - {subsource}', 'success')
                    else:
                        flash('Error: Original lead not found for duplicate creation', 'error')
            else:
                supabase.table('lead_master').insert(lead_data).execute()
                
                # Track the initial call attempt for fresh leads
                if lead_status:
                    track_cre_call_attempt(
                        uid=uid,
                        cre_name=cre_name,
                        call_no='first',
                        lead_status=lead_status,
                        call_was_recorded=True,  # Fresh leads always have first_call_date recorded
                        follow_up_date=follow_up_date if follow_up_date else None,
                        remarks=remark if remark else None
                    )
                
                # Create PS followup if PS is assigned during lead creation
                if ps_name:
                    ps_user = next((ps for ps in ps_users if ps['name'] == ps_name), None)
                    if ps_user:
                        create_or_update_ps_followup(lead_data, ps_name, ps_user['branch'])
                        
                        # Send email notification to PS
                        try:
                            socketio.start_background_task(send_email_to_ps, ps_user['email'], ps_user['name'], lead_data, cre_name)
                            flash(f'Lead added successfully and assigned to {ps_name}! Email notification sent.', 'success')
                        except Exception as e:
                            print(f"Error sending email: {e}")
                            flash(f'Lead added successfully and assigned to {ps_name}! (Email notification failed)', 'warning')
                    else:
                        flash('Lead added successfully! (PS assignment failed)', 'warning')
                else:
                    flash('Lead added successfully!', 'success')
                
                # Trigger auto-assign for new leads if they're not already assigned
                if not cre_name:  # Only auto-assign if no CRE was manually assigned
                    try:
                        print(f"🔄 New lead added from {source}, triggering auto-assign...")
                        # Run auto-assign in background to avoid blocking the response
                        socketio.start_background_task(auto_assign_new_leads_for_source, source)
                        print(f"✅ Auto-assign triggered for {source}")
                    except Exception as e:
                        print(f"❌ Error triggering auto-assign for new lead: {e}")
            
            return redirect(url_for('cre_dashboard'))
        except Exception as e:
            flash(f'Error adding lead: {str(e)}', 'error')
            return render_template('add_lead.html', branches=branches, ps_users=ps_users)
    return render_template('add_lead.html', branches=branches, ps_users=ps_users)

# Add import for optimized operations
from optimized_lead_operations import create_optimized_operations

@app.route('/add_lead_optimized', methods=['POST'])
@require_cre
def add_lead_optimized():
    """
    Optimized lead creation endpoint with improved performance
    """
    try:
        from datetime import datetime
        
        # Get form data
        customer_name = request.form.get('customer_name', '').strip()
        customer_mobile_number = request.form.get('customer_mobile_number', '').strip()
        source = request.form.get('source', '').strip()
        subsource = request.form.get('subsource', '').strip()
        lead_status = request.form.get('lead_status', '').strip()
        lead_category = request.form.get('lead_category', '').strip()
        model_interested = request.form.get('model_interested', '').strip()
        branch = request.form.get('branch', '').strip()
        ps_name = request.form.get('ps_name', '').strip()
        final_status = request.form.get('final_status', 'Pending').strip()
        follow_up_date = request.form.get('follow_up_date', '').strip()
        remark = request.form.get('remark', '').strip()
        date_now = datetime.now().strftime('%Y-%m-%d')
        
        # Validation
        if not customer_name or not customer_mobile_number or not source or not subsource:
            return jsonify({
                'success': False,
                'message': 'Please fill all required fields'
            })
        
        # Validate follow_up_date is required when final_status is Pending
        if final_status == 'Pending' and not follow_up_date:
            return jsonify({
                'success': False,
                'message': 'Follow-up date is required when final status is Pending'
            })
        
        # Normalize phone number
        normalized_phone = ''.join(filter(str.isdigit, customer_mobile_number))
        
        # CRE name from session
        cre_name = session.get('cre_name')
        
        # Prepare lead data
        lead_data = {
            'date': date_now,
            'customer_name': customer_name,
            'customer_mobile_number': normalized_phone,
            'source': source,
            'sub_source': subsource,
            'lead_status': lead_status,
            'lead_category': lead_category,
            'model_interested': model_interested,
            'branch': branch,
            'ps_name': ps_name if ps_name else None,
            'final_status': final_status,
            'follow_up_date': follow_up_date if follow_up_date else None,
            'assigned': 'Yes' if cre_name else 'No',
            'cre_assigned_at': datetime.now().isoformat() if cre_name else None,
            'ps_assigned_at': datetime.now().isoformat() if ps_name else None,
            'created_at': datetime.now().isoformat(),
            'updated_at': datetime.now().isoformat(),
            'first_remark': remark,
            'cre_name': cre_name,
            'first_call_date': get_ist_timestamp()
        }
        
        # Get PS branch if PS is assigned
        ps_branch = None
        if ps_name:
            ps_users = safe_get_data('ps_users')
            ps_user = next((ps for ps in ps_users if ps['name'] == ps_name), None)
            if ps_user:
                ps_branch = ps_user['branch']
        
        # Use optimized operations
        optimized_ops = create_optimized_operations(supabase)
        result = optimized_ops.create_lead_optimized(lead_data, cre_name, ps_name, ps_branch)
        
        if result['success']:
            return jsonify({
                'success': True,
                'message': result['message'],
                'uid': result['uid'],
                'execution_time': f"{result['execution_time']:.3f}s"
            })
        else:
            return jsonify({
                'success': False,
                'message': result['message'],
                'execution_time': f"{result['execution_time']:.3f}s"
            })
            
    except Exception as e:
        print(f"Error in optimized lead creation: {e}")
        return jsonify({
            'success': False,
            'message': f'Error creating lead: {str(e)}'
        })

@app.route('/add_lead_with_cre', methods=['POST'])
@require_admin  # <-- Change to @require_cre if you want CREs to use it, or create a custom decorator for both
def add_lead_with_cre():
    """
    Add a new lead with minimal required columns.
    - Checks for duplicate by last 10 digits of phone number.
    - If duplicate, returns UID of existing lead.
    - Only fills: uid, customer_name, customer_mobile_number, source, date, assigned.
    - Source is always 'Google(Web)', UID uses 'G' as the source character.
    """
    try:
        # Check if this is an AJAX request
        is_ajax = request.headers.get('Content-Type') == 'application/json'
        
        if is_ajax:
            # Handle JSON request
            data = request.get_json()
            customer_name = data.get('customer_name', '').strip()
            customer_mobile_number = data.get('customer_mobile_number', '').strip()
            source = data.get('source', 'GOOGLE').strip()
            subsource = data.get('subsource', '').strip()
            assigned_cre_id = data.get('assigned_cre')
        else:
            # Handle form request
            customer_name = request.form.get('customer_name', '').strip()
            customer_mobile_number = request.form.get('customer_mobile_number', '').strip()
            source = request.form.get('source', 'GOOGLE').strip()
            subsource = request.form.get('subsource', '').strip()
            assigned_cre_id = request.form.get('assigned_cre')
        
        assigned = "Yes"
        date_now = datetime.now().strftime('%Y-%m-%d')

        # Validate required fields
        if not customer_name or not customer_mobile_number or not source or not subsource:
            if is_ajax:
                return jsonify({
                    'success': False,
                    'message': 'Customer name, mobile number, source, and subsource are required'
                })
            else:
                flash('Customer name, mobile number, source, and subsource are required', 'error')
                return redirect('/assign_leads')

        # Normalize phone number to last 10 digits
        mobile_digits = ''.join(filter(str.isdigit, customer_mobile_number))[-10:]
        
        if len(mobile_digits) != 10:
            if is_ajax:
                return jsonify({
                    'success': False,
                    'message': 'Invalid mobile number. Please provide a 10-digit number.'
                })
            else:
                flash('Invalid mobile number. Please provide a 10-digit number.', 'error')
                return redirect('/assign_leads')

        # Check if this is a duplicate with new source from form
        if is_ajax:
            is_duplicate_new_source = data.get('is_duplicate_new_source', '').strip() == 'true'
        else:
            is_duplicate_new_source = request.form.get('is_duplicate_new_source', '').strip() == 'true'
        
        # Check for duplicate by phone number and source-subsource combination
        existing_leads = supabase.table('lead_master').select('*').eq('customer_mobile_number', mobile_digits).execute()
        duplicate_leads = supabase.table('duplicate_leads').select('*').eq('customer_mobile_number', mobile_digits).execute()
        
        original_lead = None
        
        # Only check for duplicates if not already confirmed as duplicate with new source
        if not is_duplicate_new_source:
            # Check in lead_master table
            if existing_leads.data:
                original_lead = existing_leads.data[0]
                # Check if this exact source-subsource combination already exists
                if original_lead.get('source') == source and original_lead.get('sub_source') == subsource:
                    if is_ajax:
                        return jsonify({
                            'success': False,
                            'message': f'Lead with this phone number and source-subsource combination already exists. UID: {original_lead["uid"]}',
                            'uid': original_lead["uid"]
                        })
                    else:
                        flash(f'Lead with this phone number and source-subsource combination already exists. UID: {original_lead["uid"]}', 'error')
                        return redirect('/assign_leads')
                else:
                    # Phone exists but with different source/subsource - this is a duplicate with new source
                    is_duplicate_new_source = True
            
            # Check in duplicate_leads table
            if duplicate_leads.data:
                duplicate_lead = duplicate_leads.data[0]
                # Check if this exact source-subsource combination already exists in any slot
                exact_match = False
                for i in range(1, 11):
                    source_field = f'source{i}'
                    sub_source_field = f'sub_source{i}'
                    
                    if (duplicate_lead.get(source_field) == source and 
                        duplicate_lead.get(sub_source_field) == subsource):
                        exact_match = True
                        break
                
                if exact_match:
                    if is_ajax:
                        return jsonify({
                            'success': False,
                            'message': f'Lead with this phone number and source-subsource combination already exists in duplicates. UID: {duplicate_lead["uid"]}',
                            'uid': duplicate_lead["uid"]
                        })
                    else:
                        flash(f'Lead with this phone number and source-subsource combination already exists in duplicates. UID: {duplicate_lead["uid"]}', 'error')
                        return redirect('/assign_leads')
                else:
                    # Phone exists in duplicates but with different source/subsource
                    is_duplicate_new_source = True
        else:
            # If is_duplicate_new_source is true, we need to get the original lead
            if existing_leads.data:
                original_lead = existing_leads.data[0]

        # Generate UID using the correct function based on source
        # Map source to UID source character
        source_mapping = {
            'GOOGLE': 'Google',
            'META': 'Meta',
            'BTL': 'BTL',
            'OEM': 'OEM'
        }
        uid_source = source_mapping.get(source, 'Google')
        
        sequence = 1
        uid = generate_uid(uid_source, mobile_digits, sequence)
        # Ensure UID is unique
        while supabase.table('lead_master').select('uid').eq('uid', uid).execute().data:
            sequence += 1
            uid = generate_uid(uid_source, mobile_digits, sequence)

        # Get assigned CRE ID from form
        if is_ajax:
            assigned_cre_id = data.get('assigned_cre')
        else:
            assigned_cre_id = request.form.get('assigned_cre')
        print(f"🔍 Raw assigned_cre_id from form: '{assigned_cre_id}' (type: {type(assigned_cre_id)})")
        
        cre_name = None
        if assigned_cre_id and assigned_cre_id.strip():
            try:
                # Convert to integer if it's a string
                cre_id = int(assigned_cre_id) if isinstance(assigned_cre_id, str) else assigned_cre_id
                print(f"🔍 Looking up CRE with ID: {cre_id}")
                
                cre_data = supabase.table('cre_users').select('name').eq('id', cre_id).execute()
                if cre_data.data:
                    cre_name = cre_data.data[0]['name']
                    print(f"✅ Found CRE: {cre_name} for ID: {cre_id}")
                else:
                    print(f"❌ No CRE found for ID: {cre_id}")
                    print(f"Available CRE IDs: {[cre['id'] for cre in supabase.table('cre_users').select('id,name').execute().data]}")
            except Exception as e:
                print(f"❌ Error fetching CRE data: {e}")
                import traceback
                traceback.print_exc()
        else:
            print("❌ No assigned_cre_id provided in form or it's empty")
            print(f"All form fields: {dict(request.form)}")

        # Prepare lead data (only required columns)
        # Set assigned based on whether CRE is assigned
        assigned_status = "Yes" if cre_name else "No"
        
        lead_data = {
            'uid': uid,
            'customer_name': customer_name,
            'customer_mobile_number': mobile_digits,
            'source': source,
            'sub_source': subsource,
            'date': date_now,   
            'assigned': assigned_status,
            'final_status': 'Pending',
            'cre_name': cre_name,
            'lead_status': 'Pending',
            'lead_category': None,  # Keep as null for assign leads page
            'cre_assigned_at': datetime.now().isoformat() if cre_name else None,
            'ps_assigned_at': None,  # Will be set when PS is assigned
            'created_at': datetime.now().isoformat(),
            'updated_at': datetime.now().isoformat(),
            'ps_assigned_at': None  # Added this line to ensure ps_assigned_at is set
        }
        print('=== DEBUG INFO ===')
        print('Form data:', dict(request.form))
        print('Assigned CRE ID:', assigned_cre_id)
        print('CRE name fetched:', cre_name)
        print('Is duplicate new source:', is_duplicate_new_source)
        print('Lead data to insert:', lead_data)
        print('==================')

        # Insert lead based on whether it's a duplicate with new source
        if is_duplicate_new_source:
            print("=== DUPLICATE HANDLING ===")
            print(f"Original lead: {original_lead}")
            print(f"Duplicate leads: {duplicate_leads.data}")
            
            if original_lead:
                # Create new duplicate record
                duplicate_data = {
                    'uid': uid,
                    'customer_mobile_number': mobile_digits,
                    'customer_name': customer_name,
                    'original_lead_id': original_lead['id'],
                    'source1': original_lead['source'],
                    'sub_source1': original_lead.get('sub_source'),
                    'date1': original_lead['date'],
                    'source2': source,
                    'sub_source2': subsource,
                    'date2': date_now,
                    'duplicate_count': 2,
                    'created_at': datetime.now().isoformat(),
                    'updated_at': datetime.now().isoformat()
                }
                print(f"Creating duplicate record: {duplicate_data}")
                result = supabase.table('duplicate_leads').insert(duplicate_data).execute()
                if result.data:
                    print("Duplicate record created successfully")
                    if is_ajax:
                        return jsonify({'success': True, 'message': 'Lead added to duplicates with new source', 'uid': uid})
                    else:
                        flash('Lead added to duplicates with new source', 'success')
                        return redirect('/assign_leads')
                else:
                    print("Failed to create duplicate record")
                    if is_ajax:
                        return jsonify({'success': False, 'message': 'Failed to add duplicate lead'})
                    else:
                        flash('Failed to add duplicate lead', 'error')
                        return redirect('/assign_leads')
            elif duplicate_leads.data:
                # Add to existing duplicate record
                duplicate_record = duplicate_leads.data[0]
                # Find next available slot
                next_slot = None
                for i in range(1, 11):
                    source_field = f'source{i}'
                    if not duplicate_record.get(source_field):
                        next_slot = i
                        break
                
                if next_slot:
                    # Update the existing duplicate record
                    update_data = {
                        f'source{next_slot}': source,
                        f'sub_source{next_slot}': subsource,
                        f'date{next_slot}': date_now,
                        'duplicate_count': duplicate_record.get('duplicate_count', 0) + 1,
                        'updated_at': datetime.now().isoformat()
                    }
                    result = supabase.table('duplicate_leads').update(update_data).eq('id', duplicate_record['id']).execute()
                    if result.data:
                        if is_ajax:
                            return jsonify({'success': True, 'message': 'Lead added to existing duplicate record', 'uid': uid})
                        else:
                            flash('Lead added to existing duplicate record', 'success')
                            return redirect('/assign_leads')
                    else:
                        if is_ajax:
                            return jsonify({'success': False, 'message': 'Failed to update duplicate record'})
                        else:
                            flash('Failed to update duplicate record', 'error')
                            return redirect('/assign_leads')
                else:
                    if is_ajax:
                        return jsonify({'success': False, 'message': 'Duplicate record is full (max 10 sources reached)'})
                    else:
                        flash('Duplicate record is full (max 10 sources reached)', 'error')
                        return redirect('/assign_leads')
            else:
                # Phone exists in duplicate_leads but not in lead_master - this shouldn't happen
                if is_ajax:
                    return jsonify({'success': False, 'message': 'Error: Original lead not found for duplicate creation'})
                else:
                    flash('Error: Original lead not found for duplicate creation', 'error')
                    return redirect('/assign_leads')
        else:
            # Insert as new lead
            print("=== FRESH LEAD INSERTION ===")
            print(f"Inserting fresh lead: {lead_data}")
            result = supabase.table('lead_master').insert(lead_data).execute()
            if result.data:
                print("Fresh lead inserted successfully")
                
                # Trigger auto-assign for new leads if they're not already assigned
                if not assigned_cre_id:  # Only auto-assign if no CRE was manually assigned
                    try:
                        print(f"🔄 New lead added from {source}, triggering auto-assign...")
                        # Run auto-assign in background to avoid blocking the response
                        socketio.start_background_task(auto_assign_new_leads_for_source, source)
                        print(f"✅ Auto-assign triggered for {source}")
                    except Exception as e:
                        print(f"❌ Error triggering auto-assign for new lead: {e}")
                
                if is_ajax:
                    return jsonify({'success': True, 'message': 'Lead added successfully', 'uid': uid})
                else:
                    flash('Lead added successfully', 'success')
                    return redirect('/assign_leads')
            else:
                print("Failed to insert fresh lead")
                if is_ajax:
                    return jsonify({'success': False, 'message': 'Failed to add lead'})
                else:
                    flash('Failed to add lead', 'error')
                    return redirect('/assign_leads')

    except Exception as e:
        print(f"Error adding lead with CRE: {e}")
        import traceback
        traceback.print_exc()
        if is_ajax:
            return jsonify({'success': False, 'message': f'Error adding lead: {str(e)}'})
        else:
            flash(f'Error adding lead: {str(e)}', 'error')
            return redirect('/assign_leads')

@app.route('/cre_dashboard')
@require_cre
def cre_dashboard():
    import time
    from datetime import datetime, date
    start_time = time.time()
    cre_name = session.get('cre_name')
    
    # Get status parameter for Won/Lost toggle
    status = request.args.get('status', 'lost')
    # Get tab and sub_tab parameters for redirection
    tab = request.args.get('tab', '')
    sub_tab = request.args.get('sub_tab', '')
    
    today = date.today()
    today_str = today.isoformat()

    # --- AUTO-INCREMENT LOGIC FOR UNATTENDED LEADS (CRE) ---
    # 1. Regular leads (lead_master)
    all_leads = safe_get_data('lead_master', {'cre_name': cre_name})
    for lead in all_leads:
        follow_up_date = lead.get('follow_up_date')
        final_status = lead.get('final_status')
        if follow_up_date and str(follow_up_date) < today_str and final_status not in ['Won', 'Lost']:
            supabase.table('lead_master').update({'follow_up_date': today_str}).eq('uid', lead.get('uid')).execute()



    # 3. Event leads (activity_leads)
    event_event_leads = safe_get_data('activity_leads', {'cre_assigned': cre_name})
    for lead in event_event_leads:
        cre_followup_date = lead.get('cre_followup_date')
        final_status = lead.get('final_status')
        if cre_followup_date and str(cre_followup_date)[:10] < today_str and final_status not in ['Won', 'Lost']:
            supabase.table('activity_leads').update({'cre_followup_date': today_str}).eq('activity_uid', lead.get('activity_uid')).execute()

    # --- Date Filter Logic ---
    filter_type = request.args.get('filter_type', 'all')
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')

    all_leads = safe_get_data('lead_master', {'cre_name': cre_name})
    
    # Apply date filtering using the 'date' column
    if filter_type == 'today':
        today_str = date.today().isoformat()
        all_leads = [lead for lead in all_leads if lead.get('date') == today_str]
    elif filter_type == 'range' and start_date_str and end_date_str:
        # Filter leads where lead['date'] falls within the range
        all_leads = [
            lead for lead in all_leads 
            if lead.get('date') and start_date_str <= lead['date'] <= end_date_str
        ]
    # For 'all' filter type, we keep all leads without date filtering

    print(f"Date filter applied - Type: {filter_type}, Leads count: {len(all_leads)}")
    if filter_type == 'range':
        print(f"Date range: {start_date_str} to {end_date_str}")

    # Fetch event leads assigned to this CRE
    event_event_leads = safe_get_data('activity_leads', {'cre_assigned': cre_name})

    print("Fetched leads for CRE:", cre_name, "Count:", len(all_leads))
    print("Status parameter from URL:", status)

    # Initialize buckets for leads for mutual exclusivity
    untouched_leads = []
    called_leads = []
    follow_up_leads = []  # New list for "Call me Back" leads
    attended_leads = []
    assigned_to_ps = []
    won_leads = []
    lost_leads = []
    
    # Get current month for filtering
    current_month = datetime.now().strftime('%Y-%m')
    print(f"DEBUG: Filtering for current month: {current_month}")

    non_contact_statuses = ['RNR', 'Busy on another Call', 'Call me Back', 'Call Disconnected', 'Call not Connected']

    for lead in all_leads:
        lead_status = (lead.get('lead_status') or '').strip()
        final_status = lead.get('final_status')
        has_first_call = lead.get('first_call_date') is not None

        if final_status == 'Won':
            # Only count leads won in current month
            if lead.get('won_timestamp'):
                won_month = str(lead.get('won_timestamp'))[:7]  # Get YYYY-MM part
                if won_month == current_month:
                    won_leads.append(lead)
                    print(f"DEBUG: Added to won leads - UID: {lead.get('uid')}, Won: {lead.get('won_timestamp')}")
            continue
        if final_status == 'Lost':
            # Only count leads lost in current month
            if lead.get('lost_timestamp'):
                lost_month = str(lead.get('lost_timestamp'))[:7]  # Get YYYY-MM part
                if lost_month == current_month:
                    lost_leads.append(lead)
                    print(f"DEBUG: Added to lost leads - UID: {lead.get('uid')}, Lost: {lead.get('lost_timestamp')}")
            continue

        # PS Assigned: any lead with a PS assigned in current month
        if lead.get('ps_name') and lead.get('ps_assigned_at'):
            ps_assigned_month = str(lead.get('ps_assigned_at'))[:7]  # Get YYYY-MM part
            if ps_assigned_month == current_month:
                assigned_to_ps.append(lead)
                print(f"DEBUG: Added to PS assigned - UID: {lead.get('uid')}, PS: {lead.get('ps_name')}, Assigned: {lead.get('ps_assigned_at')}")

        # Pending Leads: final_status == 'Pending' AND has first_call_date
        if final_status == 'Pending' and has_first_call:
            attended_leads.append(lead)

        # Untouched Leads (Fresh leads that are still pending)
        if not has_first_call and lead_status == 'Pending':
            untouched_leads.append(lead)
            continue

        # Called Fresh Leads: Non-contact status on FIRST update only
        if lead_status in non_contact_statuses and not has_first_call:
            # Separate "Call me Back" leads into follow_up_leads
            if lead_status == 'Call me Back':
                follow_up_leads.append(lead)
            else:
                called_leads.append(lead)
            continue

    print(f"Won leads count (current month): {len(won_leads)}")
    print(f"Lost leads count (current month): {len(lost_leads)}")
    print(f"PS assigned count (current month): {len(assigned_to_ps)}")
    
    untouched_count = len(untouched_leads)
    called_count = len(called_leads)
    follow_up_count = len(follow_up_leads)
    total_fresh_leads = untouched_count + called_count + follow_up_count

    fresh_leads_sorted = sorted(
        untouched_leads + called_leads + follow_up_leads,
        key=lambda l: l.get('date') or '',  # Sort by 'date' column
        reverse=True
    )

    # Get today's followups
    today = date.today()
    today_str = today.isoformat()
    todays_followups = [
        lead for lead in all_leads
        if (lead.get('follow_up_date') and str(lead.get('follow_up_date')).startswith(today_str)
        and lead.get('final_status') not in ['Won', 'Lost'])
    ]

    # Add event leads with today's cre_followup_date to the follow-up list
    event_leads_today = []
    for lead in event_event_leads:
        cre_followup_date = lead.get('cre_followup_date')
        if cre_followup_date and str(cre_followup_date)[:10] == today_str:
            event_lead_row = {
                'is_event_lead': True,
                'activity_uid': lead.get('activity_uid'),
                'customer_name': lead.get('customer_name'),
                'customer_phone_number': lead.get('customer_phone_number'),
                'lead_status': lead.get('lead_status'),
                'location': lead.get('location'),
                'activity_name': lead.get('activity_name'),
            }
            event_leads_today.append(event_lead_row)
    
    todays_followups.extend(event_leads_today)

    print(f"[PERF] cre_dashboard TOTAL took {time.time() - start_time:.3f} seconds")
    print(f"=== DASHBOARD SUMMARY ===")
    print(f"Current Month Filter: {current_month}")
    print(f"Fresh Leads: {total_fresh_leads} (Untouched: {untouched_count}, Called: {called_count}, Follow-up: {follow_up_count})")
    print(f"Today's Follow-ups: {len(todays_followups)}")
    print(f"Pending Leads: {len(attended_leads)}")
    print(f"PS Assigned (Current Month): {len(assigned_to_ps)}")
    print(f"Won Leads (Current Month): {len(won_leads)}")
    print(f"Lost Leads (Current Month): {len(lost_leads)}")
    print(f"========================")
    
    return render_template(
        'cre_dashboard.html',
        untouched_count=untouched_count,
        called_count=called_count,
        follow_up_count=follow_up_count,
        total_fresh_leads=total_fresh_leads,
        fresh_leads_sorted=fresh_leads_sorted,
        untouched_leads=untouched_leads,
        called_leads=called_leads,
        follow_up_leads=follow_up_leads,
        pending_leads=attended_leads,
        todays_followups=todays_followups,
        attended_leads=attended_leads,
        assigned_to_ps=assigned_to_ps,
        won_leads=won_leads,
        lost_leads=lost_leads,
        event_event_leads=event_event_leads,
        filter_type=filter_type,  # Pass filter type to template
        start_date=start_date_str,  # Pass start date to template
        end_date=end_date_str,  # Pass end date to template
        status=status,  # Pass status parameter to template
        return_tab=tab,  # Pass tab parameter to template
        return_sub_tab=sub_tab  # Pass sub_tab parameter to template
    )

@app.route('/update_lead/<uid>', methods=['GET', 'POST'])
@require_cre
def update_lead(uid):
    # Get return_tab parameter for redirection
    return_tab = request.args.get('return_tab', '')
    try:
        # Get lead data
        lead_result = supabase.table('lead_master').select('*').eq('uid', uid).execute()
        if not lead_result.data:
            flash('Lead not found', 'error')
            return redirect(url_for('cre_dashboard'))

        lead_data = lead_result.data[0]

        # Verify this lead belongs to the current CRE
        if lead_data.get('cre_name') != session.get('cre_name'):
            flash('Access denied - This lead is not assigned to you', 'error')
            return redirect(url_for('cre_dashboard'))

        # Get next call info
        next_call, completed_calls = get_next_call_info(lead_data)

        # Get PS users for branch selection
        ps_users = safe_get_data('ps_users')

        # Model options
        rizta_models = [
            'Rizta S Mono (2.9 kWh)',
            'Rizta S Super Matte (2.9 kWh)',
            'Rizta Z Mono (2.9 kWh)',
            'Rizta Z Duo (2.9 kWh)',
            'Rizta Z Super Matte (2.9 kWh)',
            'Rizta Z Mono (3.7 kWh)',
            'Rizta Z Duo (3.7 kWh)',
            'Rizta Z Super Matte (3.7 kWh)'
        ]

        x450_models = [
            '450 X (2.9 kWh)',
            '450 X (3.7 kWh)',
            '450 X (2.9 kWh) Pro Pack',
            '450 X (3.7 kWh) Pro Pack',
            '450 Apex STD'
        ]

        branches = ['PORUR', 'NUNGAMBAKKAM', 'TIRUVOTTIYUR']

        lead_statuses = [
            'Busy on another Call', 'RNR', 'Call me Back', 'Interested',
            'Not Interested', 'Did Not Inquire', 'Lost to Competition',
            'Lost to Co Dealer', 'Call Disconnected', 'Wrong Number'
        ]

        if request.method == 'POST':
            update_data = {}
            lead_status = request.form.get('lead_status', '')
            follow_up_date = request.form.get('follow_up_date', '')
            call_remark = request.form.get('call_remark', '')

            # Lock follow_up_date and set final_status for certain statuses
            lock_statuses = ['Booked', 'Retailed']
            lost_statuses = ['Not Interested', 'Lost to Codealer', 'Lost to Competition']
            if lead_status in lock_statuses:
                update_data['lead_status'] = lead_status
                update_data['follow_up_date'] = follow_up_date or lead_data.get('follow_up_date')
                update_data['final_status'] = 'Won' if lead_status in ['Booked', 'Retailed'] else lead_data.get('final_status')
            elif lead_status in lost_statuses:
                update_data['lead_status'] = lead_status
                update_data['follow_up_date'] = follow_up_date or lead_data.get('follow_up_date')
                update_data['final_status'] = 'Lost'
            else:
                if lead_status:
                    update_data['lead_status'] = lead_status

            if request.form.get('customer_name'):
                update_data['customer_name'] = request.form['customer_name']

            if request.form.get('source'):
                update_data['source'] = request.form['source']

            if request.form.get('lead_category'):
                update_data['lead_category'] = request.form['lead_category']

            if request.form.get('model_interested'):
                update_data['model_interested'] = request.form['model_interested']

            if request.form.get('branch'):
                update_data['branch'] = request.form['branch']

            # Handle PS assignment and email notification
            ps_name = request.form.get('ps_name')
            if ps_name and ps_name != lead_data.get('ps_name'):
                update_data['ps_name'] = ps_name
                update_data['assigned'] = 'Yes'
                update_data['ps_assigned_at'] = datetime.now().isoformat()

                # Get PS details for followup creation
                ps_user = next((ps for ps in ps_users if ps['name'] == ps_name), None)
                if ps_user:
                    # Create PS followup record
                    updated_lead_data = {**lead_data, **update_data}
                    create_or_update_ps_followup(updated_lead_data, ps_name, ps_user['branch'])

                # Send email to PS
                try:
                    if ps_user:
                        lead_data_for_email = {**lead_data, **update_data}
                        socketio.start_background_task(send_email_to_ps, ps_user['email'], ps_user['name'], lead_data_for_email, session.get('cre_name'))
                        flash(f'Lead assigned to {ps_name} and email notification sent', 'success')
                    else:
                        flash(f'Lead assigned to {ps_name}', 'success')
                except Exception as e:
                    print(f"Error sending email: {e}")
                    flash(f'Lead assigned to {ps_name} (email notification failed)', 'warning')

            if follow_up_date:
                update_data['follow_up_date'] = follow_up_date

            if request.form.get('final_status'):
                final_status = request.form['final_status']
                update_data['final_status'] = final_status
                if final_status == 'Won':
                    update_data['won_timestamp'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                elif final_status == 'Lost':
                    update_data['lost_timestamp'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            # Handle call dates and remarks - record for all statuses including RNR
            if request.form.get('call_date') and call_remark:
                combined_remark = f"{lead_status}, {call_remark}"
                # For all 7 calls, use the correct schema columns
                call_names = ['first', 'second', 'third', 'fourth', 'fifth', 'sixth', 'seventh']
                if next_call in call_names:
                    # Always use current timestamp, ignore the form date input
                    timestamp = get_ist_timestamp()
                    print(f"DEBUG: Setting {next_call}_call_date to '{timestamp}'")
                    update_data[f'{next_call}_call_date'] = timestamp
                    update_data[f'{next_call}_remark'] = combined_remark
                    
                    # Set first_call_date when the first call is made
                    if next_call == 'first' and not lead_data.get('first_call_date'):
                        update_data['first_call_date'] = get_ist_timestamp()
                # Notification removed - no longer sending notifications between PS and CRE

            try:
                # Track the call attempt before updating the lead - ALWAYS track regardless of status
                if lead_status:
                    # Determine if this attempt resulted in a recorded call
                    call_was_recorded = bool(request.form.get('call_date') and call_remark)
                    # Track the attempt - this ensures RNR and other statuses are properly tracked
                    track_cre_call_attempt(
                        uid=uid,
                        cre_name=session.get('cre_name'),
                        call_no=next_call,
                        lead_status=lead_status,
                        call_was_recorded=call_was_recorded,
                        follow_up_date=follow_up_date if follow_up_date else None,
                        remarks=call_remark if call_remark else None
                    )

                if update_data:
                    update_data = normalize_call_dates(update_data)
                    supabase.table('lead_master').update(update_data).eq('uid', uid).execute()

                    # Keep ps_followup_master.final_status in sync if final_status is updated and PS followup exists
                    if 'final_status' in update_data and lead_data.get('ps_name'):
                        ps_result = supabase.table('ps_followup_master').select('lead_uid').eq('lead_uid', uid).execute()
                        if ps_result.data:
                            supabase.table('ps_followup_master').update({'final_status': update_data['final_status']}).eq('lead_uid', uid).execute()

                    # Log lead update
                    auth_manager.log_audit_event(
                        user_id=session.get('user_id'),
                        user_type=session.get('user_type'),
                        action='LEAD_UPDATED',
                        resource='lead_master',
                        resource_id=uid,
                        details={'updated_fields': list(update_data.keys())}
                    )

                    flash('Lead updated successfully', 'success')
                else:
                    flash('No changes to update', 'info')
                
                # Redirect based on return_tab parameter
                if return_tab:
                    if return_tab in ['untouched-leads', 'called-leads', 'follow-up-leads']:
                        return redirect(url_for('cre_dashboard', tab='fresh-leads', sub_tab=return_tab))
                    elif return_tab == 'followups':
                        return redirect(url_for('cre_dashboard', tab='followups'))
                    elif return_tab == 'pending':
                        return redirect(url_for('cre_dashboard', tab='pending'))
                    elif return_tab == 'ps-assigned':
                        return redirect(url_for('cre_dashboard', tab='ps-assigned'))
                    elif return_tab == 'won-leads':
                        return redirect(url_for('cre_dashboard', tab='won-leads'))
                    else:
                        return redirect(url_for('cre_dashboard'))
                else:
                    return redirect(url_for('cre_dashboard'))
            except Exception as e:
                flash(f'Error updating lead: {str(e)}', 'error')

        # Fetch PS call summary from ps_followup_master
        ps_call_summary = {}
        if lead_data.get('ps_name'):
            ps_result = supabase.table('ps_followup_master').select('*').eq('lead_uid', uid).execute()
            if ps_result.data:
                ps_followup = ps_result.data[0]
                call_order = ['first', 'second', 'third']
                for call in call_order:
                    date_key = f"{call}_call_date"
                    remark_key = f"{call}_call_remark"
                    ps_call_summary[call] = {
                        "date": ps_followup.get(date_key),
                        "remark": ps_followup.get(remark_key)
                    }

        return render_template('update_lead.html',
                               lead=lead_data,
                               ps_users=ps_users,
                               rizta_models=rizta_models,
                               x450_models=x450_models,
                               branches=branches,
                               lead_statuses=lead_statuses,
                               next_call=next_call,
                               completed_calls=completed_calls,
                               today=date.today(),
                               ps_call_summary=ps_call_summary)

    except Exception as e:
        flash(f'Error loading lead: {str(e)}', 'error')
        return redirect(url_for('cre_dashboard'))

@app.route('/ps_dashboard')
@require_ps
def ps_dashboard():
    start_time = time.time()
    ps_name = session.get('ps_name')
    
    # Debug session data
    print(f"[DEBUG] PS Dashboard - Session data: {dict(session)}")
    print(f"[DEBUG] PS Dashboard - ps_name from session: {ps_name}")
    
    # Fallback to username if ps_name is not available
    if not ps_name:
        ps_name = session.get('username')
        print(f"[DEBUG] PS Dashboard - Using username as fallback: {ps_name}")
    
    # If still no name, try to get from database
    if not ps_name:
        user_id = session.get('user_id')
        if user_id:
            try:
                ps_user = supabase.table('ps_users').select('name,username').eq('id', user_id).execute()
                if ps_user.data:
                    ps_name = ps_user.data[0].get('name') or ps_user.data[0].get('username')
                    print(f"[DEBUG] PS Dashboard - Retrieved from DB: {ps_name}")
            except Exception as e:
                print(f"[DEBUG] PS Dashboard - Error fetching PS data: {e}")
    
    if not ps_name:
        flash('Error: Could not determine PS name. Please log in again.', 'error')
        return redirect(url_for('logout'))
    
    today = datetime.now().date()
    today_str = today.isoformat()

    # --- AUTO-INCREMENT LOGIC FOR UNATTENDED LEADS ---
    # 1. ps_followup_master
    assigned_leads = safe_get_data('ps_followup_master', {'ps_name': ps_name})
    for lead in assigned_leads:
        follow_up_date = lead.get('follow_up_date')
        final_status = lead.get('final_status')
        if follow_up_date and str(follow_up_date) < today_str and final_status not in ['Won', 'Lost']:
            supabase.table('ps_followup_master').update({'follow_up_date': today_str}).eq('lead_uid', lead.get('lead_uid')).execute()



    # 3. activity_leads
    event_leads = safe_get_data('activity_leads', {'ps_name': ps_name})
    for lead in event_leads:
        ps_followup_date_ts = lead.get('ps_followup_date_ts')
        final_status = lead.get('final_status')
        if ps_followup_date_ts and str(ps_followup_date_ts)[:10] < today_str and final_status not in ['Won', 'Lost']:
            if not lead.get('ps_first_call_date'):
                supabase.table('activity_leads').update({'ps_followup_date_ts': today_str}).eq('activity_uid', lead.get('activity_uid')).execute()

    # Get filter parameters from query string
    filter_type = request.args.get('filter_type', 'all')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    status_filter = request.args.get('status_filter', '')
    status = request.args.get('status', 'lost')  # <-- Add this line
    tab = request.args.get('tab', 'fresh-leads')  # <-- Add tab parameter
    category_filter = request.args.get('category_filter', '')  # <-- Add category filter parameter

    try:
        t0 = time.time()
        # Get all assigned leads for this PS
        print(f"[DEBUG] PS Dashboard - Fetching leads for PS: {ps_name}")
        assigned_leads = safe_get_data('ps_followup_master', {'ps_name': ps_name})
        print(f"[DEBUG] PS Dashboard - Assigned leads count: {len(assigned_leads) if assigned_leads else 0}")

        # Apply date filtering to assigned leads using ps_assigned_at for ps_followup_master table
        filtered_leads = filter_leads_by_date(assigned_leads, filter_type, 'ps_assigned_at')
        print(f"[DEBUG] PS Dashboard - Filtered leads count: {len(filtered_leads) if filtered_leads else 0}")

        # Initialize lists for different lead categories
        todays_followups_regular = []

        todays_followups_event = []  # Separate list for event followups
        fresh_leads = []
        pending_leads = []
        attended_leads = []
        won_leads = []
        lost_leads = []
        event_leads = []  # Separate list for event leads

        # Define statuses that should be excluded from Today's Follow-up and Pending Leads
        excluded_statuses = ['Lost to Codealer', 'Lost to Competition', 'Dropped', 'Booked', 'Retailed']



        print(f"[PERF] ps_dashboard: data fetching took {time.time() - t0:.3f} seconds")
        print(f"[DEBUG] Total assigned leads fetched: {len(assigned_leads)}")
        print(f"[DEBUG] Total filtered leads: {len(filtered_leads)}")
        print(f"[DEBUG] PS Name: {ps_name}")
        print(f"[DEBUG] Filter type: {filter_type}")

        t1 = time.time()
        # --- Process Regular Assigned Leads ---
        for lead in filtered_leads:
            lead_dict = dict(lead)  # Make a copy to avoid mutating the original
            lead_dict['lead_uid'] = lead.get('lead_uid')  # Ensure lead_uid for template compatibility

            final_status = lead.get('final_status')
            lead_status = lead.get('lead_status')

            # DEBUG: Print lead details for troubleshooting
            print(f"[DEBUG] Processing lead: {lead.get('lead_uid')} | final_status: {final_status} | lead_status: {lead_status} | first_call_date: {lead.get('first_call_date')} | ps_name: {lead.get('ps_name')}")

            # Categorize leads based on final_status and first_call_date
            if final_status == 'Won':
                won_leads.append(lead_dict)
            elif final_status == 'Lost':
                lost_leads.append(lead_dict)
            elif final_status == 'Pending':
                # Check if lead has been called (has first_call_date)
                if lead.get('first_call_date'):
                    # Lead has been called, goes to pending_leads
                    print(f"[DEBUG] Found lead with final_status == 'Pending' and first_call_date: {lead.get('lead_uid')} | first_call_date: {lead.get('first_call_date')} | lead_status: {lead_status}")
                    if not lead_status or lead_status not in excluded_statuses:
                        # Set default lead_category if missing
                        if 'lead_category' not in lead_dict or not lead_dict['lead_category']:
                            lead_dict['lead_category'] = 'Not Set'
                        pending_leads.append(lead_dict)
                        print(f"[DEBUG] Added to pending_leads (Pending status with first_call_date): {lead.get('lead_uid')}")
                    else:
                        print(f"[DEBUG] Lead {lead.get('lead_uid')} excluded from pending_leads due to lead_status: {lead_status}")
                else:
                    # Lead hasn't been called yet, goes to fresh_leads
                    print(f"[DEBUG] Found lead with final_status == 'Pending' but no first_call_date: {lead.get('lead_uid')} | lead_status: {lead_status}")
                    if not lead_status or lead_status not in excluded_statuses:
                        # Set default lead_category if missing
                        if 'lead_category' not in lead_dict or not lead_dict['lead_category']:
                            lead_dict['lead_category'] = 'Not Set'
                        fresh_leads.append(lead_dict)
                        print(f"[DEBUG] Added to fresh_leads (Pending status without first_call_date): {lead.get('lead_uid')}")
                    else:
                        print(f"[DEBUG] Lead {lead.get('lead_uid')} excluded from fresh_leads due to lead_status: {lead_status}")
            elif not final_status:  # Include leads with no final_status
                if lead.get('first_call_date'):
                    # Lead has been called, goes to pending_leads
                    print(f"[DEBUG] Found lead with no final_status but has first_call_date: {lead.get('lead_uid')} | first_call_date: {lead.get('first_call_date')} | lead_status: {lead_status}")
                    if not lead_status or lead_status not in excluded_statuses:
                        # Set default lead_category if missing
                        if 'lead_category' not in lead_dict or not lead_dict['lead_category']:
                            lead_dict['lead_category'] = 'Not Set'
                        pending_leads.append(lead_dict)
                        print(f"[DEBUG] Added to pending_leads (no final_status with first_call_date): {lead.get('lead_uid')}")
                    else:
                        print(f"[DEBUG] Lead {lead.get('lead_uid')} excluded from pending_leads due to lead_status: {lead_status}")
                else:
                    # Lead hasn't been called yet, goes to fresh_leads
                    print(f"[DEBUG] Found lead with no final_status and no first_call_date: {lead.get('lead_uid')} | lead_status: {lead_status}")
                    if not lead_status or lead_status not in excluded_statuses:
                        # Set default lead_category if missing
                        if 'lead_category' not in lead_dict or not lead_dict['lead_category']:
                            lead_dict['lead_category'] = 'Not Set'
                        fresh_leads.append(lead_dict)
                        print(f"[DEBUG] Added to fresh_leads (no final_status without first_call_date): {lead.get('lead_uid')}")
                    else:
                        print(f"[DEBUG] Lead {lead.get('lead_uid')} excluded from fresh_leads due to lead_status: {lead_status}")

            # Add to today's followups if applicable (exclude Won/Lost and specific statuses)
            follow_up_date = lead.get('follow_up_date')
            if (follow_up_date and
                str(follow_up_date).startswith(today_str) and
                final_status not in ['Won', 'Lost'] and
                (not lead_status or lead_status not in excluded_statuses)):
                todays_followups_regular.append(lead_dict)



        print(f"[PERF] ps_dashboard: regular leads processing took {time.time() - t1:.3f} seconds")
        print(f"[DEBUG] Fresh leads count: {len(fresh_leads)}")
        print(f"[DEBUG] Attended leads count: {len(attended_leads)}")
        print(f"[DEBUG] Won leads count: {len(won_leads)}")
        print(f"[DEBUG] Lost leads count: {len(lost_leads)}")
        print(f"[DEBUG] Today's followups (regular) count: {len(todays_followups_regular)}")
        
        # Debug: Check lead_category values in attended_leads
        print(f"[DEBUG] Attended leads with lead_category:")
        for lead in attended_leads:
            category = lead.get('lead_category')
            source = lead.get('source', 'Unknown')
            print(f"  - {lead.get('lead_uid')}: {category} (Source: {source})")
            
        # Debug: Check raw lead_category values from database
        print(f"[DEBUG] Raw lead_category values from database:")
        for lead in assigned_leads:
            category = lead.get('lead_category')
            print(f"  - {lead.get('lead_uid')}: {category} (Type: {type(category)})")





        t3 = time.time()
        # Fetch event leads for this PS
        fetched_event_leads = safe_get_data('activity_leads', {'ps_name': ps_name})
        
        # Fetch walk-in leads for this PS
        print(f"[DEBUG] PS Dashboard - Fetching walk-in leads for PS: {ps_name}")
        walkin_leads = safe_get_data('walkin_table', {'ps_assigned': ps_name})
        print(f"[DEBUG] PS Dashboard - Walk-in leads count: {len(walkin_leads) if walkin_leads else 0}")
        if walkin_leads:
            print(f"[DEBUG] Walk-in leads details:")
            for lead in walkin_leads:
                print(f"  - ID: {lead.get('id')}, UID: {lead.get('uid')}, Status: {lead.get('status')}, First Call: {lead.get('first_call_date')}")



        # --- Process Event Leads ---
        print(f"[DEBUG] Processing {len(fetched_event_leads)} event leads")
        for lead in fetched_event_leads:
            lead_dict = dict(lead)  # Make a copy to avoid mutating the original
            lead_dict['lead_uid'] = lead.get('activity_uid') or lead.get('uid')
            lead_dict['customer_mobile_number'] = lead.get('customer_phone_number')  # For template compatibility
            
            # For event leads, use 'final_status' instead of 'ps_final_status'
            final_status = lead.get('final_status')
            lead_status = lead.get('lead_status')
            ps_first_call_date = lead.get('ps_first_call_date')
            
            print(f"[DEBUG] Event lead: {lead_dict['lead_uid']} | final_status: {final_status} | lead_status: {lead_status} | ps_first_call_date: {ps_first_call_date}")

            # Always add event leads to event_leads list first
            if 'lead_category' not in lead_dict or not lead_dict['lead_category']:
                lead_dict['lead_category'] = 'Not Set'
            event_leads.append(lead_dict)
            print(f"[DEBUG] Event lead {lead_dict['lead_uid']} added to event_leads list")
            
            # Also categorize event leads for other tabs based on final_status
            if final_status == 'Won':
                won_leads.append(lead_dict)
                print(f"[DEBUG] Event lead {lead_dict['lead_uid']} also added to won_leads")
            elif final_status == 'Lost':
                lost_leads.append(lead_dict)
                print(f"[DEBUG] Event lead {lead_dict['lead_uid']} also added to lost_leads")
            elif final_status == 'Pending':
                # Check if event lead has been called (has ps_first_call_date)
                if ps_first_call_date:
                    # Event lead has been called, goes to pending_leads
                    print(f"[DEBUG] Event lead with final_status == 'Pending' and ps_first_call_date: {lead_dict['lead_uid']} | ps_first_call_date: {ps_first_call_date} | lead_status: {lead_status}")
                    if not lead_status or lead_status not in excluded_statuses:
                        pending_leads.append(lead_dict)
                        print(f"[DEBUG] Event lead {lead_dict['lead_uid']} also added to pending_leads (Pending status with ps_first_call_date)")
                    else:
                        print(f"[DEBUG] Event lead {lead_dict['lead_uid']} excluded from pending_leads due to lead_status: {lead_status}")
                else:
                    # Event lead hasn't been called yet, goes to fresh_leads
                    print(f"[DEBUG] Event lead with final_status == 'Pending' but no ps_first_call_date: {lead_dict['lead_uid']} | lead_status: {lead_status}")
                    if not lead_status or lead_status not in excluded_statuses:
                        fresh_leads.append(lead_dict)
                        print(f"[DEBUG] Event lead {lead_dict['lead_uid']} also added to fresh_leads (Pending status without ps_first_call_date)")
                    else:
                        print(f"[DEBUG] Event lead {lead_dict['lead_uid']} excluded from fresh_leads due to lead_status: {lead_status}")
            elif not final_status:  # Include leads with no final_status
                if ps_first_call_date:
                    # Event lead has been called, goes to pending_leads
                    print(f"[DEBUG] Event lead with no final_status but has ps_first_call_date: {lead_dict['lead_uid']} | ps_first_call_date: {ps_first_call_date} | lead_status: {lead_status}")
                    if not lead_status or lead_status not in excluded_statuses:
                        pending_leads.append(lead_dict)
                        print(f"[DEBUG] Event lead {lead_dict['lead_uid']} also added to pending_leads (no final_status with ps_first_call_date)")
                    else:
                        print(f"[DEBUG] Event lead {lead_dict['lead_uid']} excluded from pending_leads due to lead_status: {lead_status}")
                else:
                    # Event lead hasn't been called yet, goes to fresh_leads
                    print(f"[DEBUG] Event lead with no final_status and no ps_first_call_date: {lead_dict['lead_uid']} | lead_status: {lead_status}")
                    if not lead_status or lead_status not in excluded_statuses:
                        fresh_leads.append(lead_dict)
                        print(f"[DEBUG] Event lead {lead_dict['lead_uid']} also added to fresh_leads (no final_status without ps_first_call_date)")
                    else:
                        print(f"[DEBUG] Event lead {lead_dict['lead_uid']} excluded from fresh_leads due to lead_status: {lead_status}")

            # Add event leads with today's ps_followup_date_ts to today's followups
            ps_followup_date_ts = lead.get('ps_followup_date_ts')
            if (ps_followup_date_ts and
                str(ps_followup_date_ts)[:10] == today_str and
                final_status not in ['Won', 'Lost'] and
                (not lead_status or lead_status not in excluded_statuses)):
                # Set follow_up_date for template compatibility
                lead_dict['follow_up_date'] = str(ps_followup_date_ts)
                todays_followups_event.append(lead_dict)
                print(f"[DEBUG] Event lead {lead_dict['lead_uid']} added to today's followups")

        print(f"[PERF] ps_dashboard: event leads processing took {time.time() - t3:.3f} seconds")

        t3_5 = time.time()
        # --- Process Walk-in Leads ---
        print(f"[DEBUG] Processing {len(walkin_leads)} walk-in leads")
        todays_followups_walkin = []
        
        for lead in walkin_leads:
            lead_dict = dict(lead)  # Make a copy to avoid mutating the original
            
            # Ensure UID is properly set for walkin leads
            if lead.get('uid'):
                lead_dict['lead_uid'] = lead.get('uid')
            else:
                # Generate UID if not present (for older records)
                lead_dict['lead_uid'] = f"W{lead.get('id')}"
            
            # Also set the uid field for consistency
            lead_dict['uid'] = lead_dict['lead_uid']
            print(f"[DEBUG] Walkin lead UID: {lead_dict['lead_uid']} for ID: {lead.get('id')}")
            lead_dict['customer_mobile_number'] = lead.get('mobile_number', '')  # For template compatibility
            lead_dict['customer_name'] = lead.get('customer_name', '')  # Ensure customer_name is set
            lead_dict['source'] = 'Walk-in'  # Set source for consistency
            lead_dict['is_walkin'] = True  # Flag to identify walk-in leads
            lead_dict['walkin_id'] = lead.get('id')  # Store the walk-in ID for URL generation
            lead_dict['cre_name'] = ''  # Set empty cre_name for walk-in leads
            lead_dict['created_at'] = lead.get('created_at', '')  # Add created_at for template compatibility
            lead_dict['lead_category'] = lead.get('lead_category', 'Not Set')  # Add lead_category for template compatibility
            
            final_status = lead.get('status')
            lead_status = lead.get('lead_status')  # Add lead_status for walk-in leads
            followup_no = lead.get('followup_no', 1)
            next_followup_date = lead.get('next_followup_date')
            
            print(f"[DEBUG] Walk-in lead: {lead_dict['lead_uid']} | status: {final_status} | lead_status: {lead_status} | followup_no: {followup_no} | next_followup_date: {next_followup_date}")

            # Categorize walk-in leads based on status
            if final_status == 'Won':
                # Add won timestamp for filtering
                lead_dict['won_timestamp'] = lead.get('updated_at') or datetime.now().isoformat()
                won_leads.append(lead_dict)
                print(f"[DEBUG] Walk-in lead {lead_dict['lead_uid']} added to won_leads")
            elif final_status == 'Lost':
                # Add lost timestamp for filtering
                lead_dict['lost_timestamp'] = lead.get('updated_at') or datetime.now().isoformat()
                lost_leads.append(lead_dict)
                print(f"[DEBUG] Walk-in lead {lead_dict['lead_uid']} added to lost_leads")
            elif final_status == 'Pending' or not final_status:
                # Add to today's followups if next_followup_date is today
                if next_followup_date and str(next_followup_date)[:10] == today_str:
                    lead_dict['follow_up_date'] = str(next_followup_date)
                    todays_followups_walkin.append(lead_dict)
                    print(f"[DEBUG] Walk-in lead {lead_dict['lead_uid']} added to today's followups")
                
                # Add to pending leads if no first call has been made yet
                first_call_date = lead.get('first_call_date')
                # Get lead_status for walk-in leads (use status field or default to None)
                lead_status = lead.get('status') or lead.get('lead_status')
                print(f"[DEBUG] Walk-in lead {lead_dict['lead_uid']} - first_call_date: {first_call_date}, lead_status: {lead_status}")
                
                # For walk-in leads, check if any call dates exist
                has_any_call = any(lead.get(f'{i}_call_date') for i in range(1, 8))
                print(f"[DEBUG] Walk-in lead {lead_dict['lead_uid']} - has_any_call: {has_any_call}")
                
                if not has_any_call and (not lead_status or lead_status not in excluded_statuses):
                    pending_leads.append(lead_dict)
                    print(f"[DEBUG] Walk-in lead {lead_dict['lead_uid']} added to pending_leads")
                else:
                    # Add to attended leads if any call has been made
                    attended_leads.append(lead_dict)
                    print(f"[DEBUG] Walk-in lead {lead_dict['lead_uid']} added to attended_leads")

        # Update the walkin_leads list with processed data for template
        processed_walkin_leads = []
        for lead in walkin_leads:
            lead_dict = dict(lead)  # Make a copy to avoid mutating the original
            
            # Ensure UID is properly set for walkin leads
            if lead.get('uid'):
                lead_dict['lead_uid'] = lead.get('uid')
            else:
                # Generate UID if not present (for older records)
                lead_dict['lead_uid'] = f"W{lead.get('id')}"
            
            # Also set the uid field for consistency
            lead_dict['uid'] = lead_dict['lead_uid']
            lead_dict['customer_mobile_number'] = lead.get('mobile_number', '')
            lead_dict['customer_name'] = lead.get('customer_name', '')
            lead_dict['source'] = 'Walk-in'
            lead_dict['is_walkin'] = True
            lead_dict['walkin_id'] = lead.get('id')
            lead_dict['cre_name'] = ''
            lead_dict['created_at'] = lead.get('created_at', '')
            lead_dict['lead_category'] = lead.get('lead_category', 'Not Set')
            
            processed_walkin_leads.append(lead_dict)
        
        # Replace the original walkin_leads with processed data
        walkin_leads = processed_walkin_leads
        
        print(f"[PERF] ps_dashboard: walk-in leads processing took {time.time() - t3_5:.3f} seconds")

        t4 = time.time()
        # Apply date filtering to lost leads using lost_timestamp
        def filter_lost_by_timestamp(leads_list):
            if filter_type == 'all':
                return leads_list
            elif filter_type == 'today':
                return [lead for lead in leads_list if lead.get('lost_timestamp') and str(lead.get('lost_timestamp')).startswith(today_str)]
            elif filter_type == 'range' and start_date and end_date:
                filtered = []
                for lead in leads_list:
                    ts = lead.get('lost_timestamp')
                    if ts:
                        try:
                            date_val = ts[:10]
                            if start_date <= date_val <= end_date:
                                filtered.append(lead)
                        except Exception:
                            continue
                return filtered
            else:
                return leads_list

        # Apply timestamp filtering to lost leads
        lost_leads = filter_lost_by_timestamp(lost_leads)

        print(f"[PERF] ps_dashboard: lost leads timestamp filtering took {time.time() - t4:.3f} seconds")

        t5 = time.time()
        # Merge today's followup lists
        todays_followups = todays_followups_regular + todays_followups_event + todays_followups_walkin

        print(f"[PERF] ps_dashboard: final processing took {time.time() - t5:.3f} seconds")

        t6 = time.time()
        # Render template with all lead categories and the new status variable
        result = render_template('ps_dashboard.html',
                               assigned_leads=assigned_leads,
                               fresh_leads=fresh_leads,
                               pending_leads=pending_leads,
                               todays_followups=todays_followups,
                               attended_leads=attended_leads,
                               won_leads=won_leads,
                               lost_leads=lost_leads,
                               event_leads=event_leads,
                               walkin_leads=walkin_leads,
                               filter_type=filter_type,
                               start_date=start_date,
                               end_date=end_date,
                               status_filter=status_filter,
                               status=status)  # <-- Add status here

        print(f"[PERF] ps_dashboard: render_template took {time.time() - t6:.3f} seconds")
        print(f"[DEBUG] FINAL COUNTS - Fresh: {len(fresh_leads)}, Pending: {len(pending_leads)}, Attended: {len(attended_leads)}, Won: {len(won_leads)}, Lost: {len(lost_leads)}, Event: {len(event_leads)}")
        print(f"[DEBUG] Fresh leads being sent to template: {[lead.get('lead_uid') for lead in fresh_leads]}")
        print(f"[DEBUG] Pending leads being sent to template: {[lead.get('lead_uid') for lead in pending_leads]}")
        print(f"[DEBUG] Won leads being sent to template: {[lead.get('lead_uid') for lead in won_leads]}")
        print(f"[DEBUG] Lost leads being sent to template: {[lead.get('lead_uid') for lead in lost_leads]}")
        print(f"[DEBUG] Event leads being sent to template: {[lead.get('lead_uid') for lead in event_leads]}")
        print(f"[DEBUG] Walk-in leads count: {len(walkin_leads) if walkin_leads else 0}")
        print(f"[DEBUG] Today's followups count: {len(todays_followups)}")
        print(f"[PERF] ps_dashboard TOTAL took {time.time() - start_time:.3f} seconds")
        
        # Add session data to template context for debugging
        template_data = {
            'assigned_leads': assigned_leads,
            'fresh_leads': fresh_leads,
            'pending_leads': pending_leads,
            'todays_followups': todays_followups,
            'attended_leads': attended_leads,
            'won_leads': won_leads,
            'lost_leads': lost_leads,
            'event_leads': event_leads,
            'walkin_leads': walkin_leads,
            'filter_type': filter_type,
            'start_date': start_date,
            'end_date': end_date,
            'status_filter': status_filter,
            'status': status,
            'tab': tab,  # Add tab parameter
            'category_filter': category_filter,  # Add category filter parameter
            'ps_name': ps_name,  # Pass ps_name directly to template
            'debug_info': {
                'session_data': dict(session),
                'ps_name': ps_name,
                'total_leads': len(assigned_leads) if assigned_leads else 0
            }
        }
        
        result = render_template('ps_dashboard.html', **template_data)
        return result

    except Exception as e:
        print(f"[PERF] ps_dashboard failed after {time.time() - start_time:.3f} seconds: {e}")
        flash(f'Error loading dashboard: {str(e)}', 'error')
        return render_template('ps_dashboard.html',
                             assigned_leads=[],
                             fresh_leads=[],
                             pending_leads=[],
                             todays_followups=[],
                             attended_leads=[],
                             won_leads=[],
                             lost_leads=[],
                             event_leads=[],
                             walkin_leads=[],
                             filter_type=filter_type,
                             start_date=start_date,
                             end_date=end_date,
                             status_filter=status_filter,
                             status=status,
                             ps_name=ps_name if 'ps_name' in locals() else 'Unknown',
                             debug_info={
                                 'session_data': dict(session),
                                 'ps_name': ps_name if 'ps_name' in locals() else 'Unknown',
                                 'error': str(e)
                             })



@app.route('/update_walkin_lead/<int:walkin_id>', methods=['GET', 'POST'])
@require_ps
def update_walkin_lead(walkin_id):
    return_tab = request.args.get('return_tab', 'walkin-leads')
    try:
        # Get the walk-in lead data
        walkin_result = supabase.table('walkin_table').select('*').eq('id', walkin_id).execute()
        if not walkin_result.data:
            flash('Walk-in lead not found', 'error')
            return redirect(url_for('ps_dashboard'))
        
        walkin_lead = walkin_result.data[0]
        ps_name = session.get('ps_name')
        
        # Check if this PS is assigned to this walk-in lead
        if walkin_lead.get('ps_assigned') != ps_name:
            flash('You are not authorized to update this walk-in lead', 'error')
            return redirect(url_for('ps_dashboard'))
        
        if request.method == 'POST':
            # Get form data
            submitted_followup_no = int(request.form.get('followup_no', 1))
            call_date = request.form.get('call_date')
            call_remark = request.form.get('call_remark')
            status = request.form.get('status', 'Pending')
            lead_category = request.form.get('lead_category', 'Warm')
            lead_status = request.form.get('lead_status', '')
            model_interested = request.form.get('model_interested', '')
            test_drive_done = request.form.get('test_drive_done', '')
            next_followup_date = request.form.get('next_followup_date')
            
            # Validate that the submitted followup number is the correct next call
            next_call, completed_calls = get_next_ps_call_info(walkin_lead)
            call_order = ['first', 'second', 'third', 'fourth', 'fifth', 'sixth', 'seventh']
            expected_followup_no = call_order.index(next_call) + 1
            
            if submitted_followup_no != expected_followup_no:
                flash(f'Invalid follow-up number. Only call #{expected_followup_no} can be recorded next.', 'error')
                redirect_url = url_for('update_walkin_lead', walkin_id=walkin_id, return_tab=return_tab)
                return redirect(redirect_url)
            
            # Validate next follow-up date is required unless status is Won/Lost
            if status not in ['Won', 'Lost'] and not next_followup_date:
                flash('Next follow-up date is required for Pending status.', 'error')
                redirect_url = url_for('update_walkin_lead', walkin_id=walkin_id, return_tab=return_tab)
                return redirect(redirect_url)
            
            # Validate test_drive_done is required
            if not test_drive_done:
                flash('Test Drive Done field is required.', 'error')
                redirect_url = url_for('update_walkin_lead', walkin_id=walkin_id, return_tab=return_tab)
                return redirect(redirect_url)
            
            followup_no = submitted_followup_no
            
            # Update the specific call field based on followup number
            update_data = {
                'status': status,
                'lead_category': lead_category,
                'lead_status': lead_status,
                'model_interested': model_interested,
                'test_drive_done': test_drive_done,
                'followup_no': followup_no,
                'updated_at': datetime.now().isoformat()
            }
            
            # Set the specific call date and remark fields
            if followup_no == 1:
                update_data['first_call_date'] = get_ist_timestamp() if call_date else None
                update_data['first_call_remark'] = call_remark
            elif followup_no == 2:
                update_data['second_call_date'] = get_ist_timestamp() if call_date else None
                update_data['second_call_remark'] = call_remark
            elif followup_no == 3:
                update_data['third_call_date'] = get_ist_timestamp() if call_date else None
                update_data['third_call_remark'] = call_remark
            elif followup_no == 4:
                update_data['fourth_call_date'] = get_ist_timestamp() if call_date else None
                update_data['fourth_call_remark'] = call_remark
            elif followup_no == 5:
                update_data['fifth_call_date'] = get_ist_timestamp() if call_date else None
                update_data['fifth_call_remark'] = call_remark
            elif followup_no == 6:
                update_data['sixth_call_date'] = get_ist_timestamp() if call_date else None
                update_data['sixth_call_remark'] = call_remark
            elif followup_no == 7:
                update_data['seventh_call_date'] = get_ist_timestamp() if call_date else None
                update_data['seventh_call_remark'] = call_remark
            
            # Set next followup date if provided
            if next_followup_date:
                update_data['next_followup_date'] = next_followup_date
            
            # Update the walk-in lead
            supabase.table('walkin_table').update(update_data).eq('id', walkin_id).execute()
            
            # Track the PS call attempt in ps_call_attempt_history
            if lead_status:
                call_was_recorded = bool(call_date and call_remark)
                track_ps_call_attempt(
                    uid=walkin_lead['uid'],
                    ps_name=ps_name,
                    call_no=next_call,
                    lead_status=lead_status,
                    call_was_recorded=call_was_recorded,
                    follow_up_date=next_followup_date if next_followup_date else None,
                    remarks=call_remark if call_remark else None
                )
            
            # Sync to alltest_drive table if test_drive_done is set
            if test_drive_done in ['Yes', 'No', True, False]:
                # Get the updated walkin lead data for syncing
                updated_walkin_result = supabase.table('walkin_table').select('*').eq('id', walkin_id).execute()
                if updated_walkin_result.data:
                    updated_walkin_data = updated_walkin_result.data[0]
                    sync_test_drive_to_alltest_drive('walkin_table', walkin_id, updated_walkin_data)
            
            flash(f'Walk-in lead updated successfully! Follow-up {followup_no} recorded.', 'success')
            redirect_url = url_for('ps_dashboard', tab=return_tab)
            return redirect(redirect_url)
        
        # For GET request, render the update form
        # Determine next call number and completed calls
        next_call, completed_calls = get_next_ps_call_info(walkin_lead)
        
        # Create a list of available call numbers for the dropdown
        # Only allow the next call number to be selected
        available_calls = []
        call_order = ['first', 'second', 'third', 'fourth', 'fifth', 'sixth', 'seventh']
        
        # Find the index of the next call
        next_call_index = call_order.index(next_call)
        
        # Only add the next call as an option
        if next_call_index < len(call_order):
            call_number = next_call_index + 1
            available_calls.append({
                'number': call_number,
                'name': next_call,
                'display': f"{call_number}{'st' if call_number == 1 else 'nd' if call_number == 2 else 'rd' if call_number == 3 else 'th'} Call"
            })
        
        models = [
            "450X", "450S", "Rizta S Mono (2.9 kWh)", "Rizta S Super Matte (2.9 kWh)",
            "Rizta Z Mono (2.9 kWh)", "Rizta Z Duo (2.9 kWh)", "Rizta Z Super Matte (2.9 kWh)",
            "Rizta Z Mono (3.7 kWh)", "Rizta Z Duo (3.7 kWh)", "Rizta Z Super Matte (3.7 kWh)",
            "450 X (2.9 kWh)", "450 X (3.7 kWh)", "450 X (2.9 kWh) Pro Pack", "450 X (3.7 kWh) Pro Pack",
            "450 Apex STD"
        ]
        return render_template(
            'update_walkin_lead.html',
            walkin_lead=walkin_lead,
            next_call=next_call,
            completed_calls=completed_calls,
            available_calls=available_calls,
            models=models
        )
        
    except Exception as e:
        flash(f'Error updating walk-in lead: {str(e)}', 'error')
        redirect_url = url_for('ps_dashboard', tab=return_tab)
        return redirect(redirect_url)


@app.route('/update_ps_lead/<uid>', methods=['GET', 'POST'])
@require_ps
def update_ps_lead(uid):
    return_tab = request.args.get('return_tab', 'fresh-leads')
    status_filter = request.args.get('status_filter', '')
    category_filter = request.args.get('category_filter', '')
    try:
        # Try to get PS followup data (regular leads)
        ps_result = supabase.table('ps_followup_master').select('*').eq('lead_uid', uid).execute()
        if ps_result.data:
            ps_data = ps_result.data[0]
            # Fetch CRE call summary from lead_master
            lead_result = supabase.table('lead_master').select('*').eq('uid', uid).execute()
            cre_call_summary = {}
            if lead_result.data:
                lead = lead_result.data[0]
                call_order = ['first', 'second', 'third', 'fourth', 'fifth', 'sixth', 'seventh']
                for call in call_order:
                    date_key = f"{call}_call_date"
                    remark_key = f"{call}_remark"
                    cre_call_summary[call] = {
                        "date": lead.get(date_key),
                        "remark": lead.get(remark_key)
                    }
            else:
                cre_call_summary = {}
            # Verify this lead belongs to the current PS
            if ps_data.get('ps_name') != session.get('ps_name'):
                flash('Access denied - This lead is not assigned to you', 'error')
                redirect_url = url_for('ps_dashboard', tab=return_tab)
                if status_filter:
                    redirect_url += f'&status_filter={status_filter}'
                return redirect(redirect_url)
            # Get next call info for PS (only 7 calls)
            next_call, completed_calls = get_next_ps_call_info(ps_data)
            if request.method == 'POST':
                update_data = {}
                lead_status = request.form.get('lead_status', '')
                follow_up_date = request.form.get('follow_up_date', '')
                call_remark = request.form.get('call_remark', '')
                lead_category = request.form.get('lead_category', '')
                model_interested = request.form.get('model_interested', '')
                test_drive_done = request.form.get('test_drive_done', '')
                
                # Always update lead_status from the form
                if lead_status:
                    update_data['lead_status'] = lead_status
                    
                    # Auto-set final_status to Won for Booked with another number
                    if lead_status == 'Booked with another number':
                        update_data['final_status'] = 'Won'
                        update_data['won_timestamp'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        
                if lead_category:
                    update_data['lead_category'] = lead_category
                if model_interested:
                    update_data['model_interested'] = model_interested
                # Test Drive Done is now required
                if not test_drive_done:
                    flash('Test Drive Done field is required', 'error')
                    return redirect(url_for('update_ps_lead', uid=uid, return_tab=return_tab))
                update_data['test_drive_done'] = test_drive_done
                if request.form.get('follow_up_date'):
                    update_data['follow_up_date'] = follow_up_date
                if request.form.get('final_status'):
                    final_status = request.form['final_status']
                    update_data['final_status'] = final_status
                    if final_status == 'Won':
                        update_data['won_timestamp'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    elif final_status == 'Lost':
                        update_data['lost_timestamp'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                # Handle call dates and remarks for the next available call
                skip_first_call_statuses = [
                    'Call not Connected',
                    'Busy on another call',
                    'RNR',
                    'Call me Back'
                ]
                if request.form.get('call_date') and lead_status not in skip_first_call_statuses:
                    update_data[f'{next_call}_call_date'] = get_ist_timestamp()
                if call_remark:
                    combined_remark = f"{lead_status}, {call_remark}"
                    update_data[f'{next_call}_call_remark'] = combined_remark
                    # Notification removed - no longer sending notifications between PS and CRE
                try:
                    # Track the PS call attempt before updating the lead
                    if lead_status:
                        call_was_recorded = bool(request.form.get('call_date') and call_remark)
                        track_ps_call_attempt(
                            uid=uid,
                            ps_name=session.get('ps_name'),
                            call_no=next_call,
                            lead_status=lead_status,
                            call_was_recorded=call_was_recorded,
                            follow_up_date=follow_up_date if follow_up_date else None,
                            remarks=call_remark if call_remark else None
                        )
                    if update_data:
                        supabase.table('ps_followup_master').update(update_data).eq('lead_uid', uid).execute()
                        
                        # Sync to alltest_drive table if test_drive_done is set
                        if test_drive_done in ['Yes', 'No', True, False]:
                            # Get the updated ps_followup data for syncing
                            updated_ps_result = supabase.table('ps_followup_master').select('*').eq('lead_uid', uid).execute()
                            if updated_ps_result.data:
                                updated_ps_data = updated_ps_result.data[0]
                                sync_test_drive_to_alltest_drive('ps_followup_master', uid, updated_ps_data)
                        
                        # Also update the main lead table final status
                        if request.form.get('final_status'):
                            final_status = request.form['final_status']
                            main_update_data = {'final_status': final_status}
                            if final_status == 'Won':
                                main_update_data['won_timestamp'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                            elif final_status == 'Lost':
                                main_update_data['lost_timestamp'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                            supabase.table('lead_master').update(main_update_data).eq('uid', uid).execute()
                        # Log PS lead update
                        auth_manager.log_audit_event(
                            user_id=session.get('user_id'),
                            user_type=session.get('user_type'),
                            action='PS_LEAD_UPDATED',
                            resource='ps_followup_master',
                            resource_id=uid,
                            details={'updated_fields': list(update_data.keys())}
                        )
                        flash('Lead updated successfully', 'success')
                    else:
                        flash('No changes to update', 'info')
                    redirect_url = url_for('ps_dashboard', tab=return_tab)
                    if status_filter:
                        redirect_url += f'&status_filter={status_filter}'
                    if category_filter:
                        redirect_url += f'&category_filter={category_filter}'
                    return redirect(redirect_url)
                except Exception as e:
                    flash(f'Error updating lead: {str(e)}', 'error')
            else:
                return render_template('update_ps_lead.html',
                                       lead=ps_data,
                                       next_call=next_call,
                                       completed_calls=completed_calls,
                                       today=date.today(),
                                       cre_call_summary=cre_call_summary)
        else:
            flash('Lead not found', 'error')
            redirect_url = url_for('ps_dashboard', tab=return_tab)
            if status_filter:
                redirect_url += f'&status_filter={status_filter}'
            return redirect(redirect_url)
    except Exception as e:
        flash(f'Error loading lead: {str(e)}', 'error')
        return redirect(url_for('ps_dashboard'))


@app.route('/logout')
@require_auth()
def logout():
    # Log logout
    auth_manager.log_audit_event(
        user_id=session.get('user_id'),
        user_type=session.get('user_type'),
        action='LOGOUT',
        details={'session_id': session.get('session_id')}
    )

    # Deactivate session
    if session.get('session_id'):
        auth_manager.deactivate_session(session.get('session_id'))

    session.clear()
    flash('You have been logged out successfully', 'info')
    return redirect(url_for('index'))

# =====================================================
# OPTIMIZED LEAD UPDATE ENDPOINTS
# =====================================================

@app.route('/update_lead_optimized/<uid>', methods=['GET', 'POST'])
@require_cre
def update_lead_optimized(uid):
    """
    Optimized lead update endpoint for CRE users
    """
    return_tab = request.args.get('return_tab', '')

    try:
        # Get lead data with optimized query
        lead_data_result = optimized_ops.get_lead_with_related_data(uid)

        if not lead_data_result or not lead_data_result['lead']:
            flash('Lead not found', 'error')
            return redirect(url_for('cre_dashboard'))

        lead_data = lead_data_result['lead']
        ps_followup = lead_data_result.get('ps_followup')

        # Verify this lead belongs to the current CRE
        if lead_data.get('cre_name') != session.get('cre_name'):
            flash('Access denied - This lead is not assigned to you', 'error')
            return redirect(url_for('cre_dashboard'))

        # Get next call info
        next_call, completed_calls = get_next_call_info(lead_data)

        # Get PS users for branch selection
        ps_users = safe_get_data('ps_users')

        # Model options and other data (same as original)
        rizta_models = [
            'Rizta S Mono (2.9 kWh)',
            'Rizta S Super Matte (2.9 kWh)',
            'Rizta Z Mono (2.9 kWh)',
            'Rizta Z Duo (2.9 kWh)',
            'Rizta Z Super Matte (2.9 kWh)',
            'Rizta Z Mono (3.7 kWh)',
            'Rizta Z Duo (3.7 kWh)',
            'Rizta Z Super Matte (3.7 kWh)'
        ]

        x450_models = [
            '450 X (2.9 kWh)',
            '450 X (3.7 kWh)',
            '450 X (2.9 kWh) Pro Pack',
            '450 X (3.7 kWh) Pro Pack',
            '450 Apex STD'
        ]

        branches = ['PORUR', 'NUNGAMBAKKAM', 'TIRUVOTTIYUR']

        lead_statuses = [
            'Busy on another Call', 'RNR', 'Call me Back', 'Interested',
            'Not Interested', 'Did Not Inquire', 'Lost to Competition',
            'Lost to Co Dealer', 'Call Disconnected', 'Wrong Number'
        ]

        if request.method == 'POST':
            # Prepare update data
            update_data = {}

            # Process form data (same logic as original)
            lead_status = request.form.get('lead_status', '')
            follow_up_date = request.form.get('follow_up_date', '')
            call_remark = request.form.get('call_remark', '')

            # Lock follow_up_date and set final_status for certain statuses
            lock_statuses = ['Booked', 'Retailed']
            lost_statuses = ['Not Interested', 'Lost to Codealer', 'Lost to Competition']

            if lead_status in lock_statuses:
                update_data['lead_status'] = lead_status
                update_data['follow_up_date'] = follow_up_date or lead_data.get('follow_up_date')
                update_data['final_status'] = 'Won' if lead_status in ['Booked', 'Retailed'] else lead_data.get('final_status')
            elif lead_status in lost_statuses:
                update_data['lead_status'] = lead_status
                update_data['follow_up_date'] = follow_up_date or lead_data.get('follow_up_date')
                update_data['final_status'] = 'Lost'
            else:
                if lead_status:
                    update_data['lead_status'] = lead_status

            # Process other form fields
            if request.form.get('customer_name'):
                update_data['customer_name'] = request.form['customer_name']

            if request.form.get('source'):
                update_data['source'] = request.form['source']

            if request.form.get('lead_category'):
                update_data['lead_category'] = request.form['lead_category']

            if request.form.get('model_interested'):
                update_data['model_interested'] = request.form['model_interested']

            if request.form.get('branch'):
                update_data['branch'] = request.form['branch']

            # Handle PS assignment
            ps_name = request.form.get('ps_name')
            if ps_name and ps_name != lead_data.get('ps_name'):
                update_data['ps_name'] = ps_name
                update_data['assigned'] = 'Yes'
                update_data['ps_assigned_at'] = datetime.now().isoformat()

                # Create PS followup record
                ps_user = next((ps for ps in ps_users if ps['name'] == ps_name), None)
                if ps_user:
                    create_or_update_ps_followup(lead_data, ps_name, ps_user['branch'])

                    # Send email notification (non-blocking)
                    try:
                        lead_data_for_email = {**lead_data, **update_data}
                        socketio.start_background_task(send_email_to_ps, ps_user['email'], ps_user['name'], lead_data_for_email, session.get('cre_name'))
                        flash(f'Lead assigned to {ps_name} and email notification sent', 'success')
                    except Exception as e:
                        print(f"Error sending email: {e}")
                        flash(f'Lead assigned to {ps_name} (email notification failed)', 'warning')

            if follow_up_date:
                update_data['follow_up_date'] = follow_up_date

            if request.form.get('final_status'):
                final_status = request.form['final_status']
                update_data['final_status'] = final_status
                if final_status == 'Won':
                    update_data['won_timestamp'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                elif final_status == 'Lost':
                    update_data['lost_timestamp'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            # Handle call dates and remarks
            if request.form.get('call_date') and call_remark:
                combined_remark = f"{lead_status}, {call_remark}"
                call_names = ['first', 'second', 'third', 'fourth', 'fifth', 'sixth', 'seventh']
                if next_call in call_names:
                    update_data[f'{next_call}_call_date'] = get_ist_timestamp()
                    update_data[f'{next_call}_remark'] = combined_remark
                    
                    # Set first_call_date when the first call is made
                    if next_call == 'first' and not lead_data.get('first_call_date'):
                        update_data['first_call_date'] = get_ist_timestamp()

            # Track call attempt (non-blocking)
            if lead_status:
                call_was_recorded = bool(request.form.get('call_date') and call_remark)
                try:
                    track_cre_call_attempt(
                        uid=uid,
                        cre_name=session.get('cre_name'),
                        call_no=next_call,
                        lead_status=lead_status,
                        call_was_recorded=call_was_recorded,
                        follow_up_date=follow_up_date if follow_up_date else None,
                        remarks=call_remark if call_remark else None
                    )
                except Exception as e:
                    print(f"Error tracking call attempt: {e}")

            # Use optimized update operation
            if update_data:
                update_data = normalize_call_dates(update_data)
                result = optimized_ops.update_lead_optimized(
                    uid=uid,
                    update_data=update_data,
                    user_type='cre',
                    user_name=session.get('cre_name')
                )

                if result['success']:
                    flash('Lead updated successfully', 'success')
                else:
                    flash(f'Error updating lead: {result.get("error", "Unknown error")}', 'error')
            else:
                flash('No changes to update', 'info')

            # Redirect based on return_tab parameter
            if return_tab:
                if return_tab in ['untouched-leads', 'called-leads', 'follow-up-leads']:
                    return redirect(url_for('cre_dashboard', tab='fresh-leads', sub_tab=return_tab))
                elif return_tab == 'followups':
                    return redirect(url_for('cre_dashboard', tab='followups'))
                elif return_tab == 'pending':
                    return redirect(url_for('cre_dashboard', tab='pending'))
                elif return_tab == 'ps-assigned':
                    return redirect(url_for('cre_dashboard', tab='ps-assigned'))
                elif return_tab == 'won-leads':
                    return redirect(url_for('cre_dashboard', tab='won-leads'))
                else:
                    return redirect(url_for('cre_dashboard'))
            else:
                return redirect(url_for('cre_dashboard'))

        # Fetch PS call summary from ps_followup_master
        ps_call_summary = {}
        if ps_followup:
            call_order = ['first', 'second', 'third']
            for call in call_order:
                date_key = f"{call}_call_date"
                remark_key = f"{call}_call_remark"
                ps_call_summary[call] = {
                    "date": ps_followup.get(date_key),
                    "remark": ps_followup.get(remark_key)
                }

        return render_template('update_lead.html',
                               lead=lead_data,
                               ps_users=ps_users,
                               rizta_models=rizta_models,
                               x450_models=x450_models,
                               branches=branches,
                               lead_statuses=lead_statuses,
                               next_call=next_call,
                               completed_calls=completed_calls,
                               today=date.today(),
                               ps_call_summary=ps_call_summary)

    except Exception as e:
        flash(f'Error loading lead: {str(e)}', 'error')
        return redirect(url_for('cre_dashboard'))

@app.route('/update_ps_lead_optimized/<uid>', methods=['GET', 'POST'])
@require_ps
def update_ps_lead_optimized(uid):
    """
    Optimized PS lead update endpoint
    """
    return_tab = request.args.get('return_tab', 'fresh-leads')
    status_filter = request.args.get('status_filter', '')
    category_filter = request.args.get('category_filter', '')

    try:
        # Get PS followup data with optimized query
        ps_result = optimized_ops.supabase.table('ps_followup_master').select('*').eq('lead_uid', uid).execute()

        if not ps_result.data:
            flash('Lead not found', 'error')
            redirect_url = url_for('ps_dashboard', tab=return_tab)
            if status_filter:
                redirect_url += f'&status_filter={status_filter}'
            return redirect(redirect_url)

        ps_data = ps_result.data[0]

        # Verify this lead belongs to the current PS
        if ps_data.get('ps_name') != session.get('ps_name'):
            flash('Access denied - This lead is not assigned to you', 'error')
            redirect_url = url_for('ps_dashboard', tab=return_tab)
            if status_filter:
                redirect_url += f'&status_filter={status_filter}'
            return redirect(redirect_url)

        # Get next call info for PS
        next_call, completed_calls = get_next_ps_call_info(ps_data)

        # Fetch CRE call summary from lead_master
        cre_call_summary = {}
        lead_result = optimized_ops.supabase.table('lead_master').select('*').eq('uid', uid).execute()
        if lead_result.data:
            lead = lead_result.data[0]
            call_order = ['first', 'second', 'third', 'fourth', 'fifth', 'sixth', 'seventh']
            for call in call_order:
                date_key = f"{call}_call_date"
                remark_key = f"{call}_remark"
                cre_call_summary[call] = {
                    "date": lead.get(date_key),
                    "remark": lead.get(remark_key)
                }

        if request.method == 'POST':
            # Prepare update data
            update_data = {}

            # Process form data
            lead_status = request.form.get('lead_status', '')
            follow_up_date = request.form.get('follow_up_date', '')
            call_remark = request.form.get('call_remark', '')
            lead_category = request.form.get('lead_category', '')
            model_interested = request.form.get('model_interested', '')

            # Always update lead_status from the form
            if lead_status:
                update_data['lead_status'] = lead_status
                
                # Auto-set final_status to Won for Booked with another number
                if lead_status == 'Booked with another number':
                    update_data['final_status'] = 'Won'
                    update_data['won_timestamp'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    
            if lead_category:
                update_data['lead_category'] = lead_category
            if model_interested:
                update_data['model_interested'] = model_interested
            if request.form.get('follow_up_date'):
                update_data['follow_up_date'] = follow_up_date
            if request.form.get('final_status'):
                final_status = request.form['final_status']
                update_data['final_status'] = final_status
                if final_status == 'Won':
                    update_data['won_timestamp'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                elif final_status == 'Lost':
                    update_data['lost_timestamp'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            # Handle call dates and remarks for the next available call
            skip_first_call_statuses = [
                'Call not Connected',
                'Busy on another call',
                'RNR',
                'Call me Back'
            ]

            if request.form.get('call_date') and lead_status not in skip_first_call_statuses:
                update_data[f'{next_call}_call_date'] = get_ist_timestamp()
            if call_remark:
                combined_remark = f"{lead_status}, {call_remark}"
                update_data[f'{next_call}_call_remark'] = combined_remark

            # Track call attempt (non-blocking)
            if lead_status:
                call_was_recorded = bool(request.form.get('call_date') and call_remark)
                try:
                    track_ps_call_attempt(
                        uid=uid,
                        ps_name=session.get('ps_name'),
                        call_no=next_call,
                        lead_status=lead_status,
                        call_was_recorded=call_was_recorded,
                        follow_up_date=follow_up_date if follow_up_date else None,
                        remarks=call_remark if call_remark else None
                    )
                except Exception as e:
                    print(f"Error tracking PS call attempt: {e}")

            # Use optimized update operation
            if update_data:
                result = optimized_ops.update_ps_lead_optimized(
                    uid=uid,
                    update_data=update_data,
                    ps_name=session.get('ps_name')
                )

                if result['success']:
                    flash('Lead updated successfully', 'success')
                else:
                    flash(f'Error updating lead: {result.get("error", "Unknown error")}', 'error')
            else:
                flash('No changes to update', 'info')

            # Redirect with filters
            redirect_url = url_for('ps_dashboard', tab=return_tab)
            if status_filter:
                redirect_url += f'&status_filter={status_filter}'
            if category_filter:
                redirect_url += f'&category_filter={category_filter}'
            return redirect(redirect_url)

        return render_template('update_ps_lead.html',
                               lead=ps_data,
                               next_call=next_call,
                               completed_calls=completed_calls,
                               today=date.today(),
                               cre_call_summary=cre_call_summary)

    except Exception as e:
        flash(f'Error loading lead: {str(e)}', 'error')
        return redirect(url_for('ps_dashboard'))

@app.route('/dashboard_leads_optimized')
@require_auth(['cre', 'ps'])
def dashboard_leads_optimized():
    """
    Optimized dashboard leads endpoint for AJAX loading
    """
    try:
        user_type = session.get('user_type')
        user_name = session.get('cre_name') if user_type == 'cre' else session.get('ps_name')

        # Get filters from request
        filters = {
            'final_status': request.args.get('final_status'),
            'lead_status': request.args.get('lead_status'),
            'page': int(request.args.get('page', 1)),
            'per_page': int(request.args.get('per_page', 50))
        }

        # Add date range filter if provided
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        if start_date and end_date:
            filters['date_range'] = (start_date, end_date)

        # Get optimized dashboard data
        result = optimized_ops.get_dashboard_leads_optimized(user_type, user_name, filters)

        if 'error' in result:
            return jsonify({'error': result['error']}), 500

        return jsonify({
            'success': True,
            'leads': result['leads'],
            'total_count': result['total_count'],
            'page': result['page'],
            'per_page': result['per_page']
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/batch_update_leads', methods=['POST'])
@require_auth(['cre', 'ps'])
def batch_update_leads():
    """
    Batch update multiple leads for better performance
    """
    try:
        updates = request.json.get('updates', [])

        if not updates:
            return jsonify({'error': 'No updates provided'}), 400

        # Add user information to each update
        user_type = session.get('user_type')
        user_name = session.get('cre_name') if user_type == 'cre' else session.get('ps_name')

        for update in updates:
            update['user_type'] = user_type
            update['user_name'] = user_name

        # Perform batch update
        result = optimized_ops.batch_update_leads(updates)

        return jsonify({
            'success': True,
            'successful': result['successful'],
            'failed': result['failed'],
            'errors': result['errors']
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/performance_metrics')
@require_admin
def performance_metrics():
    """
    Get performance metrics for monitoring
    """
    try:
        # Get basic metrics
        metrics = {
            'total_leads': len(safe_get_data('lead_master')),
            'total_ps_followups': len(safe_get_data('ps_followup_master')),
            'active_cre_users': len([u for u in safe_get_data('cre_users') if u.get('is_active')]),
            'active_ps_users': len([u for u in safe_get_data('ps_users') if u.get('is_active')])
        }

        return jsonify({
            'success': True,
            'metrics': metrics,
            'timestamp': datetime.now().isoformat()
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500










def ensure_static_directories():
    """Ensure static directories exist"""
    try:
        static_dir = os.path.join(os.path.dirname(__file__), 'static')
        images_dir = os.path.join(static_dir, 'images')

        os.makedirs(static_dir, exist_ok=True)
        os.makedirs(images_dir, exist_ok=True)

        print(f"Static directories ensured: {images_dir}")
        return images_dir
    except Exception as e:
        print(f"Error creating static directories: {e}")
        return None









# Initialize AuthManager
auth_manager = AuthManager(supabase)
# Store auth_manager in app config instead of direct attribute
app.config['AUTH_MANAGER'] = auth_manager

# Initialize rate limiter
limiter = Limiter(
    key_func=get_remote_address,
    app=app,
    default_limits=["1000 per minute"]  # Use in-memory backend for local/dev
)

# Upload folder configuration
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'csv', 'xlsx', 'xls'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max file size

# Create upload folder if it doesn't exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def send_email_to_ps(ps_email, ps_name, lead_data, cre_name):
    """Send email notification to PS when a lead is assigned"""
    try:
        if not EMAIL_USER or not EMAIL_PASSWORD:
            print("Email credentials not configured. Skipping email notification.")
            return False

        # Create message
        msg = MIMEMultipart()
        msg['From'] = EMAIL_USER
        msg['To'] = ps_email
        msg['Subject'] = f"New Lead Assigned - {lead_data['customer_name']}"

        # Email body
        body = f"""
        Dear {ps_name},

        A new lead has been assigned to you by {cre_name}.

        Lead Details:
        - Customer Name: {lead_data['customer_name']}
        - Mobile Number: {lead_data['customer_mobile_number']}
        - Source: {lead_data['source']}
        - Lead Category: {lead_data.get('lead_category', 'Not specified')}
        - Model Interested: {lead_data.get('model_interested', 'Not specified')}
        - Branch: {lead_data.get('branch', 'Not specified')}

        Please log in to the CRM system to view and update this lead.

        Best regards,
        Ather CRM System
        """

        msg.attach(MIMEText(body, 'plain'))

        # Send email
        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        server.starttls()
        server.login(EMAIL_USER, EMAIL_PASSWORD)
        text = msg.as_string()
        server.sendmail(EMAIL_USER, ps_email, text)
        server.quit()

        print(f"Email sent successfully to {ps_email}")
        return True

    except Exception as e:
        print(f"Error sending email: {e}")
        return False


def read_csv_file(filepath):
    """Read CSV file and return list of dictionaries with memory optimization"""
    data = []
    try:
        with open(filepath, 'r', encoding='utf-8') as file:
            csv_reader = csv.DictReader(file)
            for row_num, row in enumerate(csv_reader):
                if row_num >= 10000:  # Limit to 10,000 rows
                    print(f"Warning: File contains more than 10,000 rows. Only processing first 10,000.")
                    break

                # Clean and validate row data
                cleaned_row = {}
                for key, value in row.items():
                    if key and value:  # Only include non-empty keys and values
                        cleaned_row[key.strip()] = str(value).strip()

                if cleaned_row:  # Only add non-empty rows
                    data.append(cleaned_row)

                # Memory management for large files
                if row_num % 1000 == 0 and row_num > 0:
                    print(f"Processed {row_num} rows...")

    except Exception as e:
        print(f"Error reading CSV file: {e}")
        raise

    return data


def read_excel_file(filepath):
    """Read Excel file and return list of dictionaries with memory optimization"""
    data = []
    try:
        workbook = openpyxl.load_workbook(filepath, read_only=True)  # Read-only mode for memory efficiency
        sheet = workbook.active

        # Get headers from first row
        headers = []
        if sheet and sheet[1]:
            for cell in sheet[1]:
                if cell and cell.value:
                    headers.append(str(cell.value).strip())
                else:
                    headers.append(None)

        # Read data rows with limit
        row_count = 0
        if sheet:
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if row_count >= 10000:  # Limit to 10,000 rows
                    print(f"Warning: File contains more than 10,000 rows. Only processing first 10,000.")
                    break

                row_data = {}
                has_data = False

                for i, value in enumerate(row):
                    if i < len(headers) and headers[i] and value is not None:
                        row_data[headers[i]] = str(value).strip()
                        has_data = True

                if has_data:  # Only add rows with actual data
                    data.append(row_data)
                    row_count += 1

                # Memory management for large files
                if row_count % 1000 == 0 and row_count > 0:
                    print(f"Processed {row_count} rows...")

        workbook.close()  # Explicitly close workbook

    except Exception as e:
        print(f"Error reading Excel file: {e}")
        raise

    return data


def batch_insert_leads(leads_data, batch_size=100):
    """Insert leads in batches to avoid overwhelming the database"""
    total_inserted = 0
    total_batches = (len(leads_data) + batch_size - 1) // batch_size

    print(f"Starting batch insert: {len(leads_data)} leads in {total_batches} batches")

    for i in range(0, len(leads_data), batch_size):
        batch = leads_data[i:i + batch_size]
        batch_num = (i // batch_size) + 1

        try:
            # Insert batch
            result = supabase.table('lead_master').insert(batch).execute()

            if result.data:
                batch_inserted = len(result.data)
                total_inserted += batch_inserted
                print(f"Batch {batch_num}/{total_batches}: Inserted {batch_inserted} leads")
            else:
                print(f"Batch {batch_num}/{total_batches}: No data returned from insert")

            # Small delay to prevent overwhelming the database
            time.sleep(0.1)  # CHANGED from eventlet.sleep(0.1)

            # Force garbage collection every 10 batches
            if batch_num % 10 == 0:
                gc.collect()

        except Exception as e:
            print(f"Error inserting batch {batch_num}: {e}")
            # Continue with next batch instead of failing completely
            continue

    print(f"Batch insert completed: {total_inserted} total leads inserted")
    return total_inserted


def generate_uid(source, mobile_number, sequence):
    """Generate UID based on source, mobile number, and sequence"""
    source_map = {
        'Google': 'G',
        'Meta': 'M',
        'Affiliate': 'A',
        'Know': 'K',
        'Whatsapp': 'W',
        'Tele': 'T',
        'Activity': 'AC',
        'Walk-in': 'W',  # Walk-in mapping
        'Walkin': 'W'    # Walkin mapping (without hyphen)
    }

    source_char = source_map.get(source, 'X')

    # Get sequence character (A-Z)
    sequence_char = chr(65 + (sequence % 26))  # A=65 in ASCII

    # Get last 4 digits of mobile number
    mobile_str = str(mobile_number).replace(' ', '').replace('-', '')
    mobile_last4 = mobile_str[-4:] if len(mobile_str) >= 4 else mobile_str.zfill(4)

    # Generate sequence number (0001, 0002, etc.)
    seq_num = f"{(sequence % 9999) + 1:04d}"

    return f"{source_char}{sequence_char}-{mobile_last4}-{seq_num}"


def get_next_call_info(lead_data):
    """Determine the next available call number and which calls are completed"""
    call_order = ['first', 'second', 'third', 'fourth', 'fifth', 'sixth', 'seventh']
    completed_calls = []
    next_call = 'first'

    for call_num in call_order:
        call_date_key = f'{call_num}_call_date'
        if lead_data.get(call_date_key):
            completed_calls.append(call_num)
        else:
            next_call = call_num
            break

    return next_call, completed_calls


def get_next_ps_call_info(ps_data):
    """Determine the next available PS call number and which calls are completed (now 7 calls)"""
    call_order = ['first', 'second', 'third', 'fourth', 'fifth', 'sixth', 'seventh']
    completed_calls = []
    next_call = 'first'

    for call_num in call_order:
        call_date_key = f'{call_num}_call_date'
        if ps_data.get(call_date_key):
            completed_calls.append(call_num)
        else:
            next_call = call_num
            break

    return next_call, completed_calls


def get_accurate_count(table_name, filters=None):
    """Get accurate count from Supabase table"""
    try:
        query = supabase.table(table_name).select('id')

        if filters:
            for key, value in filters.items():
                if value is not None:
                    query = query.eq(key, value)

        result = query.execute()

        # Count the returned data
        return len(result.data) if result.data else 0

    except Exception as e:
        print(f"Error getting count from {table_name}: {e}")
        return 0


def safe_get_data(table_name, filters=None, select_fields='*', limit=10000):
    """Safely get data from Supabase with error handling"""
    try:
        query = supabase.table(table_name).select(select_fields)

        if filters:
            for key, value in filters.items():
                if value is not None:
                    query = query.eq(key, value)

        # Add limit to prevent default 1000 row limitation
        if limit:
            query = query.limit(limit)

        result = query.execute()
        return result.data or []
    except Exception as e:
        print(f"Error fetching data from {table_name}: {e}")
        return []


def sync_test_drive_to_alltest_drive(source_table, original_id, lead_data):
    """
    Sync test drive data to alltest_drive table when test_drive_done is updated
    """
    try:
        # Check if test_drive_done is not null and is Yes/No or True/False
        test_drive_done = lead_data.get('test_drive_done')
        
        # Handle both boolean and string values
        if test_drive_done is None:
            return
        
        # Convert boolean to string if needed
        if test_drive_done is True:
            test_drive_done = 'Yes'
        elif test_drive_done is False:
            test_drive_done = 'No'
        elif test_drive_done not in ['Yes', 'No']:
            return
        
        # Check if record already exists in alltest_drive
        existing_record = supabase.table('alltest_drive').select('*').eq('source_table', source_table).eq('original_id', str(original_id)).execute()
        
        # Prepare data for alltest_drive table
        alltest_drive_data = {
            'source_table': source_table,
            'original_id': str(original_id),
            'test_drive_done': test_drive_done,
            'updated_at': datetime.now().isoformat()
        }
        
        # Map fields based on source table
        if source_table == 'walkin_table':
            alltest_drive_data.update({
                'customer_name': lead_data.get('customer_name'),
                'mobile_number': lead_data.get('mobile_number'),
                'lead_status': lead_data.get('lead_status'),
                'lead_category': lead_data.get('lead_category'),
                'model_interested': lead_data.get('model_interested'),
                'final_status': lead_data.get('status'),
                'ps_name': lead_data.get('ps_assigned'),
                'branch': lead_data.get('branch'),
                'created_at': lead_data.get('created_at')
            })
            # For walkin_table, use uid instead of id
            alltest_drive_data['original_id'] = lead_data.get('uid', str(original_id))
        elif source_table == 'ps_followup_master':
            alltest_drive_data.update({
                'customer_name': lead_data.get('customer_name'),
                'mobile_number': lead_data.get('customer_mobile_number'),
                'lead_status': lead_data.get('lead_status'),
                'lead_category': lead_data.get('lead_category'),
                'model_interested': lead_data.get('model_interested'),
                'final_status': lead_data.get('final_status'),
                'ps_name': lead_data.get('ps_name'),
                'branch': lead_data.get('branch') or lead_data.get('ps_branch'),
                'created_at': lead_data.get('created_at'),
                'lead_source': lead_data.get('lead_source'),
                'cre_name': lead_data.get('cre_name')
            })
        elif source_table == 'activity_leads':
            alltest_drive_data.update({
                'customer_name': lead_data.get('customer_name'),
                'mobile_number': lead_data.get('customer_phone_number'),
                'lead_status': lead_data.get('lead_status'),
                'lead_category': lead_data.get('lead_category'),
                'model_interested': lead_data.get('interested_model'),
                'final_status': lead_data.get('final_status'),
                'ps_name': lead_data.get('ps_name'),
                'branch': lead_data.get('location'),
                'created_at': lead_data.get('created_at'),
                'remarks': lead_data.get('remarks'),
                'activity_name': lead_data.get('activity_name'),
                'activity_location': lead_data.get('activity_location'),
                'customer_location': lead_data.get('customer_location'),
                'customer_profession': lead_data.get('customer_profession'),
                'gender': lead_data.get('gender')
            })
        
        # Insert or update record in alltest_drive table
        if existing_record.data:
            # Update existing record
            supabase.table('alltest_drive').update(alltest_drive_data).eq('source_table', source_table).eq('original_id', str(original_id)).execute()
        else:
            # Insert new record
            supabase.table('alltest_drive').insert(alltest_drive_data).execute()
            
        print(f"Successfully synced test drive data for {source_table} - {original_id}")
        
    except Exception as e:
        print(f"Error syncing test drive data to alltest_drive: {e}")


def create_or_update_ps_followup(lead_data, ps_name, ps_branch):
    from datetime import datetime
    try:
        existing = supabase.table('ps_followup_master').select('*').eq('lead_uid', lead_data['uid']).execute()
        ps_followup_data = {
            'lead_uid': lead_data['uid'],
            'ps_name': ps_name,
            'ps_branch': ps_branch,
            'customer_name': lead_data.get('customer_name'),
            'customer_mobile_number': lead_data.get('customer_mobile_number'),
            'source': lead_data.get('source'),
            'cre_name': lead_data.get('cre_name'),
            'lead_category': lead_data.get('lead_category'),
            'model_interested': lead_data.get('model_interested'),
            'final_status': 'Pending',
            'ps_assigned_at': datetime.now().isoformat(),  # Always set when PS is assigned
            'created_at': lead_data.get('created_at') or datetime.now().isoformat(),
            'first_call_date': None  # Ensure fresh leads start without first_call_date
        }
        if existing.data:
            supabase.table('ps_followup_master').update(ps_followup_data).eq('lead_uid', lead_data['uid']).execute()
        else:
            supabase.table('ps_followup_master').insert(ps_followup_data).execute()
    except Exception as e:
        print(f"Error creating/updating PS followup: {e}")

def track_cre_call_attempt(uid, cre_name, call_no, lead_status, call_was_recorded=False, follow_up_date=None, remarks=None):
    """Track CRE call attempt in the history table and update TAT for first attempt"""
    try:
        # Get the next attempt number for this call
        attempt_result = supabase.table('cre_call_attempt_history').select('attempt').eq('uid', uid).eq('call_no', call_no).order('attempt', desc=True).limit(1).execute()
        next_attempt = 1
        if attempt_result.data:
            next_attempt = attempt_result.data[0]['attempt'] + 1

        # Fetch the current final_status from lead_master
        final_status = None
        lead_result = supabase.table('lead_master').select('final_status').eq('uid', uid).limit(1).execute()
        if lead_result.data and 'final_status' in lead_result.data[0]:
            final_status = lead_result.data[0]['final_status']

        # Prepare attempt data
        attempt_data = {
            'uid': uid,
            'call_no': call_no,
            'attempt': next_attempt,
            'status': lead_status,
            'cre_name': cre_name,
            'call_was_recorded': call_was_recorded,
            'follow_up_date': follow_up_date,
            'remarks': remarks,
            'final_status': final_status
        }

        # Insert the attempt record
        insert_result = supabase.table('cre_call_attempt_history').insert(attempt_data).execute()
        print(f"Tracked call attempt: {uid} - {call_no} call, attempt {next_attempt}, status: {lead_status}")

        # --- TAT Calculation and Update ---
        if call_no == 'first' and next_attempt == 1:
            # Fetch lead's created_at
            lead_result = supabase.table('lead_master').select('created_at').eq('uid', uid).limit(1).execute()
            if lead_result.data and lead_result.data[0].get('created_at'):
                created_at_str = lead_result.data[0]['created_at']
                from datetime import datetime
                try:
                    if 'T' in created_at_str:
                        created_at = datetime.fromisoformat(created_at_str.replace('Z', '+00:00'))
                    else:
                        created_at = datetime.strptime(created_at_str, '%Y-%m-%d %H:%M:%S')
                except Exception:
                    created_at = None
                # Get updated_at from inserted attempt (if available), else use now
                updated_at_str = None
                if insert_result.data and insert_result.data[0].get('updated_at'):
                    updated_at_str = insert_result.data[0]['updated_at']
                else:
                    from datetime import datetime
                    updated_at_str = datetime.now().isoformat()
                try:
                    if 'T' in updated_at_str:
                        updated_at = datetime.fromisoformat(updated_at_str.replace('Z', '+00:00'))
                    else:
                        updated_at = datetime.strptime(updated_at_str, '%Y-%m-%d %H:%M:%S')
                except Exception:
                    updated_at = datetime.now()
                if created_at:
                    tat_seconds = (updated_at - created_at).total_seconds()
                    # Update lead_master with TAT
                    supabase.table('lead_master').update({'tat': tat_seconds}).eq('uid', uid).execute()
                    print(f"TAT updated for lead {uid}: {tat_seconds} seconds")
    except Exception as e:
        print(f"Error tracking CRE call attempt: {e}")

def track_ps_call_attempt(uid, ps_name, call_no, lead_status, call_was_recorded=False, follow_up_date=None, remarks=None):
    """Track PS call attempt in the history table"""
    try:
        # Get the next attempt number for this call
        attempt_result = supabase.table('ps_call_attempt_history').select('attempt').eq('uid', uid).eq('call_no', call_no).order('attempt', desc=True).limit(1).execute()
        next_attempt = 1
        if attempt_result.data:
            next_attempt = attempt_result.data[0]['attempt'] + 1

        # Fetch the current final_status from multiple tables
        final_status = None
        
        # Try all possible tables and field combinations
        tables_to_check = [
            # (table_name, uid_field, status_field, uid_value)
            ('ps_followup_master', 'lead_uid', 'final_status', uid),
            ('lead_master', 'uid', 'final_status', uid),
            ('walkin_table', 'uid', 'status', uid),
            ('activity_leads', 'activity_uid', 'final_status', uid),
            # Also try with different UID formats for walkin and activity
            ('walkin_table', 'uid', 'status', uid.replace('WB-', 'W')),
            ('activity_leads', 'activity_uid', 'final_status', uid.replace('WB-', 'A'))
        ]
        
        for table_name, uid_field, status_field, uid_value in tables_to_check:
            try:
                result = supabase.table(table_name).select(status_field).eq(uid_field, uid_value).limit(1).execute()
                if result.data and result.data[0].get(status_field):
                    final_status = result.data[0][status_field]
                    print(f"Found final_status '{final_status}' in {table_name} for {uid_value}")
                    break
            except Exception as e:
                print(f"Error checking {table_name}: {e}")
                continue
        
        # If still not found, set a default
        if not final_status:
            final_status = 'Pending'
            print(f"No final_status found for {uid}, defaulting to 'Pending'")

        # Prepare attempt data
        attempt_data = {
            'uid': uid,
            'call_no': call_no,
            'attempt': next_attempt,
            'status': lead_status,
            'ps_name': ps_name,
            'call_was_recorded': call_was_recorded,
            'follow_up_date': follow_up_date,
            'remarks': remarks,
            'final_status': final_status
        }

        # Insert the attempt record
        supabase.table('ps_call_attempt_history').insert(attempt_data).execute()
        print(f"Tracked PS call attempt: {uid} - {call_no} call, attempt {next_attempt}, status: {lead_status}")
    except Exception as e:
        print(f"Error tracking PS call attempt: {e}")


def filter_leads_by_date(leads, filter_type, date_field='created_at'):
    """Filter leads based on date range"""
    if filter_type == 'all':
        return leads

    today = datetime.now().date()

    if filter_type == 'today':
        start_date = today
        end_date = today
    elif filter_type == 'mtd':  # Month to Date
        start_date = today.replace(day=1)
        end_date = today
    elif filter_type == 'week':
        start_date = today - timedelta(days=today.weekday())  # Start of current week (Monday)
        end_date = today
    elif filter_type == 'month':
        start_date = today - timedelta(days=30)
        end_date = today
    elif filter_type == 'quarter':
        start_date = today - timedelta(days=90)
        end_date = today
    elif filter_type == 'year':
        start_date = today - timedelta(days=365)
        end_date = today
    else:
        return leads

    filtered_leads = []
    for lead in leads:
        lead_date_str = lead.get(date_field)
        if lead_date_str:
            try:
                # Handle different date formats
                if 'T' in lead_date_str:  # ISO format with time
                    lead_date = datetime.fromisoformat(lead_date_str.replace('Z', '+00:00')).date()
                else:  # Date only format
                    lead_date = datetime.strptime(lead_date_str, '%Y-%m-%d').date()

                if start_date <= lead_date <= end_date:
                    filtered_leads.append(lead)
            except (ValueError, TypeError):
                # If date parsing fails, include the lead
                filtered_leads.append(lead)
        else:
            # If no date field, include the lead
            filtered_leads.append(lead)

    return filtered_leads


def fix_missing_timestamps():
    """
    Fix missing timestamps for existing leads that have final_status but missing won_timestamp or lost_timestamp
    """
    try:
        # Fix lead_master table
        # Get leads with final_status = 'Won' but no won_timestamp
        won_leads = supabase.table('lead_master').select('uid, final_status, won_timestamp, lost_timestamp, updated_at').eq('final_status', 'Won').is_('won_timestamp', 'null').execute()
        
        for lead in won_leads.data:
            supabase.table('lead_master').update({
                'won_timestamp': lead.get('updated_at') or datetime.now().isoformat()
            }).eq('uid', lead['uid']).execute()
        
        # Get leads with final_status = 'Lost' but no lost_timestamp
        lost_leads = supabase.table('lead_master').select('uid, final_status, won_timestamp, lost_timestamp, updated_at').eq('final_status', 'Lost').is_('lost_timestamp', 'null').execute()
        
        for lead in lost_leads.data:
            supabase.table('lead_master').update({
                'lost_timestamp': lead.get('updated_at') or datetime.now().isoformat()
            }).eq('uid', lead['uid']).execute()
        
        # Fix ps_followup_master table
        # Get PS leads with final_status = 'Won' but no won_timestamp
        ps_won_leads = supabase.table('ps_followup_master').select('lead_uid, final_status, won_timestamp, lost_timestamp, updated_at').eq('final_status', 'Won').is_('won_timestamp', 'null').execute()
        
        for lead in ps_won_leads.data:
            supabase.table('ps_followup_master').update({
                'won_timestamp': lead.get('updated_at') or datetime.now().isoformat()
            }).eq('lead_uid', lead['lead_uid']).execute()
        
        # Get PS leads with final_status = 'Lost' but no lost_timestamp
        ps_lost_leads = supabase.table('ps_followup_master').select('lead_uid, final_status, won_timestamp, lost_timestamp, updated_at').eq('final_status', 'Lost').is_('lost_timestamp', 'null').execute()
        
        for lead in ps_lost_leads.data:
            supabase.table('ps_followup_master').update({
                'lost_timestamp': lead.get('updated_at') or datetime.now().isoformat()
            }).eq('lead_uid', lead['lead_uid']).execute()
        
        print(f"Fixed {len(won_leads.data)} won leads, {len(lost_leads.data)} lost leads in lead_master")
        print(f"Fixed {len(ps_won_leads.data)} won leads, {len(ps_lost_leads.data)} lost leads in ps_followup_master")
        
    except Exception as e:
        print(f"Error fixing timestamps: {str(e)}")


# DUPLICATE SECTION REMOVED - This was causing route conflicts


def get_all_branches():
    # Replace with a DB fetch if you have a branches table
    return ['PORUR', 'NUNGAMBAKKAM', 'TIRUVOTTIYUR']


@app.route('/admin_duplicate_leads', methods=['GET'])
@require_admin
def admin_duplicate_leads():
    # Get filters and pagination params
    search_uid = request.args.get('search_uid', '').strip().lower()
    search_source = request.args.get('search_source', '').strip().lower()
    search_name = request.args.get('search_name', '').strip().lower()
    from_date = request.args.get('from_date', '').strip()
    to_date = request.args.get('to_date', '').strip()
    date_range_type = request.args.get('date_range_type', 'all_time')
    # Get filters and pagination params
    try:
        page = int(request.args.get('page', 1))
    except (ValueError, TypeError):
        page = 1
    page_size = 50
    offset = (page - 1) * page_size

    # Build Supabase query with filters
    query = supabase.table('duplicate_leads').select('*')
    if search_uid:
        query = query.ilike('uid', f'%{search_uid}%')
    if search_name:
        query = query.ilike('customer_name', f'%{search_name}%')
    if search_source:
        # No ilike for array fields, so filter after fetch for sources
        pass
    # Date filter will be applied after fetch (since last_enquiry_date is computed)

    # Fetch total count for pagination (without limit)
    total_result = query.execute()
    total_count = len(total_result.data or [])

    # Fetch only the current page
    query = query.range(offset, offset + page_size - 1)
    result = query.execute()
    duplicate_leads = result.data or []

    # Prepare leads for display (apply source/date filter in Python if needed)
    leads_display = []
    for lead in duplicate_leads:
        sources = []
        sub_sources = []
        dates = []
        for i in range(1, 11):
            src = lead.get(f'source{i}')
            sub = lead.get(f'sub_source{i}')
            dt = lead.get(f'date{i}')
            if src:
                sources.append(src)
            if sub and sub.strip() and sub.strip().lower() != 'unknown':
                sub_sources.append(sub)
            if dt:
                dates.append(dt)
        last_enquiry_date = max([d for d in dates if d], default=None)
        days_old = None
        if last_enquiry_date:
            try:
                last_date = datetime.strptime(last_enquiry_date, '%Y-%m-%d').date()
                days_old = (date.today() - last_date).days
            except Exception:
                days_old = None
        cre_name = None
        assigned = False
        lm_result = supabase.table('lead_master').select('cre_name').eq('uid', lead['uid']).execute()
        if lm_result.data and lm_result.data[0].get('cre_name'):
            cre_name = lm_result.data[0]['cre_name']
            assigned = True
        # Filter by Source (any source)
        if search_source and not any(search_source in (s or '').lower() for s in sources):
            continue
        # Date range filter (if select_date)
        if date_range_type == 'select_date' and from_date and to_date:
            if last_enquiry_date:
                try:
                    last_date = datetime.strptime(last_enquiry_date, '%Y-%m-%d').date()
                    from_dt = datetime.strptime(from_date, '%Y-%m-%d').date()
                    to_dt = datetime.strptime(to_date, '%Y-%m-%d').date()
                    if not (from_dt <= last_date <= to_dt):
                        continue
                except Exception as e:
                    continue
            else:
                continue
        leads_display.append({
            'uid': lead.get('uid'),
            'customer_name': lead.get('customer_name'),
            'customer_mobile_number': lead.get('customer_mobile_number'),
            'sources': sources,
            'sub_sources': sub_sources,
            'last_enquiry_date': last_enquiry_date,
            'days_old': days_old,
            'cre_name': cre_name,
            'assigned': assigned
        })
    # Calculate total pages
    total_pages = (total_count + page_size - 1) // page_size
    current_args = request.args.to_dict()
    # Previous page URL
    prev_url = None
    if page > 1:
        prev_args = current_args.copy()
        prev_args['page'] = page - 1
        prev_url = url_for('admin_duplicate_leads', **prev_args)
    # Next page URL
    next_url = None
    if page < total_pages:
        next_args = current_args.copy()
        next_args['page'] = page + 1
        next_url = url_for('admin_duplicate_leads', **next_args)
    # Page URLs for numbered links
    page_urls = []
    for p in range(1, total_pages + 1):
        page_args = current_args.copy()
        page_args['page'] = p
        page_urls.append(url_for('admin_duplicate_leads', **page_args))
    return render_template('admin_duplicate_leads.html', duplicate_leads=leads_display, page=page, total_pages=total_pages, total_count=total_count, prev_url=prev_url, next_url=next_url, page_urls=page_urls)


@app.route('/convert_duplicate_to_fresh/<uid>', methods=['POST'])
@require_admin
def convert_duplicate_to_fresh(uid):
    # Get duplicate lead
    result = supabase.table('duplicate_leads').select('*').eq('uid', uid).execute()
    if not result.data:
        flash('Duplicate lead not found', 'error')
        return redirect(url_for('admin_duplicate_leads'))
    dup_lead = result.data[0]
    # Generate new UID (use existing logic or append timestamp)
    new_uid = f"{uid}-NEW-{int(datetime.now().timestamp())}"
    # Ensure UID is at most 20 characters (DB limit)
    if len(new_uid) > 20:
        new_uid = new_uid[:20]
    # Prepare new lead data for lead_master
    lead_data = {
        'uid': new_uid,
        'customer_name': dup_lead.get('customer_name'),
        'customer_mobile_number': dup_lead.get('customer_mobile_number'),
        'source': dup_lead.get('source1'),
        'sub_source': dup_lead.get('sub_source1'),
        'date': dup_lead.get('date1'),
        'assigned': 'No',
        'lead_status': 'Pending',
        'final_status': 'Pending',
        'created_at': datetime.now().isoformat(),
        'updated_at': datetime.now().isoformat()
    }
    try:
        supabase.table('lead_master').insert(lead_data).execute()
        supabase.table('duplicate_leads').delete().eq('uid', uid).execute()
        flash('Duplicate lead converted to fresh lead successfully!', 'success')
    except Exception as e:
        flash(f'Error converting lead: {str(e)}', 'error')
    return redirect(url_for('admin_duplicate_leads'))


@app.route('/lead_transfer')
@require_admin
def lead_transfer():
    """Lead Transfer Dashboard for Admin"""
    try:
        # Get all branches
        branches = get_all_branches()
        
        return render_template('lead_transfer.html', branches=branches)
    except Exception as e:
        print(f"Error in lead_transfer: {str(e)}")
        return render_template('lead_transfer.html', branches=[])


@app.route('/api/cre_pending_leads')
@require_admin
def api_cre_pending_leads():
    """API to get pending leads summary for CRE transfer from lead_master only"""
    try:
        # Get all pending leads from lead_master grouped by CRE only
        lead_result = supabase.table('lead_master').select('cre_name').eq('final_status', 'Pending').execute()
        lead_pending_leads = lead_result.data if lead_result.data else []
        
        # Group leads by CRE and count them from lead_master only
        cre_summary = {}
        
        # Count from lead_master only
        for lead in lead_pending_leads:
            cre_name = lead.get('cre_name', 'Unassigned')
            if cre_name not in cre_summary:
                cre_summary[cre_name] = 0
            cre_summary[cre_name] += 1
        
        # Format the data for display (exclude Unassigned leads)
        formatted_summary = []
        for cre_name, count in cre_summary.items():
            # Skip unassigned leads
            if cre_name and cre_name != 'Unassigned' and cre_name != 'null':
                formatted_summary.append({
                    'cre_name': cre_name,
                    'pending_count': count
                })
        
        return jsonify({'success': True, 'data': formatted_summary})
    except Exception as e:
        print(f"Error in api_cre_pending_leads: {str(e)}")
        return jsonify({'success': False, 'message': 'Error fetching pending leads'})


@app.route('/api/ps_pending_leads')
@require_admin
def api_ps_pending_leads():
    """API to get pending leads summary for PS transfer from multiple tables"""
    try:
        # Get branch filter from request
        branch_filter = request.args.get('branch', '')
        
        # Build query for ps_followup_master
        ps_followup_query = supabase.table('ps_followup_master').select('ps_name, ps_branch').eq('final_status', 'Pending')
        if branch_filter:
            ps_followup_query = ps_followup_query.eq('ps_branch', branch_filter)
        
        ps_followup_result = ps_followup_query.execute()
        ps_followup_leads = ps_followup_result.data if ps_followup_result.data else []
        
        # Get pending leads from walkin_table where status is 'Pending' (if table exists)
        walkin_leads = []
        try:
            walkin_query = supabase.table('walkin_table').select('ps_assigned, branch').eq('status', 'Pending')
            if branch_filter:
                walkin_query = walkin_query.eq('branch', branch_filter)
            walkin_result = walkin_query.execute()
            walkin_leads = walkin_result.data if walkin_result.data else []
        except Exception as e:
            print(f"Warning: walkin_table not found or error: {str(e)}")
            walkin_leads = []
        
        # Get pending leads from activity_leads where final_status is 'Pending' (if table exists)
        activity_leads = []
        try:
            activity_query = supabase.table('activity_leads').select('ps_name, location').eq('final_status', 'Pending')
            if branch_filter:
                activity_query = activity_query.eq('location', branch_filter)
            activity_result = activity_query.execute()
            activity_leads = activity_result.data if activity_result.data else []
        except Exception as e:
            print(f"Warning: activity_leads table not found or error: {str(e)}")
            activity_leads = []
        
        # Group leads by PS and count them from all tables
        ps_summary = {}
        
        # Count from ps_followup_master
        for lead in ps_followup_leads:
            ps_name = lead.get('ps_name', 'Unassigned')
            ps_branch = lead.get('ps_branch', '')
            key = f"{ps_name}|{ps_branch}"
            if key not in ps_summary:
                ps_summary[key] = {
                    'ps_name': ps_name,
                    'ps_branch': ps_branch,
                    'ps_followup_count': 0,
                    'walkin_count': 0,
                    'activity_count': 0,
                    'total_count': 0
                }
            ps_summary[key]['ps_followup_count'] += 1
            ps_summary[key]['total_count'] += 1
        
        # Count from walkin_table
        for lead in walkin_leads:
            ps_name = lead.get('ps_assigned', 'Unassigned')  # Changed from ps_name to ps_assigned
            ps_branch = lead.get('branch', '')
            key = f"{ps_name}|{ps_branch}"
            if key not in ps_summary:
                ps_summary[key] = {
                    'ps_name': ps_name,
                    'ps_branch': ps_branch,
                    'ps_followup_count': 0,
                    'walkin_count': 0,
                    'activity_count': 0,
                    'total_count': 0
                }
            ps_summary[key]['walkin_count'] += 1
            ps_summary[key]['total_count'] += 1
        
        # Count from activity_leads
        for lead in activity_leads:
            ps_name = lead.get('ps_name', 'Unassigned')
            ps_branch = lead.get('location', '')  # location is branch in activity_leads
            key = f"{ps_name}|{ps_branch}"
            if key not in ps_summary:
                ps_summary[key] = {
                    'ps_name': ps_name,
                    'ps_branch': ps_branch,
                    'ps_followup_count': 0,
                    'walkin_count': 0,
                    'activity_count': 0,
                    'total_count': 0
                }
            ps_summary[key]['activity_count'] += 1
            ps_summary[key]['total_count'] += 1
        
        # Format the data for display (exclude Unassigned leads)
        formatted_summary = []
        for key, summary in ps_summary.items():
            # Skip unassigned leads
            if summary['ps_name'] and summary['ps_name'] != 'Unassigned' and summary['ps_name'] != 'null':
                formatted_summary.append(summary)
        
        return jsonify({'success': True, 'data': formatted_summary})
    except Exception as e:
        print(f"Error in api_ps_pending_leads: {str(e)}")
        return jsonify({'success': False, 'message': 'Error fetching pending leads'})


@app.route('/api/transfer_cre_lead', methods=['POST'])
@require_admin
def api_transfer_cre_lead():
    """API to transfer CRE lead to another CRE"""
    try:
        data = request.get_json()
        lead_uid = data.get('lead_uid')
        new_cre_name = data.get('new_cre_name')
        
        if not lead_uid or not new_cre_name:
            return jsonify({'success': False, 'message': 'Missing required parameters'})
        
        # Verify the CRE exists and is active, and get the name
        cre_result = supabase.table('cre_users').select('name, username').eq('username', new_cre_name).eq('is_active', True).execute()
        if not cre_result.data:
            return jsonify({'success': False, 'message': 'CRE not found or inactive'})
        
        # Get the target CRE's name (not username)
        target_cre_name = cre_result.data[0].get('name')
        
        # Update the lead in lead_master
        lead_result = supabase.table('lead_master').update({'cre_name': target_cre_name}).eq('uid', lead_uid).execute()
        
        # Update the lead in ps_followup_master if it exists there
        ps_followup_result = supabase.table('ps_followup_master').update({'cre_name': target_cre_name}).eq('lead_uid', lead_uid).execute()
        
        if lead_result.data or ps_followup_result.data:
            return jsonify({'success': True, 'message': 'Lead transferred successfully'})
        else:
            return jsonify({'success': False, 'message': 'Lead not found or transfer failed'})
            
    except Exception as e:
        print(f"Error in api_transfer_cre_lead: {str(e)}")
        return jsonify({'success': False, 'message': 'Error transferring lead'})


@app.route('/api/bulk_transfer_cre_leads', methods=['POST'])
@require_admin
def api_bulk_transfer_cre_leads():
    """API to bulk transfer all pending leads from one CRE to another"""
    try:
        data = request.get_json()
        from_cre_name = data.get('from_cre_name')
        to_cre_name = data.get('to_cre_name')
        
        if not from_cre_name or not to_cre_name:
            return jsonify({'success': False, 'message': 'Missing required parameters'})
        
        if from_cre_name == to_cre_name:
            return jsonify({'success': False, 'message': 'Cannot transfer to the same CRE'})
        
        # Verify the target CRE exists and is active, and get the name
        cre_result = supabase.table('cre_users').select('name, username').eq('username', to_cre_name).eq('is_active', True).execute()
        if not cre_result.data:
            return jsonify({'success': False, 'message': 'Target CRE not found or inactive'})
        
        # Get the target CRE's name (not username)
        target_cre_name = cre_result.data[0].get('name')
        
        # Update all pending leads in lead_master
        lead_result = supabase.table('lead_master').update({'cre_name': target_cre_name}).eq('cre_name', from_cre_name).eq('final_status', 'Pending').execute()
        
        # Update all pending leads in ps_followup_master
        ps_followup_result = supabase.table('ps_followup_master').update({'cre_name': target_cre_name}).eq('cre_name', from_cre_name).eq('final_status', 'Pending').execute()
        
        return jsonify({'success': True, 'message': 'Bulk transfer completed successfully'})
        
    except Exception as e:
        print(f"Error in api_bulk_transfer_cre_leads: {str(e)}")
        return jsonify({'success': False, 'message': 'Error during bulk transfer'})


@app.route('/export_leads')
@require_admin
def export_leads():
    """Export leads dashboard"""
    try:
        # Get all leads for counting
        all_leads = safe_get_data('lead_master')

        # Count by final status
        won_count = len([l for l in all_leads if l.get('final_status') == 'Won'])
        lost_count = len([l for l in all_leads if l.get('final_status') == 'Lost'])
        pending_count = len([l for l in all_leads if l.get('final_status') == 'Pending'])
        in_progress_count = len([l for l in all_leads if l.get('final_status') == 'In Progress'])

        # Count by lead category
        hot_count = len([l for l in all_leads if l.get('lead_category') == 'Hot'])
        warm_count = len([l for l in all_leads if l.get('lead_category') == 'Warm'])
        cold_count = len([l for l in all_leads if l.get('lead_category') == 'Cold'])
        not_interested_count = len([l for l in all_leads if l.get('lead_category') == 'Not Interested'])

        # Count test drive leads from alltest_drive table
        test_drive_leads = safe_get_data('alltest_drive')
        test_drive_yes_count = len([l for l in test_drive_leads if l.get('test_drive_done') == 'Yes'])
        test_drive_no_count = len([l for l in test_drive_leads if l.get('test_drive_done') == 'No'])
        test_drive_total_count = len(test_drive_leads)

        # Log export dashboard access
        auth_manager.log_audit_event(
            user_id=session.get('user_id'),
            user_type=session.get('user_type'),
            action='EXPORT_DASHBOARD_ACCESS',
            resource='export_leads'
        )

        return render_template('export_leads.html',
                               won_count=won_count,
                               lost_count=lost_count,
                               pending_count=pending_count,
                               in_progress_count=in_progress_count,
                               hot_count=hot_count,
                               warm_count=warm_count,
                               cold_count=cold_count,
                               not_interested_count=not_interested_count,
                               test_drive_yes_count=test_drive_yes_count,
                               test_drive_no_count=test_drive_no_count,
                               test_drive_total_count=test_drive_total_count)

    except Exception as e:
        flash(f'Error loading export dashboard: {str(e)}', 'error')
        return redirect(url_for('admin_dashboard'))


if __name__ == '__main__':
    print(" Starting Ather CRM System...")
    print("📱 Server will be available at: http://127.0.0.1:5000")
    print("🌐 You can also try: http://localhost:5000")
    socketio.run(app, host='0.0.0.0', port=5000, debug=True, use_reloader=False)
